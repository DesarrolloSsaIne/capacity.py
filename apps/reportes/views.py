from django.shortcuts import render, redirect

from apps.actividades.models import Ges_Actividad, Ges_Actividad_Historia
from apps.estructura.models import Ges_Niveles
from django.db.models import Q
from django.views.generic import TemplateView
from django.http import HttpResponse
import datetime
from apps.periodos.models import Glo_Periodos
from openpyxl import Workbook
from django.http import HttpResponseRedirect
from django.db.models import Subquery, OuterRef, Count, Sum
from django.db.models.functions import Extract
from django.http import HttpResponse



# Create your views here.
class GeneraReportCurvaEjecucion(TemplateView):
    template_name = 'reportes/report_seguimiento.html'


    def post(self, request, *args, **kwargs):
        template_name = 'reportes/report_seguimiento.html'
        unidad_filtro = request.POST['unidad_filtro']
        mes_filtro = request.POST['mes_filtro']
        nivel = str(unidad_filtro[0:1])
        id_nivel = str(unidad_filtro[2:5])
        request.session['nivel']=nivel
        request.session['id_nivel'] = id_nivel
        request.session['mes_filtro'] = mes_filtro

        mes_seleccionado = ''

        if mes_filtro=='1':
            mes_seleccionado='Enero'
        if mes_filtro=='2':
            mes_seleccionado='Febrero'
        if mes_filtro=='3':
            mes_seleccionado='Marzo'
        if mes_filtro=='4':
            mes_seleccionado='Abril'
        if mes_filtro=='5':
            mes_seleccionado='Mayo'
        if mes_filtro=='6':
            mes_seleccionado='Junio'
        if mes_filtro=='7':
            mes_seleccionado='Julio'
        if mes_filtro=='8':
            mes_seleccionado='Agosto'
        if mes_filtro=='9':
            mes_seleccionado='Septiembre'
        if mes_filtro=='10':
            mes_seleccionado='Octubre'
        if mes_filtro=='11':
            mes_seleccionado='Noviembre'
        if mes_filtro=='12':
            mes_seleccionado='Diciembre'

        unidad_seleccionada=''

        if nivel == '2':
            unidad_seleccionada= Ges_Actividad_Historia.objects.filter(id_controlador__id_jefatura__id_nivel__id_cuarto_nivel__tercer_nivel__segundo_nivel=id_nivel)

        if nivel == '3':
            unidad_seleccionada= Ges_Actividad_Historia.objects.filter(id_controlador__id_jefatura__id_nivel__id_tercer_nivel=id_nivel)

        if nivel == '4':
            unidad_seleccionada= Ges_Actividad_Historia.objects.filter(id_controlador__id_jefatura__id_nivel__id_cuarto_nivel=id_nivel)



        try:
            periodo_actual = Glo_Periodos.objects.get(id_estado=1)
        except Glo_Periodos.DoesNotExist:
            return None


        lista_datos_unidades = Ges_Niveles.objects.values(
            'orden_nivel',
            'id_cuarto_nivel__descripcion_nivel',
            'id_tercer_nivel__descripcion_nivel',
            'id_segundo_nivel__descripcion_nivel',
            # 'id_primer_nivel__descripcion_nivel',

            'id_cuarto_nivel_id',
            'id_tercer_nivel_id',
            'id_segundo_nivel_id',
            # 'id_primer_nivel_id',
            # # N4
            # 'id_controlador__id_jefatura__id_nivel__id_cuarto_nivel__tercer_nivel__descripcion_nivel',  # N3
            # 'id_controlador__id_jefatura__id_nivel__id_cuarto_nivel__descripcion_nivel',  # N2
            #
            # 'id_controlador__id_jefatura__id_nivel__id_cuarto_nivel__tercer_nivel__segundo_nivel__id',  # N4_id
            # 'id_controlador__id_jefatura__id_nivel__id_cuarto_nivel__tercer_nivel__id',  # N3_id
            # 'id_controlador__id_jefatura__id_nivel__id_cuarto_nivel__id',  # N2_id
        ).exclude(orden_nivel=1)



        # # # # # # # # # # # # # # # # # # # # # #  INICIO GRÁFICO CEAP  # # # # # # # # # # # # # # # # # # # # # #

        meses = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12]
        mesesacum = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11]

        ValMeses = []
        ValMesesAcum = []

        ValMesesEjec = []
        ValMesesAcumEjec = []

        estados_excluidos = [9,10,6,5]

        mydate = datetime.datetime.now()
        Mes = mydate.month

        if nivel == '2':
            for i in meses:
                val = Ges_Actividad.objects.filter(
                    Q(id_controlador__id_jefatura__id_nivel__id_cuarto_nivel__tercer_nivel__segundo_nivel=id_nivel) & Q(
                        fecha_termino_actividad__month=i) & Q(id_periodo=periodo_actual)).exclude(id_estado_actividad_id__in=estados_excluidos).count()
                ValMeses.append(val)


        if nivel == '3':
            for i in meses:
                val = Ges_Actividad.objects.filter(
                    Q(id_controlador__id_jefatura__id_nivel__id_tercer_nivel=id_nivel) & Q(
                        fecha_termino_actividad__month=i) & Q(id_periodo=periodo_actual)).exclude(id_estado_actividad_id__in=estados_excluidos).count()
                ValMeses.append(val)

        if nivel == '4':
            for i in meses:
                val = Ges_Actividad.objects.filter(
                    Q(id_controlador__id_jefatura__id_nivel__id_cuarto_nivel=id_nivel) & Q(
                        fecha_termino_actividad__month=i) & Q(id_periodo=periodo_actual)).exclude(id_estado_actividad_id__in=estados_excluidos).count()
                ValMeses.append(val)



        for i in mesesacum:

            if i == 0:
                ValMesesAcum.append(ValMeses[i])
            else:
                ValMesesAcum.append(ValMeses[i] + ValMesesAcum[i - 1])

        if nivel == '2':
            for i in meses:
                val = Ges_Actividad.objects.filter(Q(id_controlador__id_jefatura__id_nivel__id_cuarto_nivel__tercer_nivel__segundo_nivel=id_nivel) & Q(fecha_termino_actividad__month=i) & Q(id_periodo=periodo_actual) & (
                            Q(id_estado_actividad_id=8) | Q(id_estado_actividad_id=7))).exclude(id_estado_actividad_id__in=estados_excluidos).count()
                ValMesesEjec.append(val)

        if nivel == '3':
            for i in meses:
                val = Ges_Actividad.objects.filter(Q(id_controlador__id_jefatura__id_nivel__id_tercer_nivel=id_nivel) & Q(fecha_termino_actividad__month=i) & Q(id_periodo=periodo_actual) & (
                            Q(id_estado_actividad_id=8) | Q(id_estado_actividad_id=7))).exclude(id_estado_actividad_id__in=estados_excluidos).count()
                ValMesesEjec.append(val)


        if nivel == '4':
            for i in meses:
                val = Ges_Actividad.objects.filter(Q(id_controlador__id_jefatura__id_nivel__id_cuarto_nivel=id_nivel) & Q(fecha_termino_actividad__month=i) & Q(id_periodo=periodo_actual) & (
                            Q(id_estado_actividad_id=8) | Q(id_estado_actividad_id=7))).exclude(id_estado_actividad_id__in=estados_excluidos).count()
                ValMesesEjec.append(val)

        for i in range(0, Mes):
            if i == 0:
                ValMesesAcumEjec.append(ValMesesEjec[i])
            else:
                ValMesesAcumEjec.append(ValMesesEjec[i] + ValMesesAcumEjec[i - 1])

        acumPlan = ValMesesAcum[Mes - 1]
        acumFin = ValMesesAcumEjec[Mes - 1]
        ValDesviacion = "{0:.2f}".format(((acumFin / acumPlan) - 1) * 100)


        if (Mes-2)>0:
            acumPlan_ant = ValMesesAcum[Mes - 2]
            acumFin_ant = ValMesesAcumEjec[Mes - 2]

            if acumPlan_ant != 0:
                ValDesviacion_mes_anterior = "{0:.2f}".format(((acumFin_ant / acumPlan_ant) - 1) * 100)
            else:
                acumPlan_ant=1
                ValDesviacion_mes_anterior = "{0:.2f}".format(((acumFin_ant / acumPlan_ant) - 1) * 100)




        else:
            ValDesviacion_mes_anterior = 0


        # # # # # # # # # # # # # # # # # # # # # #  FIN GRÁFICO CEAP  # # # # # # # # # # # # # # # # # # # # # #


        # # # # # # # # # # # # # # # # # # # # # #  INICIO DISTRIBUCION POR CARGO  # # # # # # # # # # # # # # # # # # # #


        porcentaje_jefe_departamento = ''
        porcentaje_jefe_subdepartamento = ''
        porcentaje_analista = ''
        porcentaje_coordinadores=''
        porcentaje_supervisores=''
        porcentaje_analistas_especialistas=''
        porcentaje_supervisores_operativos=''
        porcentaje_operativos=''
        porcentaje_asistentes=''
        porcentaje_auxiliares=''
        porcentaje_total_unidad=''


        if nivel == '2':

            suma_total_unidad= list(Ges_Actividad_Historia.objects.filter(
                Q(id_controlador__id_jefatura__id_nivel__id_cuarto_nivel__tercer_nivel__segundo_nivel=id_nivel)  & Q(id_periodo=periodo_actual) &
                Q(id_periodo_seguimiento__fecha_termino_corte__month=mes_filtro)).exclude(id_estado_actividad_id__in=estados_excluidos).aggregate(
                Sum('id_actividad__total_horas')).values())[
                0]
            suma_total_unidad_finalizadas = list(Ges_Actividad_Historia.objects.filter(
                Q(id_controlador__id_jefatura__id_nivel__id_cuarto_nivel__tercer_nivel__segundo_nivel=id_nivel)  & Q(id_periodo=periodo_actual) &
                Q(id_periodo_seguimiento__fecha_termino_corte__month=mes_filtro) & Q(id_estado_actividad_id=7)).aggregate(
                Sum('id_actividad__total_horas')).values())[
                0]

            if suma_total_unidad_finalizadas == None:
                suma_total_unidad_finalizadas = 0

            if suma_total_unidad != None:
                porcentaje_total_unidad = "{0:.1f}".format((suma_total_unidad_finalizadas * 100) / suma_total_unidad)
            else:
                porcentaje_total_unidad = 0

            suma_horas_jefe_departamento = list(Ges_Actividad_Historia.objects.filter(
                Q(id_controlador__id_jefatura__id_nivel__id_cuarto_nivel__tercer_nivel__segundo_nivel=id_nivel) & Q(
                    id_actividad__id_familia_cargo_id=1) & Q(id_periodo=periodo_actual) &
                Q(id_periodo_seguimiento__fecha_termino_corte__month=mes_filtro)).exclude(id_estado_actividad_id__in=estados_excluidos).aggregate(
                Sum('id_actividad__total_horas')).values())[
                0]

            suma_horas_jefe_departamento_finalizadas = list(Ges_Actividad_Historia.objects.filter(
                Q(id_controlador__id_jefatura__id_nivel__id_cuarto_nivel__tercer_nivel__segundo_nivel=id_nivel) & Q(
                    id_actividad__id_familia_cargo_id=1) & Q(id_periodo=periodo_actual) &
                Q(id_periodo_seguimiento__fecha_termino_corte__month=mes_filtro) & Q(id_estado_actividad_id=7)).aggregate(
                Sum('id_actividad__total_horas')).values())[
                0]

            if suma_horas_jefe_departamento_finalizadas == None:
                suma_horas_jefe_departamento_finalizadas = 0

            if suma_horas_jefe_departamento != None:
                porcentaje_jefe_departamento = "{0:.1f}".format((suma_horas_jefe_departamento_finalizadas * 100) / suma_horas_jefe_departamento)
            else:
                porcentaje_jefe_departamento = 0


           ########################################################################################



            suma_horas_jefe_subdepartamento = list(Ges_Actividad_Historia.objects.filter(
                Q(id_controlador__id_jefatura__id_nivel__id_cuarto_nivel__tercer_nivel__segundo_nivel=id_nivel) & Q(
                    id_actividad__id_familia_cargo_id=2) & Q(id_periodo=periodo_actual) &
                Q(id_periodo_seguimiento__fecha_termino_corte__month=mes_filtro)).exclude(id_estado_actividad_id__in=estados_excluidos).aggregate(
                Sum('id_actividad__total_horas')).values())[
                0]

            suma_horas_jefe_subdepartamento_finalizadas = list(Ges_Actividad_Historia.objects.filter(
                Q(id_controlador__id_jefatura__id_nivel__id_cuarto_nivel__tercer_nivel__segundo_nivel=id_nivel) & Q(
                    id_actividad__id_familia_cargo_id=2) & Q(id_periodo=periodo_actual) &
                Q(id_periodo_seguimiento__fecha_termino_corte__month=mes_filtro) & Q(id_estado_actividad_id=7)).aggregate(
                Sum('id_actividad__total_horas')).values())[
                0]

            if suma_horas_jefe_subdepartamento_finalizadas == None:
                suma_horas_jefe_subdepartamento_finalizadas = 0

            if suma_horas_jefe_subdepartamento != None:
                porcentaje_jefe_subdepartamento = "{0:.1f}".format(
                    (suma_horas_jefe_subdepartamento_finalizadas * 100) / suma_horas_jefe_subdepartamento)
            else:
                porcentaje_jefe_subdepartamento = 0

         ########################################################################################

            suma_horas_coordinadores = list(Ges_Actividad_Historia.objects.filter(
                Q(id_controlador__id_jefatura__id_nivel__id_cuarto_nivel__tercer_nivel__segundo_nivel=id_nivel) & Q(
                    id_actividad__id_familia_cargo_id=3) & Q(id_periodo=periodo_actual) &
                Q(id_periodo_seguimiento__fecha_termino_corte__month=mes_filtro)).exclude(id_estado_actividad_id__in=estados_excluidos).aggregate(
                Sum('id_actividad__total_horas')).values())[
                0]

            suma_horas_coordinadores_finalizadas = list(Ges_Actividad_Historia.objects.filter(
                Q(id_controlador__id_jefatura__id_nivel__id_cuarto_nivel__tercer_nivel__segundo_nivel=id_nivel) & Q(
                    id_actividad__id_familia_cargo_id=3) & Q(id_periodo=periodo_actual) &
                Q(id_periodo_seguimiento__fecha_termino_corte__month=mes_filtro) & Q(
                    id_estado_actividad_id=7)).aggregate(
                Sum('id_actividad__total_horas')).values())[
                0]

            if suma_horas_coordinadores_finalizadas == None:
                suma_horas_coordinadores_finalizadas = 0

            if suma_horas_coordinadores != None:
                porcentaje_coordinadores = "{0:.1f}".format(
                    (suma_horas_coordinadores_finalizadas * 100) / suma_horas_coordinadores)
            else:
                porcentaje_coordinadores = 0

            ########################################################################################

            suma_horas_supervisores = list(Ges_Actividad_Historia.objects.filter(
                Q(id_controlador__id_jefatura__id_nivel__id_cuarto_nivel__tercer_nivel__segundo_nivel=id_nivel) & Q(
                    id_actividad__id_familia_cargo_id=4) & Q(id_periodo=periodo_actual) &
                Q(id_periodo_seguimiento__fecha_termino_corte__month=mes_filtro)).exclude(id_estado_actividad_id__in=estados_excluidos).aggregate(
                Sum('id_actividad__total_horas')).values())[
                0]

            suma_horas_supervisores_finalizadas = list(Ges_Actividad_Historia.objects.filter(
                Q(id_controlador__id_jefatura__id_nivel__id_cuarto_nivel__tercer_nivel__segundo_nivel=id_nivel) & Q(
                    id_actividad__id_familia_cargo_id=4) & Q(id_periodo=periodo_actual) &
                Q(id_periodo_seguimiento__fecha_termino_corte__month=mes_filtro) & Q(
                    id_estado_actividad_id=7)).aggregate(
                Sum('id_actividad__total_horas')).values())[
                0]

            if suma_horas_supervisores_finalizadas == None:
                suma_horas_supervisores_finalizadas = 0

            if suma_horas_supervisores != None:
                porcentaje_supervisores = "{0:.1f}".format(
                    (suma_horas_supervisores_finalizadas * 100) / suma_horas_supervisores)
            else:
                porcentaje_supervisores= 0

            ########################################################################################


            suma_horas_analistas_especialistas = list(Ges_Actividad_Historia.objects.filter(
                Q(id_controlador__id_jefatura__id_nivel__id_cuarto_nivel__tercer_nivel__segundo_nivel=id_nivel) & Q(
                    id_actividad__id_familia_cargo_id=5) & Q(id_periodo=periodo_actual) &
                Q(id_periodo_seguimiento__fecha_termino_corte__month=mes_filtro)).exclude(id_estado_actividad_id__in=estados_excluidos).aggregate(
                Sum('id_actividad__total_horas')).values())[
                0]

            suma_horas_analistas_especialistas_finalizadas = list(Ges_Actividad_Historia.objects.filter(
                Q(id_controlador__id_jefatura__id_nivel__id_cuarto_nivel__tercer_nivel__segundo_nivel=id_nivel) & Q(
                    id_actividad__id_familia_cargo_id=5) & Q(id_periodo=periodo_actual) &
                Q(id_periodo_seguimiento__fecha_termino_corte__month=mes_filtro) & Q(
                    id_estado_actividad_id=7)).aggregate(
                Sum('id_actividad__total_horas')).values())[
                0]

            if suma_horas_analistas_especialistas_finalizadas == None:
                suma_horas_analistas_especialistas_finalizadas = 0

            if suma_horas_analistas_especialistas != None:
                porcentaje_analistas_especialistas= "{0:.1f}".format(
                    (suma_horas_analistas_especialistas_finalizadas * 100) / suma_horas_analistas_especialistas)
            else:
                porcentaje_analistas_especialistas= 0

            ########################################################################################

            suma_horas_analistas= list(Ges_Actividad_Historia.objects.filter(Q(id_controlador__id_jefatura__id_nivel__id_cuarto_nivel__tercer_nivel__segundo_nivel=id_nivel) & Q(id_actividad__id_familia_cargo_id=6) & Q(id_periodo=periodo_actual) &
                                                                    Q(id_periodo_seguimiento__fecha_termino_corte__month=mes_filtro)).exclude(id_estado_actividad_id__in=estados_excluidos).aggregate(
                    Sum('id_actividad__total_horas')).values())[
                0]

            suma_horas_analistas_finalizadas = list(Ges_Actividad_Historia.objects.filter(
                Q(id_controlador__id_jefatura__id_nivel__id_cuarto_nivel__tercer_nivel__segundo_nivel=id_nivel) & Q(
                    id_actividad__id_familia_cargo_id=6) & Q(id_periodo=periodo_actual) &
                Q(id_periodo_seguimiento__fecha_termino_corte__month=mes_filtro) & Q(id_estado_actividad_id=7)).aggregate(
                Sum('id_actividad__total_horas')).values())[
                0]

            if suma_horas_analistas_finalizadas == None:
                suma_horas_analistas_finalizadas=0

            if suma_horas_analistas != None:
                porcentaje_analista = "{0:.1f}".format((suma_horas_analistas_finalizadas * 100) / suma_horas_analistas)
            else:
                porcentaje_analista = 0


             ################################################################

            suma_horas_supervisores_operativos= list(Ges_Actividad_Historia.objects.filter(
                Q(id_controlador__id_jefatura__id_nivel__id_cuarto_nivel__tercer_nivel__segundo_nivel=id_nivel) &
                Q(id_actividad__id_familia_cargo_id=7) & Q(id_periodo=periodo_actual) &
                 Q(id_periodo_seguimiento__fecha_termino_corte__month=mes_filtro)).exclude(id_estado_actividad_id__in=estados_excluidos).aggregate(
                    Sum('id_actividad__total_horas')).values())[
                0]

            suma_horas_supervisores_operativos_finalizadas = list(Ges_Actividad_Historia.objects.filter(
                Q(id_controlador__id_jefatura__id_nivel__id_cuarto_nivel__tercer_nivel__segundo_nivel=id_nivel) & Q(
                    id_actividad__id_familia_cargo_id=7) & Q(id_periodo=periodo_actual) &
                Q(id_periodo_seguimiento__fecha_termino_corte__month=mes_filtro) & Q(id_estado_actividad_id=7)).aggregate(
                Sum('id_actividad__total_horas')).values())[
                0]

            if suma_horas_supervisores_operativos_finalizadas == None:
                suma_horas_supervisores_operativos_finalizadas=0

            if suma_horas_supervisores_operativos != None:
                porcentaje_supervisores_operativos = "{0:.1f}".format((suma_horas_supervisores_operativos_finalizadas * 100) / suma_horas_supervisores_operativos)
            else:
                porcentaje_supervisores_operativos = 0


             ################################################################

            suma_horas_operativos= list(Ges_Actividad_Historia.objects.filter(
                Q(id_controlador__id_jefatura__id_nivel__id_cuarto_nivel__tercer_nivel__segundo_nivel=id_nivel) &
                Q(id_actividad__id_familia_cargo_id=8) & Q(id_periodo=periodo_actual) &
                 Q(id_periodo_seguimiento__fecha_termino_corte__month=mes_filtro)).exclude(id_estado_actividad_id__in=estados_excluidos).aggregate(
                    Sum('id_actividad__total_horas')).values())[
                0]

            suma_horas_operativos_finalizadas = list(Ges_Actividad_Historia.objects.filter(
                Q(id_controlador__id_jefatura__id_nivel__id_cuarto_nivel__tercer_nivel__segundo_nivel=id_nivel) & Q(
                    id_actividad__id_familia_cargo_id=8) & Q(id_periodo=periodo_actual) &
                Q(id_periodo_seguimiento__fecha_termino_corte__month=mes_filtro) & Q(id_estado_actividad_id=7)).aggregate(
                Sum('id_actividad__total_horas')).values())[
                0]

            if suma_horas_operativos_finalizadas == None:
                suma_horas_operativos_finalizadas=0

            if suma_horas_operativos != None:
                porcentaje_operativos = "{0:.1f}".format((suma_horas_operativos_finalizadas * 100) / suma_horas_operativos)
            else:
                porcentaje_operativos = 0


             ################################################################

            suma_horas_asistentes= list(Ges_Actividad_Historia.objects.filter(
                Q(id_controlador__id_jefatura__id_nivel__id_cuarto_nivel__tercer_nivel__segundo_nivel=id_nivel) &
                Q(id_actividad__id_familia_cargo_id=9) & Q(id_periodo=periodo_actual) &
                 Q(id_periodo_seguimiento__fecha_termino_corte__month=mes_filtro)).exclude(id_estado_actividad_id__in=estados_excluidos).aggregate(
                    Sum('id_actividad__total_horas')).values())[
                0]

            suma_horas_asistentes_finalizadas = list(Ges_Actividad_Historia.objects.filter(
                Q(id_controlador__id_jefatura__id_nivel__id_cuarto_nivel__tercer_nivel__segundo_nivel=id_nivel) & Q(
                    id_actividad__id_familia_cargo_id=9) & Q(id_periodo=periodo_actual) &
                Q(id_periodo_seguimiento__fecha_termino_corte__month=mes_filtro) & Q(id_estado_actividad_id=7)).aggregate(
                Sum('id_actividad__total_horas')).values())[
                0]

            if suma_horas_asistentes_finalizadas == None:
                suma_horas_asistentes_finalizadas=0

            if suma_horas_asistentes != None:
                porcentaje_asistentes = "{0:.1f}".format((suma_horas_asistentes_finalizadas * 100) / suma_horas_asistentes)
            else:
                porcentaje_asistentes = 0

                ################################################################

            suma_horas_auxiliares = list(Ges_Actividad_Historia.objects.filter(
                Q(id_controlador__id_jefatura__id_nivel__id_cuarto_nivel__tercer_nivel__segundo_nivel=id_nivel) &
                Q(id_actividad__id_familia_cargo_id=10) & Q(id_periodo=periodo_actual) &
                Q(id_periodo_seguimiento__fecha_termino_corte__month=mes_filtro)).exclude(id_estado_actividad_id__in=estados_excluidos).aggregate(
                Sum('id_actividad__total_horas')).values())[
                0]

            suma_horas_auxiliares_finalizadas = list(Ges_Actividad_Historia.objects.filter(
                Q(id_controlador__id_jefatura__id_nivel__id_cuarto_nivel__tercer_nivel__segundo_nivel=id_nivel) & Q(
                    id_actividad__id_familia_cargo_id=10) & Q(id_periodo=periodo_actual) &
                Q(id_periodo_seguimiento__fecha_termino_corte__month=mes_filtro) & Q(
                    id_estado_actividad_id=7)).aggregate(
                Sum('id_actividad__total_horas')).values())[
                0]

            if suma_horas_auxiliares_finalizadas == None:
                suma_horas_auxiliares_finalizadas = 0

            if suma_horas_auxiliares != None:
                porcentaje_auxiliares= "{0:.1f}".format(
                    (suma_horas_auxiliares_finalizadas * 100) / suma_horas_auxiliares)
            else:
                porcentaje_auxiliares = 0


             ################################################################################################################################
            ################################################################################################################################
            ################################################################################################################################

        if nivel == '3':

            suma_total_unidad = list(Ges_Actividad_Historia.objects.filter(
                Q(id_controlador__id_jefatura__id_nivel__id_tercer_nivel=id_nivel) & Q(
                    id_periodo=periodo_actual) &
                Q(id_periodo_seguimiento__fecha_termino_corte__month=mes_filtro)).exclude(id_estado_actividad_id__in=estados_excluidos).aggregate(
                Sum('id_actividad__total_horas')).values())[
                0]
            suma_total_unidad_finalizadas = list(Ges_Actividad_Historia.objects.filter(
                Q(id_controlador__id_jefatura__id_nivel__id_tercer_nivel=id_nivel) & Q(id_periodo=periodo_actual) &
                Q(id_periodo_seguimiento__fecha_termino_corte__month=mes_filtro) & Q(
                    id_estado_actividad_id=7)).aggregate(
                Sum('id_actividad__total_horas')).values())[
                0]

            if suma_total_unidad_finalizadas == None:
                suma_total_unidad_finalizadas = 0

            if suma_total_unidad != None:
                porcentaje_total_unidad = "{0:.1f}".format((suma_total_unidad_finalizadas * 100) / suma_total_unidad)
            else:
                porcentaje_total_unidad = 0



            suma_horas_jefe_departamento = list(Ges_Actividad_Historia.objects.filter(
                Q(id_controlador__id_jefatura__id_nivel__id_tercer_nivel=id_nivel) & Q(
                    id_actividad__id_familia_cargo_id=1) & Q(id_periodo=periodo_actual) &
                Q(id_periodo_seguimiento__fecha_termino_corte__month=mes_filtro)).exclude(id_estado_actividad_id__in=estados_excluidos).aggregate(
                Sum('id_actividad__total_horas')).values())[
                0]

            suma_horas_jefe_departamento_finalizadas = list(Ges_Actividad_Historia.objects.filter(
                Q(id_controlador__id_jefatura__id_nivel__id_tercer_nivel=id_nivel) & Q(
                    id_actividad__id_familia_cargo_id=1) & Q(id_periodo=periodo_actual) &
                Q(id_periodo_seguimiento__fecha_termino_corte__month=mes_filtro) & Q(id_estado_actividad_id=7)).aggregate(
                Sum('id_actividad__total_horas')).values())[
                0]

            if suma_horas_jefe_departamento_finalizadas == None:
                suma_horas_jefe_departamento_finalizadas = 0

            if suma_horas_jefe_departamento != None:
                porcentaje_jefe_departamento = "{0:.1f}".format((suma_horas_jefe_departamento_finalizadas * 100) / suma_horas_jefe_departamento)
            else:
                porcentaje_jefe_departamento = 0


           ########################################################################################



            suma_horas_jefe_subdepartamento = list(Ges_Actividad_Historia.objects.filter(
                Q(id_controlador__id_jefatura__id_nivel__id_tercer_nivel=id_nivel) & Q(
                    id_actividad__id_familia_cargo_id=2) & Q(id_periodo=periodo_actual) &
                Q(id_periodo_seguimiento__fecha_termino_corte__month=mes_filtro)).exclude(id_estado_actividad_id__in=estados_excluidos).aggregate(
                Sum('id_actividad__total_horas')).values())[
                0]

            suma_horas_jefe_subdepartamento_finalizadas = list(Ges_Actividad_Historia.objects.filter(
                Q(id_controlador__id_jefatura__id_nivel__id_tercer_nivel=id_nivel) & Q(
                    id_actividad__id_familia_cargo_id=2) & Q(id_periodo=periodo_actual) &
                Q(id_periodo_seguimiento__fecha_termino_corte__month=mes_filtro) & Q(id_estado_actividad_id=7)).aggregate(
                Sum('id_actividad__total_horas')).values())[
                0]

            if suma_horas_jefe_subdepartamento_finalizadas == None:
                suma_horas_jefe_subdepartamento_finalizadas = 0

            if suma_horas_jefe_subdepartamento != None:
                porcentaje_jefe_subdepartamento = "{0:.1f}".format(
                    (suma_horas_jefe_subdepartamento_finalizadas * 100) / suma_horas_jefe_subdepartamento)
            else:
                porcentaje_jefe_subdepartamento = 0

         ########################################################################################

            suma_horas_coordinadores = list(Ges_Actividad_Historia.objects.filter(
                Q(id_controlador__id_jefatura__id_nivel__id_tercer_nivel=id_nivel) & Q(
                    id_actividad__id_familia_cargo_id=3) & Q(id_periodo=periodo_actual) &
                Q(id_periodo_seguimiento__fecha_termino_corte__month=mes_filtro)).exclude(id_estado_actividad_id__in=estados_excluidos).aggregate(
                Sum('id_actividad__total_horas')).values())[
                0]

            suma_horas_coordinadores_finalizadas = list(Ges_Actividad_Historia.objects.filter(
                Q(id_controlador__id_jefatura__id_nivel__id_tercer_nivel=id_nivel) & Q(
                    id_actividad__id_familia_cargo_id=3) & Q(id_periodo=periodo_actual) &
                Q(id_periodo_seguimiento__fecha_termino_corte__month=mes_filtro) & Q(
                    id_estado_actividad_id=7)).aggregate(
                Sum('id_actividad__total_horas')).values())[
                0]

            if suma_horas_coordinadores_finalizadas == None:
                suma_horas_coordinadores_finalizadas = 0

            if suma_horas_coordinadores != None:
                porcentaje_coordinadores = "{0:.1f}".format(
                    (suma_horas_coordinadores_finalizadas * 100) / suma_horas_coordinadores)
            else:
                porcentaje_coordinadores = 0

            ########################################################################################

            suma_horas_supervisores = list(Ges_Actividad_Historia.objects.filter(
                Q(id_controlador__id_jefatura__id_nivel__id_tercer_nivel=id_nivel) & Q(
                    id_actividad__id_familia_cargo_id=4) & Q(id_periodo=periodo_actual) &
                Q(id_periodo_seguimiento__fecha_termino_corte__month=mes_filtro)).exclude(id_estado_actividad_id__in=estados_excluidos).aggregate(
                Sum('id_actividad__total_horas')).values())[
                0]

            suma_horas_supervisores_finalizadas = list(Ges_Actividad_Historia.objects.filter(
                Q(id_controlador__id_jefatura__id_nivel__id_tercer_nivel=id_nivel) & Q(
                    id_actividad__id_familia_cargo_id=4) & Q(id_periodo=periodo_actual) &
                Q(id_periodo_seguimiento__fecha_termino_corte__month=mes_filtro) & Q(
                    id_estado_actividad_id=7)).aggregate(
                Sum('id_actividad__total_horas')).values())[
                0]

            if suma_horas_supervisores_finalizadas == None:
                suma_horas_supervisores_finalizadas = 0

            if suma_horas_supervisores != None:
                porcentaje_supervisores = "{0:.1f}".format(
                    (suma_horas_supervisores_finalizadas * 100) / suma_horas_supervisores)
            else:
                porcentaje_supervisores= 0

            ########################################################################################


            suma_horas_analistas_especialistas = list(Ges_Actividad_Historia.objects.filter(
                Q(id_controlador__id_jefatura__id_nivel__id_tercer_nivel=id_nivel) & Q(
                    id_actividad__id_familia_cargo_id=5) & Q(id_periodo=periodo_actual) &
                Q(id_periodo_seguimiento__fecha_termino_corte__month=mes_filtro)).exclude(id_estado_actividad_id__in=estados_excluidos).aggregate(
                Sum('id_actividad__total_horas')).values())[
                0]

            suma_horas_analistas_especialistas_finalizadas = list(Ges_Actividad_Historia.objects.filter(
                Q(id_controlador__id_jefatura__id_nivel__id_tercer_nivel=id_nivel) & Q(
                    id_actividad__id_familia_cargo_id=5) & Q(id_periodo=periodo_actual) &
                Q(id_periodo_seguimiento__fecha_termino_corte__month=mes_filtro) & Q(
                    id_estado_actividad_id=7)).aggregate(
                Sum('id_actividad__total_horas')).values())[
                0]

            if suma_horas_analistas_especialistas_finalizadas == None:
                suma_horas_analistas_especialistas_finalizadas = 0

            if suma_horas_analistas_especialistas != None:
                porcentaje_analistas_especialistas= "{0:.1f}".format(
                    (suma_horas_analistas_especialistas_finalizadas * 100) / suma_horas_analistas_especialistas)
            else:
                porcentaje_analistas_especialistas= 0

            ########################################################################################

            suma_horas_analistas= list(Ges_Actividad_Historia.objects.filter(Q(id_controlador__id_jefatura__id_nivel__id_tercer_nivel=id_nivel) & Q(id_actividad__id_familia_cargo_id=6) & Q(id_periodo=periodo_actual) &
                                                                    Q(id_periodo_seguimiento__fecha_termino_corte__month=mes_filtro)).exclude(id_estado_actividad_id__in=estados_excluidos).exclude(id_estado_actividad_id__in=estados_excluidos).aggregate(
                    Sum('id_actividad__total_horas')).values())[
                0]

            suma_horas_analistas_finalizadas = list(Ges_Actividad_Historia.objects.filter(
                Q(id_controlador__id_jefatura__id_nivel__id_tercer_nivel=id_nivel) & Q(
                    id_actividad__id_familia_cargo_id=6) & Q(id_periodo=periodo_actual) &
                Q(id_periodo_seguimiento__fecha_termino_corte__month=mes_filtro) & Q(id_estado_actividad_id=7)).aggregate(
                Sum('id_actividad__total_horas')).values())[
                0]

            if suma_horas_analistas_finalizadas == None:
                suma_horas_analistas_finalizadas=0

            if suma_horas_analistas != None:
                porcentaje_analista = "{0:.1f}".format((suma_horas_analistas_finalizadas * 100) / suma_horas_analistas)
            else:
                porcentaje_analista = 0


             ################################################################

            suma_horas_supervisores_operativos= list(Ges_Actividad_Historia.objects.filter(
                Q(id_controlador__id_jefatura__id_nivel__id_tercer_nivel=id_nivel) &
                Q(id_actividad__id_familia_cargo_id=7) & Q(id_periodo=periodo_actual) &
                 Q(id_periodo_seguimiento__fecha_termino_corte__month=mes_filtro)).exclude(id_estado_actividad_id__in=estados_excluidos).aggregate(
                    Sum('id_actividad__total_horas')).values())[
                0]

            suma_horas_supervisores_operativos_finalizadas = list(Ges_Actividad_Historia.objects.filter(
                Q(id_controlador__id_jefatura__id_nivel__id_tercer_nivel=id_nivel) & Q(
                    id_actividad__id_familia_cargo_id=7) & Q(id_periodo=periodo_actual) &
                Q(id_periodo_seguimiento__fecha_termino_corte__month=mes_filtro) & Q(id_estado_actividad_id=7)).aggregate(
                Sum('id_actividad__total_horas')).values())[
                0]

            if suma_horas_supervisores_operativos_finalizadas == None:
                suma_horas_supervisores_operativos_finalizadas=0

            if suma_horas_supervisores_operativos != None:
                porcentaje_supervisores_operativos = "{0:.1f}".format((suma_horas_supervisores_operativos_finalizadas * 100) / suma_horas_supervisores_operativos)
            else:
                porcentaje_supervisores_operativos = 0


             ################################################################

            suma_horas_operativos= list(Ges_Actividad_Historia.objects.filter(
                Q(id_controlador__id_jefatura__id_nivel__id_tercer_nivel=id_nivel) &
                Q(id_actividad__id_familia_cargo_id=8) & Q(id_periodo=periodo_actual) &
                 Q(id_periodo_seguimiento__fecha_termino_corte__month=mes_filtro)).exclude(id_estado_actividad_id__in=estados_excluidos).aggregate(
                    Sum('id_actividad__total_horas')).values())[
                0]

            suma_horas_operativos_finalizadas = list(Ges_Actividad_Historia.objects.filter(
                Q(id_controlador__id_jefatura__id_nivel__id_tercer_nivel=id_nivel) & Q(
                    id_actividad__id_familia_cargo_id=8) & Q(id_periodo=periodo_actual) &
                Q(id_periodo_seguimiento__fecha_termino_corte__month=mes_filtro) & Q(id_estado_actividad_id=7)).aggregate(
                Sum('id_actividad__total_horas')).values())[
                0]

            if suma_horas_operativos_finalizadas == None:
                suma_horas_operativos_finalizadas=0

            if suma_horas_operativos != None:
                porcentaje_operativos = "{0:.1f}".format((suma_horas_operativos_finalizadas * 100) / suma_horas_operativos)
            else:
                porcentaje_operativos = 0


             ################################################################

            suma_horas_asistentes= list(Ges_Actividad_Historia.objects.filter(
                Q(id_controlador__id_jefatura__id_nivel__id_tercer_nivel=id_nivel) &
                Q(id_actividad__id_familia_cargo_id=9) & Q(id_periodo=periodo_actual) &
                 Q(id_periodo_seguimiento__fecha_termino_corte__month=mes_filtro)).exclude(id_estado_actividad_id__in=estados_excluidos).aggregate(
                    Sum('id_actividad__total_horas')).values())[
                0]

            suma_horas_asistentes_finalizadas = list(Ges_Actividad_Historia.objects.filter(
                Q(id_controlador__id_jefatura__id_nivel__id_tercer_nivel=id_nivel) & Q(
                    id_actividad__id_familia_cargo_id=9) & Q(id_periodo=periodo_actual) &
                Q(id_periodo_seguimiento__fecha_termino_corte__month=mes_filtro) & Q(id_estado_actividad_id=7)).aggregate(
                Sum('id_actividad__total_horas')).values())[
                0]

            if suma_horas_asistentes_finalizadas == None:
                suma_horas_asistentes_finalizadas=0

            if suma_horas_asistentes != None:
                porcentaje_asistentes = "{0:.1f}".format((suma_horas_asistentes_finalizadas * 100) / suma_horas_asistentes)
            else:
                porcentaje_asistentes = 0

                ################################################################

            suma_horas_auxiliares = list(Ges_Actividad_Historia.objects.filter(
                Q(id_controlador__id_jefatura__id_nivel__id_tercer_nivel=id_nivel) &
                Q(id_actividad__id_familia_cargo_id=10) & Q(id_periodo=periodo_actual) &
                Q(id_periodo_seguimiento__fecha_termino_corte__month=mes_filtro)).exclude(id_estado_actividad_id__in=estados_excluidos).aggregate(
                Sum('id_actividad__total_horas')).values())[
                0]

            suma_horas_auxiliares_finalizadas = list(Ges_Actividad_Historia.objects.filter(
                Q(id_controlador__id_jefatura__id_nivel__id_tercer_nivel=id_nivel) & Q(
                    id_actividad__id_familia_cargo_id=10) & Q(id_periodo=periodo_actual) &
                Q(id_periodo_seguimiento__fecha_termino_corte__month=mes_filtro) & Q(
                    id_estado_actividad_id=7)).aggregate(
                Sum('id_actividad__total_horas')).values())[
                0]

            if suma_horas_auxiliares_finalizadas == None:
                suma_horas_auxiliares_finalizadas = 0

            if suma_horas_auxiliares != None:
                porcentaje_auxiliares= "{0:.1f}".format(
                    (suma_horas_auxiliares_finalizadas * 100) / suma_horas_auxiliares)
            else:
                porcentaje_auxiliares = 0


             ################################################################################################################################
            ################################################################################################################################
            ################################################################################################################################

        if nivel == '4':

            suma_total_unidad = list(Ges_Actividad_Historia.objects.filter(
                Q(id_controlador__id_jefatura__id_nivel__id_cuarto_nivel=id_nivel) & Q(
                    id_periodo=periodo_actual) &
                Q(id_periodo_seguimiento__fecha_termino_corte__month=mes_filtro)).exclude(id_estado_actividad_id__in=estados_excluidos).aggregate(
                Sum('id_actividad__total_horas')).values())[
                0]
            suma_total_unidad_finalizadas = list(Ges_Actividad_Historia.objects.filter(
                Q(id_controlador__id_jefatura__id_nivel__id_cuarto_nivel=id_nivel) & Q(
                    id_periodo=periodo_actual) &
                Q(id_periodo_seguimiento__fecha_termino_corte__month=mes_filtro) & Q(
                    id_estado_actividad_id=7)).aggregate(
                Sum('id_actividad__total_horas')).values())[
                0]

            if suma_total_unidad_finalizadas == None:
                suma_total_unidad_finalizadas = 0

            if suma_total_unidad != None:
                porcentaje_total_unidad = "{0:.1f}".format((suma_total_unidad_finalizadas * 100) / suma_total_unidad)
            else:
                porcentaje_total_unidad = 0

            suma_horas_jefe_departamento = list(Ges_Actividad_Historia.objects.filter(
                Q(id_controlador__id_jefatura__id_nivel__id_cuarto_nivel=id_nivel) & Q(
                    id_actividad__id_familia_cargo_id=1) & Q(id_periodo=periodo_actual) &
                Q(id_periodo_seguimiento__fecha_termino_corte__month=mes_filtro)).exclude(id_estado_actividad_id__in=estados_excluidos).aggregate(
                Sum('id_actividad__total_horas')).values())[
                0]

            suma_horas_jefe_departamento_finalizadas = list(Ges_Actividad_Historia.objects.filter(
                Q(id_controlador__id_jefatura__id_nivel__id_cuarto_nivel=id_nivel) & Q(
                    id_actividad__id_familia_cargo_id=1) & Q(id_periodo=periodo_actual) &
                Q(id_periodo_seguimiento__fecha_termino_corte__month=mes_filtro) & Q(id_estado_actividad_id=7)).aggregate(
                Sum('id_actividad__total_horas')).values())[
                0]

            if suma_horas_jefe_departamento_finalizadas == None:
                suma_horas_jefe_departamento_finalizadas = 0

            if suma_horas_jefe_departamento != None:
                porcentaje_jefe_departamento = "{0:.1f}".format((suma_horas_jefe_departamento_finalizadas * 100) / suma_horas_jefe_departamento)
            else:
                porcentaje_jefe_departamento = 0


           ########################################################################################

            suma_horas_jefe_subdepartamento = list(Ges_Actividad_Historia.objects.filter(
                Q(id_controlador__id_jefatura__id_nivel__id_cuarto_nivel=id_nivel) & Q(
                    id_actividad__id_familia_cargo_id=2) & Q(id_periodo=periodo_actual) &
                Q(id_periodo_seguimiento__fecha_termino_corte__month=mes_filtro)).exclude(id_estado_actividad_id__in=estados_excluidos).aggregate(
                Sum('id_actividad__total_horas')).values())[
                0]

            suma_horas_jefe_subdepartamento_finalizadas = list(Ges_Actividad_Historia.objects.filter(
                Q(id_controlador__id_jefatura__id_nivel__id_cuarto_nivel=id_nivel) & Q(
                    id_actividad__id_familia_cargo_id=2) & Q(id_periodo=periodo_actual) &
                Q(id_periodo_seguimiento__fecha_termino_corte__month=mes_filtro) & Q(id_estado_actividad_id=7)).aggregate(
                Sum('id_actividad__total_horas')).values())[
                0]

            if suma_horas_jefe_subdepartamento_finalizadas == None:
                suma_horas_jefe_subdepartamento_finalizadas = 0

            if suma_horas_jefe_subdepartamento != None:
                porcentaje_jefe_subdepartamento = "{0:.1f}".format(
                    (suma_horas_jefe_subdepartamento_finalizadas * 100) / suma_horas_jefe_subdepartamento)
            else:
                porcentaje_jefe_subdepartamento = 0

         ########################################################################################

            suma_horas_coordinadores = list(Ges_Actividad_Historia.objects.filter(
                Q(id_controlador__id_jefatura__id_nivel__id_cuarto_nivel=id_nivel) & Q(
                    id_actividad__id_familia_cargo_id=3) & Q(id_periodo=periodo_actual) &
                Q(id_periodo_seguimiento__fecha_termino_corte__month=mes_filtro)).exclude(id_estado_actividad_id__in=estados_excluidos).aggregate(
                Sum('id_actividad__total_horas')).values())[
                0]

            suma_horas_coordinadores_finalizadas = list(Ges_Actividad_Historia.objects.filter(
                Q(id_controlador__id_jefatura__id_nivel__id_cuarto_nivel=id_nivel) & Q(
                    id_actividad__id_familia_cargo_id=3) & Q(id_periodo=periodo_actual) &
                Q(id_periodo_seguimiento__fecha_termino_corte__month=mes_filtro) & Q(
                    id_estado_actividad_id=7)).aggregate(
                Sum('id_actividad__total_horas')).values())[
                0]

            if suma_horas_coordinadores_finalizadas == None:
                suma_horas_coordinadores_finalizadas = 0

            if suma_horas_coordinadores != None:
                porcentaje_coordinadores = "{0:.1f}".format(
                    (suma_horas_coordinadores_finalizadas * 100) / suma_horas_coordinadores)
            else:
                porcentaje_coordinadores = 0

            ########################################################################################

            suma_horas_supervisores = list(Ges_Actividad_Historia.objects.filter(
                Q(id_controlador__id_jefatura__id_nivel__id_cuarto_nivel=id_nivel) & Q(
                    id_actividad__id_familia_cargo_id=4) & Q(id_periodo=periodo_actual) &
                Q(id_periodo_seguimiento__fecha_termino_corte__month=mes_filtro)).exclude(id_estado_actividad_id__in=estados_excluidos).aggregate(
                Sum('id_actividad__total_horas')).values())[
                0]

            suma_horas_supervisores_finalizadas = list(Ges_Actividad_Historia.objects.filter(
                Q(id_controlador__id_jefatura__id_nivel__id_cuarto_nivel=id_nivel) & Q(
                    id_actividad__id_familia_cargo_id=4) & Q(id_periodo=periodo_actual) &
                Q(id_periodo_seguimiento__fecha_termino_corte__month=mes_filtro) & Q(
                    id_estado_actividad_id=7)).aggregate(
                Sum('id_actividad__total_horas')).values())[
                0]

            if suma_horas_supervisores_finalizadas == None:
                suma_horas_supervisores_finalizadas = 0

            if suma_horas_supervisores != None:
                porcentaje_supervisores = "{0:.1f}".format(
                    (suma_horas_supervisores_finalizadas * 100) / suma_horas_supervisores)
            else:
                porcentaje_supervisores= 0

            ########################################################################################


            suma_horas_analistas_especialistas = list(Ges_Actividad_Historia.objects.filter(
                Q(id_controlador__id_jefatura__id_nivel__id_cuarto_nivel=id_nivel) & Q(
                    id_actividad__id_familia_cargo_id=5) & Q(id_periodo=periodo_actual) &
                Q(id_periodo_seguimiento__fecha_termino_corte__month=mes_filtro)).exclude(id_estado_actividad_id__in=estados_excluidos).aggregate(
                Sum('id_actividad__total_horas')).values())[
                0]

            suma_horas_analistas_especialistas_finalizadas = list(Ges_Actividad_Historia.objects.filter(
                Q(id_controlador__id_jefatura__id_nivel__id_cuarto_nivel=id_nivel) & Q(
                    id_actividad__id_familia_cargo_id=5) & Q(id_periodo=periodo_actual) &
                Q(id_periodo_seguimiento__fecha_termino_corte__month=mes_filtro) & Q(
                    id_estado_actividad_id=7)).aggregate(
                Sum('id_actividad__total_horas')).values())[
                0]

            if suma_horas_analistas_especialistas_finalizadas == None:
                suma_horas_analistas_especialistas_finalizadas = 0

            if suma_horas_analistas_especialistas != None:
                porcentaje_analistas_especialistas= "{0:.1f}".format(
                    (suma_horas_analistas_especialistas_finalizadas * 100) / suma_horas_analistas_especialistas)
            else:
                porcentaje_analistas_especialistas= 0

            ########################################################################################

            suma_horas_analistas= list(Ges_Actividad_Historia.objects.filter(Q(id_controlador__id_jefatura__id_nivel__id_cuarto_nivel=id_nivel) & Q(id_actividad__id_familia_cargo_id=6) & Q(id_periodo=periodo_actual) &
                                                                    Q(id_periodo_seguimiento__fecha_termino_corte__month=mes_filtro)).exclude(id_estado_actividad_id__in=estados_excluidos).aggregate(
                    Sum('id_actividad__total_horas')).values())[
                0]

            suma_horas_analistas_finalizadas = list(Ges_Actividad_Historia.objects.filter(
                Q(id_controlador__id_jefatura__id_nivel__id_cuarto_nivel=id_nivel) & Q(
                    id_actividad__id_familia_cargo_id=6) & Q(id_periodo=periodo_actual) &
                Q(id_periodo_seguimiento__fecha_termino_corte__month=mes_filtro) & Q(id_estado_actividad_id=7)).aggregate(
                Sum('id_actividad__total_horas')).values())[
                0]

            if suma_horas_analistas_finalizadas == None:
                suma_horas_analistas_finalizadas=0

            if suma_horas_analistas != None:
                porcentaje_analista = "{0:.1f}".format((suma_horas_analistas_finalizadas * 100) / suma_horas_analistas)
            else:
                porcentaje_analista = 0


             ################################################################

            suma_horas_supervisores_operativos= list(Ges_Actividad_Historia.objects.filter(
                Q(id_controlador__id_jefatura__id_nivel__id_cuarto_nivel=id_nivel) &
                Q(id_actividad__id_familia_cargo_id=7) & Q(id_periodo=periodo_actual) &
                 Q(id_periodo_seguimiento__fecha_termino_corte__month=mes_filtro)).exclude(id_estado_actividad_id__in=estados_excluidos).aggregate(
                    Sum('id_actividad__total_horas')).values())[
                0]

            suma_horas_supervisores_operativos_finalizadas = list(Ges_Actividad_Historia.objects.filter(
                Q(id_controlador__id_jefatura__id_nivel__id_cuarto_nivel=id_nivel) & Q(
                    id_actividad__id_familia_cargo_id=7) & Q(id_periodo=periodo_actual) &
                Q(id_periodo_seguimiento__fecha_termino_corte__month=mes_filtro) & Q(id_estado_actividad_id=7)).aggregate(
                Sum('id_actividad__total_horas')).values())[
                0]

            if suma_horas_supervisores_operativos_finalizadas == None:
                suma_horas_supervisores_operativos_finalizadas=0

            if suma_horas_supervisores_operativos != None:
                porcentaje_supervisores_operativos = "{0:.1f}".format((suma_horas_supervisores_operativos_finalizadas * 100) / suma_horas_supervisores_operativos)
            else:
                porcentaje_supervisores_operativos = 0


             ################################################################

            suma_horas_operativos= list(Ges_Actividad_Historia.objects.filter(
                Q(id_controlador__id_jefatura__id_nivel__id_cuarto_nivel=id_nivel) &
                Q(id_actividad__id_familia_cargo_id=8) & Q(id_periodo=periodo_actual) &
                 Q(id_periodo_seguimiento__fecha_termino_corte__month=mes_filtro)).exclude(id_estado_actividad_id__in=estados_excluidos).aggregate(
                    Sum('id_actividad__total_horas')).values())[
                0]

            suma_horas_operativos_finalizadas = list(Ges_Actividad_Historia.objects.filter(
                Q(id_controlador__id_jefatura__id_nivel__id_cuarto_nivel=id_nivel) & Q(
                    id_actividad__id_familia_cargo_id=8) & Q(id_periodo=periodo_actual) &
                Q(id_periodo_seguimiento__fecha_termino_corte__month=mes_filtro) & Q(id_estado_actividad_id=7)).aggregate(
                Sum('id_actividad__total_horas')).values())[
                0]

            if suma_horas_operativos_finalizadas == None:
                suma_horas_operativos_finalizadas=0

            if suma_horas_operativos != None:
                porcentaje_operativos = "{0:.1f}".format((suma_horas_operativos_finalizadas * 100) / suma_horas_operativos)
            else:
                porcentaje_operativos = 0


             ################################################################

            suma_horas_asistentes= list(Ges_Actividad_Historia.objects.filter(
                Q(id_controlador__id_jefatura__id_nivel__id_cuarto_nivel=id_nivel) &
                Q(id_actividad__id_familia_cargo_id=9) & Q(id_periodo=periodo_actual) &
                 Q(id_periodo_seguimiento__fecha_termino_corte__month=mes_filtro)).exclude(id_estado_actividad_id__in=estados_excluidos).aggregate(
                    Sum('id_actividad__total_horas')).values())[
                0]

            suma_horas_asistentes_finalizadas = list(Ges_Actividad_Historia.objects.filter(
                Q(id_controlador__id_jefatura__id_nivel__id_cuarto_nivel=id_nivel) & Q(
                    id_actividad__id_familia_cargo_id=9) & Q(id_periodo=periodo_actual) &
                Q(id_periodo_seguimiento__fecha_termino_corte__month=mes_filtro) & Q(id_estado_actividad_id=7)).aggregate(
                Sum('id_actividad__total_horas')).values())[
                0]

            if suma_horas_asistentes_finalizadas == None:
                suma_horas_asistentes_finalizadas=0

            if suma_horas_asistentes != None:
                porcentaje_asistentes = "{0:.1f}".format((suma_horas_asistentes_finalizadas * 100) / suma_horas_asistentes)
            else:
                porcentaje_asistentes = 0

                ################################################################

            suma_horas_auxiliares = list(Ges_Actividad_Historia.objects.filter(
                Q(id_controlador__id_jefatura__id_nivel__id_cuarto_nivel=id_nivel) &
                Q(id_actividad__id_familia_cargo_id=10) & Q(id_periodo=periodo_actual) &
                Q(id_periodo_seguimiento__fecha_termino_corte__month=mes_filtro)).exclude(id_estado_actividad_id__in=estados_excluidos).aggregate(
                Sum('id_actividad__total_horas')).values())[
                0]

            suma_horas_auxiliares_finalizadas = list(Ges_Actividad_Historia.objects.filter(
                Q(id_controlador__id_jefatura__id_nivel__id_cuarto_nivel=id_nivel) & Q(
                    id_actividad__id_familia_cargo_id=10) & Q(id_periodo=periodo_actual) &
                Q(id_periodo_seguimiento__fecha_termino_corte__month=mes_filtro) & Q(
                    id_estado_actividad_id=7)).aggregate(
                Sum('id_actividad__total_horas')).values())[
                0]

            if suma_horas_auxiliares_finalizadas == None:
                suma_horas_auxiliares_finalizadas = 0

            if suma_horas_auxiliares != None:
                porcentaje_auxiliares= "{0:.1f}".format(
                    (suma_horas_auxiliares_finalizadas * 100) / suma_horas_auxiliares)
            else:
                porcentaje_auxiliares = 0


             ################################################################


        # # # # # # # # # # # # # # # # # # # # # #  FIN DISTRIBUCION POR CARGO  # # # # # # # # # # # # # # # # # # # # # #

        # # # # # # # # # # # # # # # # # # # # # #  CONTEXT COMPARTIDO # # # # # # # # # # # # # # # # # # # # # #
        context = {"ValMesesAcum": ValMesesAcum,
                   "ValMesesAcumEjec": ValMesesAcumEjec,
                   "ValDesviacion": ValDesviacion,
                   'ValDesviacion_mes_anterior':ValDesviacion_mes_anterior,
                   "lista_datos_unidades": lista_datos_unidades,
                   'unidad':unidad_seleccionada,
                   'nivel':nivel, 'mes_seleccionado':mes_seleccionado,
                   'porcentaje_analista': porcentaje_analista,
                   'porcentaje_jefe_departamento':porcentaje_jefe_departamento,
                   'porcentaje_jefe_subdepartamento':porcentaje_jefe_subdepartamento,
                   'porcentaje_coordinadores':porcentaje_coordinadores,
                   'porcentaje_supervisores':porcentaje_supervisores,
                   'porcentaje_analistas_especialistas':porcentaje_analistas_especialistas,
                   'porcentaje_supervisores_operativos':porcentaje_supervisores_operativos,
                   'porcentaje_operativos':porcentaje_operativos,
                   'porcentaje_asistentes':porcentaje_asistentes,
                   'porcentaje_auxiliares':porcentaje_auxiliares,
                   'porcentaje_total_unidad':porcentaje_total_unidad,


                   'unidad_filtro':unidad_filtro, 'mes_filtro':mes_filtro,
                   }
        # # # # # # # # # # # # # # # # # # # # # # FIN  CONTEXT COMPARTIDO # # # # # # # # # # # # # # # # # # # # # #


        #########################################DATOS####################################

        # lista_datos = Ges_Actividad.objects.filter(id_periodo=periodo_actual)
        if nivel == '4':
            lista_datos = Ges_Actividad_Historia.objects.filter(Q(id_controlador__id_jefatura__id_nivel__id_cuarto_nivel=id_nivel) &
                                                                Q(id_periodo=periodo_actual) &
                                                                Q(id_periodo_seguimiento__fecha_termino_corte__month=mes_filtro))
            context['object_list'] = lista_datos


        if nivel == '3':
            lista_datos = Ges_Actividad_Historia.objects.filter(Q(id_controlador__id_jefatura__id_nivel__id_tercer_nivel=id_nivel) &
                                                                Q(id_periodo=periodo_actual) &
                                                                Q(id_periodo_seguimiento__fecha_termino_corte__month=mes_filtro))
            context['object_list'] = lista_datos

        if nivel == '2':
            lista_datos = Ges_Actividad_Historia.objects.filter(Q(id_controlador__id_jefatura__id_nivel__id_cuarto_nivel__tercer_nivel__segundo_nivel=id_nivel) &
                                                                Q(id_periodo=periodo_actual) &
                                                                Q(id_periodo_seguimiento__fecha_termino_corte__month=mes_filtro))
            context['object_list'] = lista_datos


        ######################################### FIN DATOS####################################

        # # # # # # # # # # # # # # # # # # # # # #  INICIO ESTADO CUMPLIMIENTO  # # # # # # # # # # # # # # # # # # # # # #

        lista_datos_count=''

        if nivel == '2':
            lista_datos_count = Ges_Actividad_Historia.objects.annotate(
                month=Extract('id_periodo_seguimiento__fecha_termino_corte', 'month')).filter(
                Q(id_periodo_seguimiento__fecha_termino_corte__month=mes_filtro) & Q(id_controlador__id_jefatura__id_nivel__id_cuarto_nivel__tercer_nivel__segundo_nivel=id_nivel) & Q(
                    id_periodo=periodo_actual)).values(
                'month',
                'id_controlador__nivel_inicial',

                'id_controlador__id_jefatura__id_nivel__descripcion_nivel').annotate(
                CountNoIniciadas=Count('id', filter=Q(id_estado_actividad=4)),
                CountFinalizadas=Count('id', filter=Q(id_estado_actividad=7)),
                CountConRetraso=Count('id', filter=Q(id_estado_actividad=1)),
                CountSinMovimiento=Count('id', filter=Q(id_estado_actividad=6)),
                CountEliminadas=Count('id', filter=Q(id_estado_actividad=9)),
                CountEnCurso=Count('id', filter=Q(id_estado_actividad=8)),
                CountTotal=Count('id', filter=(
                            ~Q(id_estado_actividad=2) & ~Q(id_estado_actividad=3) & ~Q(id_estado_actividad=5) & ~Q(
                        id_estado_actividad=10)))

            )

        if nivel == '3':
            lista_datos_count = Ges_Actividad_Historia.objects.annotate(
                month=Extract('id_periodo_seguimiento__fecha_termino_corte', 'month')).filter(
                Q(id_periodo_seguimiento__fecha_termino_corte__month=mes_filtro) &
                Q(id_controlador__id_jefatura__id_nivel__id_tercer_nivel=id_nivel) & Q(
                    id_periodo=periodo_actual)).values(
                'month',
                'id_controlador__nivel_inicial',

                'id_controlador__id_jefatura__id_nivel__descripcion_nivel').annotate(
                CountNoIniciadas=Count('id', filter=Q(id_estado_actividad=4)),
                CountFinalizadas=Count('id', filter=Q(id_estado_actividad=7)),
                CountConRetraso=Count('id', filter=Q(id_estado_actividad=1)),
                CountSinMovimiento=Count('id', filter=Q(id_estado_actividad=6)),
                CountEliminadas=Count('id', filter=Q(id_estado_actividad=9)),
                CountEnCurso=Count('id', filter=Q(id_estado_actividad=8)),
                CountTotal=Count('id', filter=(
                            ~Q(id_estado_actividad=2) & ~Q(id_estado_actividad=3) & ~Q(id_estado_actividad=5) & ~Q(
                        id_estado_actividad=10)))

            )

        if nivel == '4':
            lista_datos_count = Ges_Actividad_Historia.objects.annotate(month=Extract('id_periodo_seguimiento__fecha_termino_corte', 'month')).filter( Q(id_periodo_seguimiento__fecha_termino_corte__month=mes_filtro) & Q(id_controlador__id_jefatura__id_nivel__id_cuarto_nivel=id_nivel) & Q(id_periodo=periodo_actual) ).values(
                'month',
                'id_controlador__nivel_inicial',

                'id_controlador__id_jefatura__id_nivel__descripcion_nivel').annotate(
                CountNoIniciadas=Count('id', filter=Q(id_estado_actividad=4)),
                CountFinalizadas=Count('id', filter=Q(id_estado_actividad=7)),
                CountConRetraso=Count('id', filter=Q(id_estado_actividad=1)),
                CountSinMovimiento=Count('id', filter=Q(id_estado_actividad=6)),
                CountEliminadas=Count('id', filter=Q(id_estado_actividad=9)),
                CountEnCurso=Count('id', filter=Q(id_estado_actividad=8)),
                CountTotal=Count('id', filter=(~Q(id_estado_actividad=2) & ~Q(id_estado_actividad=3) & ~Q(id_estado_actividad=5) & ~Q(id_estado_actividad=10)))

            )


        context['object_count'] = lista_datos_count


        # # # # # # # # # # # # # # # # # # # # # #  FIN ESTADO CUMPLIMIENTO  # # # # # # # # # # # # # # # # # # # # # #






        meses = Ges_Actividad_Historia.objects.annotate(month=Extract('id_periodo_seguimiento__fecha_termino_corte', 'month')).filter(Q(id_periodo=periodo_actual)).values('month').distinct().order_by('-month')
        context['meses_periodo'] = meses



        return render(request, template_name, context)





    def get_context_data(self, **kwargs):
        context = super(GeneraReportCurvaEjecucion, self).get_context_data(**kwargs)

        try:
            periodo_actual = Glo_Periodos.objects.get(id_estado=1)
        except Glo_Periodos.DoesNotExist:
            return None




        lista_datos_unidades_select = Ges_Niveles.objects.values(
            'orden_nivel',
            'id_cuarto_nivel__descripcion_nivel',
            'id_tercer_nivel__descripcion_nivel',
            'id_segundo_nivel__descripcion_nivel',
            # 'id_primer_nivel__descripcion_nivel',

            'id_cuarto_nivel_id',
            'id_tercer_nivel_id',
            'id_segundo_nivel_id',
            # 'id_primer_nivel_id',
            # # N4
            # 'id_controlador__id_jefatura__id_nivel__id_cuarto_nivel__tercer_nivel__descripcion_nivel',  # N3
            # 'id_controlador__id_jefatura__id_nivel__id_cuarto_nivel__descripcion_nivel',  # N2
            #
            # 'id_controlador__id_jefatura__id_nivel__id_cuarto_nivel__tercer_nivel__segundo_nivel__id',  # N4_id
            # 'id_controlador__id_jefatura__id_nivel__id_cuarto_nivel__tercer_nivel__id',  # N3_id
            # 'id_controlador__id_jefatura__id_nivel__id_cuarto_nivel__id',  # N2_id
        ).exclude(orden_nivel=1)


        context['lista_datos_unidades'] = lista_datos_unidades_select


        seguimiento_cerrado= Ges_Actividad_Historia.objects.annotate(
            month=Extract('id_periodo_seguimiento__fecha_termino_corte', 'month')).filter(
            Q(id_periodo=periodo_actual)).values(
            'id_controlador__nivel_inicial',

        ).distinct()

        context['seguimiento_cerrado'] = seguimiento_cerrado # Verifica que exista un periodo cerrado


        meses = Ges_Actividad_Historia.objects.annotate(month=Extract('id_periodo_seguimiento__fecha_termino_corte', 'month')).filter(Q(id_periodo=periodo_actual)).values('month').distinct().order_by('-month')

        context['meses_periodo'] = meses

        return context


def export_users_xls(request):
    try:
        periodo_actual = Glo_Periodos.objects.get(id_estado=1)
    except Glo_Periodos.DoesNotExist:
        return None

    nivel = request.session['nivel']
    id_nivel = request.session['id_nivel']
    mes_filtro = request.session['mes_filtro']

    #actividades = Ges_Actividad.objects.all()

    actividades=''

    if nivel=='4':
        actividades = Ges_Actividad_Historia.objects.filter(Q(id_controlador__id_jefatura__id_nivel__id_cuarto_nivel__tercer_nivel__segundo_nivel=id_nivel) &
                                                            Q(id_periodo=periodo_actual) &
                                                            Q(id_periodo_seguimiento__fecha_termino_corte__month=mes_filtro))

    if nivel=='3':
        actividades = Ges_Actividad_Historia.objects.filter(Q(id_controlador__id_jefatura__id_nivel__id_tercer_nivel=id_nivel) &
                                                            Q(id_periodo=periodo_actual) &
                                                            Q(id_periodo_seguimiento__fecha_termino_corte__month=mes_filtro))


    if nivel=='2':
        actividades = Ges_Actividad_Historia.objects.filter(Q(id_controlador__id_jefatura__id_nivel__id_cuarto_nivel=id_nivel) &
                                                            Q(id_periodo=periodo_actual) &
                                                            Q(id_periodo_seguimiento__fecha_termino_corte__month=mes_filtro))

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-reporte_seguimiento.xlsx'.format(
        date=datetime.datetime.now().strftime('%d/%m/%Y'),
    )
    workbook = Workbook()


    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'reporte_seguimiento'

    # Define the titles for columns

    columns = ['Actividad',
               'Objetivo Vinculado',
               'Periodicidad',
               'Producto Estadístico',
               'Hora x Actividad',
               'Volumen',
               'N° Personas Asignadas',
               'Total Horas',
               'Cargo',
               'Fecha Incio Actividad',
               'Fecha Término Actividad',
               'Estado Actividad',
               'Fecha Real Finalización',
               'Reprogramación Fecha Inicio',
               'Reprogramación Fecha Término',
               'Justificación Desviación',
               'Dependencia 1',
               'Dependencia 2',
               'Área',
               'Eje'
               ]

    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    for actividad in actividades:
        row_num += 1

        row =''

        if nivel=='4':

            row = [

                actividad.id_actividad.descripcion_actividad,
                str(actividad.id_actividad.id_objetivo_operativo) ,
                str(actividad.id_actividad.id_periodicidad),
                str(actividad.id_actividad.id_producto_estadistico),
                actividad.id_actividad.horas_actividad,
                actividad.id_actividad.volumen,
                actividad.id_actividad.personas_asignadas,
                actividad.id_actividad.total_horas,
                str(actividad.id_actividad.id_familia_cargo),
                actividad.id_actividad.fecha_inicio_actividad,
                actividad.id_actividad.fecha_termino_actividad,
                str(actividad.id_estado_actividad),
                actividad.fecha_real_termino,
                actividad.fecha_reprogramacion_inicio,
                actividad.fecha_reprogramacion_termino,
                actividad.justificacion,
                str(actividad.id_actividad.id_objetivo_operativo.id_cuarto_nivel.tercer_nivel.segundo_nivel),
                str(actividad.id_actividad.id_objetivo_operativo.id_cuarto_nivel.tercer_nivel),
                str(actividad.id_actividad.id_objetivo_operativo.id_cuarto_nivel),
                str(actividad.id_actividad.id_objetivo_operativo.id_objetivo_tacticotn.id_objetivo_tactico.id_objetivo_estrategico.ges_eje),

            ]

        if nivel=='3':

            row = [

                actividad.id_actividad.descripcion_actividad,
                str(actividad.id_actividad.id_objetivo_tacticotn) ,
                str(actividad.id_actividad.id_periodicidad),
                str(actividad.id_actividad.id_producto_estadistico),
                actividad.id_actividad.horas_actividad,
                actividad.id_actividad.volumen,
                actividad.id_actividad.personas_asignadas,
                actividad.id_actividad.total_horas,
                str(actividad.id_actividad.id_familia_cargo),
                actividad.id_actividad.fecha_inicio_actividad,
                actividad.id_actividad.fecha_termino_actividad,
                str(actividad.id_estado_actividad),
                actividad.fecha_real_termino,
                actividad.fecha_reprogramacion_inicio,
                actividad.fecha_reprogramacion_termino,
                actividad.justificacion,
                str(actividad.id_actividad.id_objetivo_tacticotn.id_tercer_nivel.segundo_nivel),
                str(actividad.id_actividad.id_objetivo_tacticotn.id_tercer_nivel),
                str('-'),
                str(actividad.id_actividad.id_objetivo_tacticotn.id_objetivo_tactico.id_objetivo_estrategico.ges_eje),


            ]



        if nivel=='2':

            row = [

                actividad.id_actividad.descripcion_actividad,
                str(actividad.id_actividad.id_objetivo_tactico) ,
                str(actividad.id_actividad.id_periodicidad),
                str(actividad.id_actividad.id_producto_estadistico),
                actividad.id_actividad.horas_actividad,
                actividad.id_actividad.volumen,
                actividad.id_actividad.personas_asignadas,
                actividad.id_actividad.total_horas,
                str(actividad.id_actividad.id_familia_cargo),
                actividad.id_actividad.fecha_inicio_actividad,
                actividad.id_actividad.fecha_termino_actividad,
                str(actividad.id_estado_actividad),
                actividad.fecha_real_termino,
                actividad.fecha_reprogramacion_inicio,
                actividad.fecha_reprogramacion_termino,
                actividad.justificacion,
                str(actividad.id_actividad.id_objetivo_tactico.id_segundo_nivel),
                str('-'),
                str('-'),
                str(actividad.id_actividad.id_objetivo_tactico.id_objetivo_estrategico.ges_eje),


            ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value


            if col_num == 10:
                cell.number_format = 'dd/mm/yyyy'
            if col_num == 11:
                cell.number_format = 'dd/mm/yyyy'
            if col_num == 13:
                cell.number_format = 'dd/mm/yyyy'
            if col_num == 14:
                cell.number_format = 'dd/mm/yyyy'
            if col_num == 15:
                cell.number_format = 'dd/mm/yyyy'

    workbook.save(response)

    return response


# def export_users_xls(request):
#     """
#     Downloads all movies as Excel file with a single worksheet
#     """
#     actividades = Ges_Actividad.objects.all()
#
#
#
#     response = HttpResponse(
#         content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
#     )
#     response['Content-Disposition'] = 'attachment; filename={date}-reporte_seguimiento.xlsx'.format(
#         date=datetime.datetime.now().strftime('%d/%m/%Y'),
#     )
#     workbook = Workbook()
#
#
#     # Get active worksheet/tab
#     worksheet = workbook.active
#     worksheet.title = 'reporte_seguimiento'
#
#     # Define the titles for columns
#
#     columns = ['Actividad',
#                'Objetivo Vinculado',
#                'Periodicidad',
#                'Producto Estadístico',
#                'Hora x Actividad',
#                'Volumen',
#                'N° Personas Asignadas',
#                'Total Horas',
#                'Cargo',
#                'Fecha Incio Actividad',
#                'Fecha Término Actividad',
#                'Estado Actividad',
#                'Fecha Real Finalización',
#                'Reprogramación Fecha Inicio',
#                'Reprogramación Fecha Término',
#                'Justificación Desviación',
#                'Dependencia 1',
#                'Dependencia 2',
#                'Área',
#                'Eje'
#                ]
#
#     row_num = 1
#
#     # Assign the titles for each cell of the header
#     for col_num, column_title in enumerate(columns, 1):
#         cell = worksheet.cell(row=row_num, column=col_num)
#         cell.value = column_title
#
#     # Iterate through all movies
#     for actividad in actividades:
#         row_num += 1
#         #
#         # Define the data for each cell in the row
#
#         Nivel=actividad.id_controlador.nivel_inicial
#
#         if Nivel==4:
#
#             row = [
#
#                 actividad.descripcion_actividad,
#                 str(actividad.id_objetivo_operativo) ,
#                 str(actividad.id_periodicidad),
#                 str(actividad.id_producto_estadistico),
#                 actividad.horas_actividad,
#                 actividad.volumen,
#                 actividad.personas_asignadas,
#                 actividad.total_horas,
#                 str(actividad.id_familia_cargo),
#                 actividad.fecha_inicio_actividad,
#                 actividad.fecha_termino_actividad,
#                 str(actividad.id_estado_actividad),
#                 actividad.fecha_real_termino,
#                 actividad.fecha_reprogramacion_inicio,
#                 actividad.fecha_reprogramacion_termino,
#                 actividad.justificacion,
#                 str(actividad.id_objetivo_operativo.id_cuarto_nivel.tercer_nivel.segundo_nivel),
#                 str(actividad.id_objetivo_operativo.id_cuarto_nivel.tercer_nivel),
#                 str(actividad.id_objetivo_operativo.id_cuarto_nivel),
#                 str(actividad.id_objetivo_operativo.id_objetivo_tacticotn.id_objetivo_tactico.id_objetivo_estrategico.ges_eje),
#
#
#             ]
#
#         if Nivel==3:
#
#             row = [
#
#                 actividad.descripcion_actividad,
#                 str(actividad.id_objetivo_tacticotn) ,
#                 str(actividad.id_periodicidad),
#                 str(actividad.id_producto_estadistico),
#                 actividad.horas_actividad,
#                 actividad.volumen,
#                 actividad.personas_asignadas,
#                 actividad.total_horas,
#                 str(actividad.id_familia_cargo),
#                 actividad.fecha_inicio_actividad,
#                 actividad.fecha_termino_actividad,
#                 str(actividad.id_estado_actividad),
#                 actividad.fecha_real_termino,
#                 actividad.fecha_reprogramacion_inicio,
#                 actividad.fecha_reprogramacion_termino,
#                 actividad.justificacion,
#                 str(actividad.id_objetivo_tacticotn.id_tercer_nivel.segundo_nivel),
#                 str(actividad.id_objetivo_tacticotn.id_tercer_nivel),
#                 str('-'),
#                 str(actividad.id_objetivo_tacticotn.id_objetivo_tactico.id_objetivo_estrategico.ges_eje),
#
#
#             ]
#
#
#
#         if Nivel==2:
#
#             row = [
#
#                 actividad.descripcion_actividad,
#                 str(actividad.id_objetivo_tactico) ,
#                 str(actividad.id_periodicidad),
#                 str(actividad.id_producto_estadistico),
#                 actividad.horas_actividad,
#                 actividad.volumen,
#                 actividad.personas_asignadas,
#                 actividad.total_horas,
#                 str(actividad.id_familia_cargo),
#                 actividad.fecha_inicio_actividad,
#                 actividad.fecha_termino_actividad,
#                 str(actividad.id_estado_actividad),
#                 actividad.fecha_real_termino,
#                 actividad.fecha_reprogramacion_inicio,
#                 actividad.fecha_reprogramacion_termino,
#                 actividad.justificacion,
#                 str(actividad.id_objetivo_tactico.id_segundo_nivel),
#                 str('-'),
#                 str('-'),
#                 str(actividad.id_objetivo_tactico.id_objetivo_estrategico.ges_eje),
#
#
#             ]
#
#
#
#
#         # Assign the data for each cell of the row
#         for col_num, cell_value in enumerate(row, 1):
#             cell = worksheet.cell(row=row_num, column=col_num)
#             cell.value = cell_value
#
#
#             if col_num == 10:
#                 cell.number_format = 'dd/mm/yyyy'
#             if col_num == 11:
#                 cell.number_format = 'dd/mm/yyyy'
#             if col_num == 13:
#                 cell.number_format = 'dd/mm/yyyy'
#             if col_num == 14:
#                 cell.number_format = 'dd/mm/yyyy'
#             if col_num == 15:
#                 cell.number_format = 'dd/mm/yyyy'
#
#     workbook.save(response)
#
#     return response







