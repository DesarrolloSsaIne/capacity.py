from django.shortcuts import render
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView
from apps.periodos.forms import periodosForm, Seguimiento_cierreform, Seguimiento_abrirform, validacion_abrirform, valida_cierreform, PeriodoAnual_cierreform
from apps.periodos.models import Glo_Periodos, Glo_Seguimiento, Glo_validacion, Glo_EstadoPeriodo, RespaldoPeriodo
from apps.controlador.models import Ges_Controlador
from apps.jefaturas.models import Ges_Jefatura
from apps.estado_seguimiento.models import Glo_EstadoSeguimiento
from apps.registration.models import logEventos
from django.contrib import messages
from django.contrib.messages.views import SuccessMessageMixin
from django.http import HttpResponseRedirect
from django.db.models.deletion import ProtectedError
from django.core.mail import EmailMessage,send_mass_mail
from openpyxl import Workbook
from django.http import HttpResponse

from django.db.models import Q
from datetime import date
from datetime import datetime
from django.utils import timezone
import mysql.connector
from mysql.connector import Error

from apps.valida_plan2.models import *
from apps.valida_plan.models import *
from apps.actividades.models import Ges_Actividad_Historia, Ges_Observaciones_valida
from apps.eje.models import *
from apps.feriados.models import *
from apps.gestion_horas.models import *
from apps.estructura.models import *
from apps.objetivos.models import *
from apps.jefaturas.models import *

from django.core.mail import send_mail
from django.conf import settings

# Create your views here.


class PeriodosList(ListView):
    model = Glo_Periodos
    template_name = 'periodos/periodo_list.html'


class PeriodosCreate(SuccessMessageMixin, CreateView):
    model = Glo_Periodos
    form_class = periodosForm
    template_name = 'periodos/periodo_form.html'


    def get_context_data(self, **kwargs):
        context = super(PeriodosCreate, self).get_context_data(**kwargs)

        try:
             estado_Periodo_Anual= Glo_Periodos.objects.get(Q(id_estado=1))
        except Glo_Periodos.DoesNotExist:
             estado_Periodo_Anual = 0

        try:
             estado_Validacion= Glo_validacion.objects.get(Q(id_estado_periodo=1) )
        except Glo_validacion.DoesNotExist:
             estado_Validacion = 0

        try:
             estado_seguimiento= Glo_Seguimiento.objects.get(Q(id_estado_seguimiento=1) )
        except Glo_Seguimiento.DoesNotExist:
             estado_seguimiento = 0

        # try:
        #     planes_no_aprobados= Ges_Controlador.objects.filter(Q(id_periodo=PeriodoActual()) & (~Q(estado_flujo_id=7) & ~Q(estado_flujo_id=2))).count()
        # except Ges_Controlador.DoesNotExist:
        #     planes_no_aprobados = 0

        context['estado_periodos'] = {'periodoAnual': estado_Periodo_Anual, 'periodoSeguimiento': estado_seguimiento, 'periodoValidacion':estado_Validacion}
        return context



    def post(self, request, *args, **kwargs):
        form_class = self.get_form_class()
        form = self.get_form(form_class)

        if form.is_valid():




            try:

                form.instance.fecha_inicio= date.today()
                form.instance.id_estado_id='1'
                form.save()
                id_nuevo_periodo = form.instance.id



            except:
                request.session['message_class'] = "alert alert-danger"
                messages.error(self.request, "Error interno: Hubo un error al intentar realizar la apertura del periodo anual, inténtelo nuevamente, si error persiste comuníquese con el administrador.")
                return HttpResponseRedirect('/periodos/listar')

            try:

                AbrirPeriodoAnual(id_nuevo_periodo)

            except:
                request.session['message_class'] = "alert alert-warning"
                messages.error(self.request, "Estimado Usuario, el periodo fue abierto exitosamente pero hubo un problema al resetear proceso actual, comuníquese con el administrador.")
                return HttpResponseRedirect('/periodos/listar')

            try:
                EnviarCorreoAbrirPeriodoAnual()
               # logEventosCreate('Cierre Periodo Anual', 'PeriodosAnualCerrar()', request.POST['id_user'], request.POST['id_user'])
                request.session['message_class'] = "alert alert-success"
                messages.success(self.request, "Estimado Usuario: El periodo anual fue abierto exitosamente, se realizó el reseteo de la información y el correo se envió a las jefaturas.!" )
                return HttpResponseRedirect('/periodos/listar')

            except:
                request.session['message_class'] = "alert alert-warning"
                messages.error(self.request, "Estimado Usuario, el respaldo y el cierre del proceso fueron realizado con éxito, pero hubo un problema al intentar enviar el correo, favor informar a las jefaturas del cierre.")
                return HttpResponseRedirect('/periodos/listar')


        else:
            request.session['message_class'] = "alert alert-danger"
            messages.error(self.request, "Error interno: No se ha actualizado el registro. Comuníquese con el administrador.")
            return HttpResponseRedirect('/periodos/listar')





class PeriodosAnualCerrar(SuccessMessageMixin, UpdateView ):
    model = Glo_Periodos
    form_class = PeriodoAnual_cierreform
    template_name = 'periodos/periodo_delete.html'

    def post(self, request, *args, **kwargs):

        self.object = self.get_object
        id_periodo = kwargs['pk']
        instancia_nivel = self.model.objects.get(id=id_periodo)
        form = self.form_class(request.POST, instance=instancia_nivel)
        id_nuevo_estado= Glo_EstadoPeriodo.objects.get(id=3)

        if form.is_valid():
            form.instance.id_estado=id_nuevo_estado
            form.instance.fecha_termino= datetime.now(tz=timezone.utc)


            try:

                respalda_periodo()
            except:
                request.session['message_class'] = "alert alert-danger"
                messages.error(self.request, "Error interno: Hubo un error al intentar realizar el respaldo, inténtelo nuevamente, si error persiste comuníquese con el administrador.")
                return HttpResponseRedirect('/periodos/listar')

            try:

                form.save()

            except:
                request.session['message_class'] = "alert alert-warning"
                messages.error(self.request, "Estimado Usuario, el respaldo fue realizado pero hubo un problema al intentar cerrar el proceso, comuníquese con el administrador.")
                return HttpResponseRedirect('/periodos/listar')

            try:
                EnviarCorreoCerrarPeriodoAnual()
               # logEventosCreate('Cierre Periodo Anual', 'PeriodosAnualCerrar()', request.POST['id_user'], request.POST['id_user'])
                request.session['message_class'] = "alert alert-success"
                messages.success(self.request, "Estimado Usuario: El periodo anual fue cerrado exitosamente, se realizó el respaldo de la información y el correo se envió a las jefaturas.!" )
                return HttpResponseRedirect('/periodos/listar')

            except:
                request.session['message_class'] = "alert alert-warning"
                messages.error(self.request, "Estimado Usuario, el respaldo y el cierre del proceso fueron realizado con éxito, pero hubo un problema al intentar enviar el correo, favor informar a las jefaturas del cierre.")
                return HttpResponseRedirect('/periodos/listar')


        else:
            request.session['message_class'] = "alert alert-danger"
            messages.error(self.request, "Error interno: No se ha actualizado el registro. Comuníquese con el administrador.")
            return HttpResponseRedirect('/periodos/listar')




class SeguimientoList(ListView):
    model = Glo_Seguimiento
    template_name = 'periodos/seguimiento_list.html'



    def get_context_data(self, **kwargs):
        context = super(SeguimientoList, self).get_context_data(**kwargs)
        #id_usuario_actual = self.request.user.id  # obtiene id usuario actual
        self.request.session['id_periodo']=self.kwargs['pk'] #guarda id  controlador


        queryset= Glo_Seguimiento.objects.filter(id_periodo=self.kwargs['pk'])


        context['object_list'] = queryset
        context['periodo'] = PeriodoActual()




        return context


class SeguimientoCerrarPeriodo(UpdateView):
    model = Glo_Seguimiento
    template_name = 'periodos/seguimiento_cerrar.html'
    form_class = Seguimiento_cierreform

    def post(self, request, *args, **kwargs):
        self.object = self.get_object
        pk = kwargs['pk']
        instancia = self.model.objects.get(id=pk)
        form = self.form_class(request.POST, instance=instancia)

        id_nuevo_estado= Glo_EstadoSeguimiento.objects.get(id=2)

        id_periodo=self.request.session['id_periodo']

        if form.is_valid():

            form.instance.id_estado_seguimiento=id_nuevo_estado
            form.instance.fecha_termino= datetime.now(tz=timezone.utc)

            form.save()

            try:

                EnviarCorreoCierreSeguimiento()

                request.session['message_class'] = "alert alert-success"  # Tipo mensaje
                messages.success(request, "La etapa de seguimiento fue cerrada y se le ha enviado un correo a las jefaturas que formulan.!")  # mensaje
                return HttpResponseRedirect('/periodos/listar_seguimiento/' + str(id_periodo))  # Redirije a la pantalla principal

            except:

                 request.session['message_class'] = "alert alert-warning" #Tipo mensaje
                 messages.success(request, "La etapa de seguimiento fue cerrada correctamente!, pero el servicio de correo tuvo un inconveniente favor comuníquese con las jefaturas que formulan para informar el cierre de esta etapa.") # mensaje
                 return HttpResponseRedirect('/periodos/listar_seguimiento/' + str(id_periodo)) # Redirije a la pantalla principal

        else:

            request.session['message_class'] = "alert alert-danger"
            messages.error(self.request, "Error interno: No se ha cerrado la etapa de seguimiento. Comuníquese con el administrador.")
            return HttpResponseRedirect('/periodos/listar_seguimiento' + str(id_periodo))



class SeguimientoAbrirPeriodo(SuccessMessageMixin, CreateView):
    model = Glo_Seguimiento
    template_name = 'periodos/seguimiento_abrir.html'
    form_class = Seguimiento_abrirform


    def get_context_data(self, **kwargs):
        context = super(SeguimientoAbrirPeriodo, self).get_context_data(**kwargs)

        try:
             estado= Glo_Seguimiento.objects.get(Q(id_estado_seguimiento=1) & Q(id_periodo=PeriodoActual()))
        except Glo_Seguimiento.DoesNotExist:
             estado = 0

        try:
             estadoValida= Glo_validacion.objects.get(Q(id_estado_periodo=1) & Q(id_periodo=PeriodoActual()))
        except Glo_validacion.DoesNotExist:
             estadoValida = 0

        try:
            planes_no_aprobados= Ges_Controlador.objects.filter(Q(id_periodo=PeriodoActual()) & (~Q(estado_flujo_id=7) & ~Q(estado_flujo_id=2))).count()
        except Ges_Controlador.DoesNotExist:
            planes_no_aprobados = 0

        context["estado_seguimiento"] = {'estadoValida': estadoValida, 'estado_seguimiento': estado, 'planes_no_aprobados':planes_no_aprobados}
        return context

    def post(self, request, *args, **kwargs):
        form_class = self.get_form_class()
        form = self.get_form(form_class)


        id_nuevo_estado= Glo_EstadoSeguimiento.objects.get(id=1)
        periodo_actual = Glo_Periodos.objects.get(id_estado=1)

        id_periodo=self.request.session['id_periodo']

        if form.is_valid():
            form.instance.id_periodo= periodo_actual
            form.instance.id_estado_seguimiento=id_nuevo_estado
            form.instance.fecha_inicio= datetime.now(tz=timezone.utc)

            form.save()

            try:
                Ges_Controlador.objects.filter(id_periodo=PeriodoActual()).update(id_estado_plan=1)
                EnviarCorreoAbrirSeguimiento()

                request.session['message_class'] = "alert alert-success"  # Tipo mensaje
                messages.success(request, "El periodo de seguimiento fue abierto correctamente! y se ha enviado un correo a las jefaturas que formulan!")  # mensaje
                return HttpResponseRedirect('/periodos/listar_seguimiento/' + str(id_periodo))  # Redirije a la pantalla principal

            except:

                 request.session['message_class'] = "alert alert-warning" #Tipo mensaje
                 messages.success(request, "El periodo de seguimiento fue abierto correctamente!, pero el servicio de correo tuvo un inconveniente favor comuníquese con la jefatura directa para informar de la apertura.") # mensaje
                 return HttpResponseRedirect('/periodos/listar_seguimiento/' + str(id_periodo)) # Redirije a la pantalla principal

        else:

            request.session['message_class'] = "alert alert-danger"
            messages.error(self.request, "Error interno: No se ha abierto la etapa de seguimiento. Comuníquese con el administrador.")
            return HttpResponseRedirect('/periodos/listar_seguimiento/' + str(id_periodo))

class ValidacionCerrarPeriodo(UpdateView):
    model = Glo_validacion
    template_name = 'periodos/validacion_cerrar.html'
    form_class = valida_cierreform

    def post(self, request, *args, **kwargs):
        self.object = self.get_object
        pk = kwargs['pk']
        instancia = self.model.objects.get(id=pk)
        form = self.form_class(request.POST, instance=instancia)

        id_nuevo_estado= Glo_EstadoSeguimiento.objects.get(id=2)

        id_periodo=self.request.session['id_periodo']

        if form.is_valid():

            form.instance.id_estado_periodo=id_nuevo_estado
            #form.instance.fecha_termino= datetime.now(tz=timezone.utc)

            form.save()

            try:

                #EnviarCorreoCierre()
                EnviarCorreoCierreValidacion()

                request.session['message_class'] = "alert alert-success"  # Tipo mensaje
                messages.success(request, "La etapa de validación fue cerrada y se le ha enviado un correo a las jefaturas que formulan.!")  # mensaje
                return HttpResponseRedirect('/periodos/listar_validacion/' + str(id_periodo))  # Redirije a la pantalla principal

            except:

                 request.session['message_class'] = "alert alert-warning" #Tipo mensaje
                 messages.success(request, "La etapa de validación fue cerrada correctamente!, pero el servicio de correo tuvo un inconveniente favor comuníquese con las jefaturas que formulan para informar el cierre de esta etapa.") # mensaje
                 return HttpResponseRedirect('/periodos/listar_validacion/' + str(id_periodo)) # Redirije a la pantalla principal

        else:

            request.session['message_class'] = "alert alert-danger"
            messages.error(self.request, "Error interno: No se ha cerrado la etapa de validación. Comuníquese con el administrador.")
            return HttpResponseRedirect('/periodos/listar_validacion/' + str(id_periodo))


class ValidacionList(ListView):
    model = Glo_validacion
    template_name = 'periodos/validacion_list.html'

    def get_context_data(self, **kwargs):
        context = super(ValidacionList, self).get_context_data(**kwargs)
        #id_usuario_actual = self.request.user.id  # obtiene id usuario actual
        self.request.session['id_periodo']=self.kwargs['pk'] #guarda id  controlador


        queryset= Glo_validacion.objects.filter(id_periodo=self.kwargs['pk']).order_by('-id')


        context['object_list'] = queryset
        context['periodo'] = PeriodoActual()




        return context

class ValidacionAbrirPeriodo(SuccessMessageMixin, CreateView):
    model = Glo_validacion
    template_name = 'periodos/validacion_abrir.html'
    form_class = validacion_abrirform


    def get_context_data(self, **kwargs):
        context = super(ValidacionAbrirPeriodo, self).get_context_data(**kwargs)

        try:
             estadoValida= Glo_validacion.objects.get(Q(id_estado_periodo=1) & Q(id_periodo=PeriodoActual()))
        except Glo_validacion.DoesNotExist:
             estadoValida = 0

        try:
            estadoSeguimiento = Glo_Seguimiento.objects.get(Q(id_estado_seguimiento=1) & Q(id_periodo=PeriodoActual()))
        except Glo_Seguimiento.DoesNotExist:
            estadoSeguimiento = 0

        context["estado_seguimiento"] = {'estadoValida': estadoValida, 'estado_seguimiento': estadoSeguimiento}

        return context

    def post(self, request, *args, **kwargs):
        form_class = self.get_form_class()
        form = self.get_form(form_class)


        id_nuevo_estado= Glo_EstadoSeguimiento.objects.get(id=1)
        periodo_actual = Glo_Periodos.objects.get(id_estado=1)

        id_periodo=self.request.session['id_periodo']

        if form.is_valid():
            form.instance.id_periodo= periodo_actual
            form.instance.id_estado_periodo=id_nuevo_estado
            form.instance.fecha_inicio= datetime.now(tz=timezone.utc)

            form.save()

            try:

                EnviarCorreoAbrirValidacion()


                request.session['message_class'] = "alert alert-success"  # Tipo mensaje
                messages.success(request, "El periodo de validación fue abierto correctamente! y se ha enviado un correo a las jefaturas que formulan!")  # mensaje
                return HttpResponseRedirect('/periodos/listar_validacion/' + str(id_periodo))  # Redirije a la pantalla principal

            except:

                 request.session['message_class'] = "alert alert-warning" #Tipo mensaje
                 messages.success(request, "El periodo de validación fue abierto correctamente!, pero el servicio de correo tuvo un inconveniente favor comuníquese con la jefatura directa para informar de la apertura.") # mensaje
                 return HttpResponseRedirect('/periodos/listar_validacion/' + str(id_periodo)) # Redirije a la pantalla principal

        else:

            request.session['message_class'] = "alert alert-danger"
            messages.error(self.request, "Error interno: No se ha abierto la etapa de validación. Comuníquese con el administrador.")
            return HttpResponseRedirect('/periodos/listar_validacion/' + str(id_periodo))


def PeriodoActual():

    try:
        periodo_actual = Glo_Periodos.objects.get(id_estado=1)
    except Glo_Periodos.DoesNotExist:
        return None
    return periodo_actual


def EnviarCorreoCierreValidacion():

    controladorPlan = Ges_Jefatura.objects.values_list('id_user__email' , flat=True).filter(id_periodo=PeriodoActual())

    ahora = datetime.now()
    fecha = ahora.strftime("%d" + " de " + "%B" + " de " + "%Y")

    idcorreoJefatura=list(controladorPlan)

    subject = 'Cierre etapa Validación'
    messageHtml = '<b>Estimada(o) Usuaria(o) del Sistema Capacity Institucional</b> ,<br> Le informamos que con fecha  '+ str(fecha) +', el proceso de <b>Validación</B> ha sido cerrado. <br><p style="font-size:12px;color:red;">correo generado automaticamente favor no responder.'

    email = EmailMessage(subject, messageHtml ,to=idcorreoJefatura)
    email.content_subtype='html'
    email.send()


def EnviarCorreoCierreSeguimiento():

    controladorPlan = Ges_Jefatura.objects.values_list('id_user__email' , flat=True).filter(id_periodo=PeriodoActual())

    ahora = datetime.now()
    fecha = ahora.strftime("%d" + " de " + "%B" + " de " + "%Y")

    idcorreoJefatura=list(controladorPlan)

    subject = 'Cierre etapa Seguimiento'
    messageHtml = '<b>Estimada(o) Usuaria(o) del Sistema Capacity Institucional</b> ,<br> Le informamos que con fecha  '+ str(fecha) +', el proceso de <b>seguimiento</B> ha sido cerrado. <br><p style="font-size:12px;color:red;">correo generado automaticamente favor no responder.'

    email = EmailMessage(subject, messageHtml ,to=idcorreoJefatura)
    email.content_subtype='html'
    email.send()



def EnviarCorreoAbrirValidacion():



    controladorPlan = Ges_Jefatura.objects.values_list('id_user__email', flat=True).filter(
        id_periodo=PeriodoActual())

    ahora = datetime.now()
    fecha = ahora.strftime("%d" + " de " + "%B" + " de " + "%Y")

    idcorreoJefatura=list(controladorPlan)

    subject = 'Apertura etapa Validación'
    messageHtml = '<b>Estimada(o) Usuaria(o) del Sistema Capacity Institucional</b> ,<br> Le informamos que con fecha  '+ str(fecha) +', el proceso de <b>VALIDACIÓN</B> ha sido abierto para que valide las actividades reportadas por el equipo. <br><p style="font-size:12px;color:red;">correo generado automaticamente favor no responder.'

    email = EmailMessage(subject, messageHtml ,to=idcorreoJefatura)
    email.content_subtype='html'
    email.send()


def EnviarCorreoAbrirSeguimiento():



    controladorPlan = Ges_Jefatura.objects.values_list('id_user__email', flat=True).filter(
        id_periodo=PeriodoActual())

    ahora = datetime.now()
    fecha = ahora.strftime("%d" + " de " + "%B" + " de " + "%Y")

    idcorreoJefatura=list(controladorPlan)

    subject = 'Apertura etapa Seguimiento'
    messageHtml = '<b>Estimada(o) Usuaria(o) del Sistema Capacity Institucional</b> ,<br> Le informamos que con fecha  '+ str(fecha) +', el proceso de <b>Seguimiento</B> ha sido abierto para que reporte el avance de sus actividades. <br><p style="font-size:12px;color:red;">correo generado automaticamente favor no responder.'

    email = EmailMessage(subject, messageHtml ,to=idcorreoJefatura)
    email.content_subtype='html'
    email.send()

def EnviarCorreoAbrirPeriodoAnual():

    controladorPlan = Ges_Jefatura.objects.values_list('id_user__email', flat=True).filter(
        id_periodo=PeriodoActual())

    ahora = datetime.now()
    fecha = ahora.strftime("%d" + " de " + "%B" + " de " + "%Y")

    #idcorreoJefatura=list(controladorPlan) Habilitar solo al publicar
    idcorreoJefatura='jmrodriguezc@ine.gob.cl'
    subject = 'Apertura Proceso Anual de Planificación'
    messageHtml = '<b>Estimada(o) Usuaria(o) del Sistema Capacity Institucional</b> ,<br> Le informamos que con fecha  '+ str(fecha) +', el <b>Periodo Anual de Planificación</B> ha sido abierto por el administrador. <br><p style="font-size:12px;color:red;">correo generado automaticamente favor no responder.'

    email = EmailMessage(subject, messageHtml,from_email=settings.EMAIL_HOST_USER ,to=['jmrodriguezc@ine.gob.cl'])
    email.content_subtype='html'
    email.send()

def EnviarCorreoCerrarPeriodoAnual():

    controladorPlan = Ges_Jefatura.objects.values_list('id_user__email', flat=True).filter(
        id_periodo=PeriodoActual())

    ahora = datetime.now()
    fecha = ahora.strftime("%d" + " de " + "%B" + " de " + "%Y")

    idcorreoJefatura=list(controladorPlan)


    subject = 'Cierre Proceso Anual de Planificación'
    messageHtml = '<b>Estimada(o) Usuaria(o) del Sistema Capacity Institucional</b> ,<br> Le informamos que con fecha  '+ str(fecha) +', el <b>Periodo Anual de Planificación</B> ha sido cerrado por el administrador. <br><p style="font-size:12px;color:red;">correo generado automaticamente favor no responder.'

    email = EmailMessage(subject,  messageHtml ,to=['jmrodriguezc@ine.gob.cl'])
    #email = EmailMessage(subject,  messageHtml,from_email=settings.EMAIL_HOST_USER ,to=idcorreoJefatura)
    email.content_subtype='html'
    email.send()

    #send_mail(subject='Apertura Proceso Anual de Planificación',message=messageHtml,from_email=settings.EMAIL_HOST_USER, recipient_list=['jmrodriguezc@ine.gob.cl'])

def ExportarBackupXls(request, *args, **kwargs):

    id_periodo = kwargs['pk']

    respaldo_periodo = RespaldoPeriodo.objects.filter(id_gp=id_periodo)


    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-RespaldoPeriodo.xlsx'.format(
        date=datetime.now().strftime('%d/%m/%Y'),
    )
    workbook = Workbook()


    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'respaldo_periodo'

    # Define the titles for columns

    columns = ['Id Primer Nivel',
               'Primer Nivel',
               'Id Segundo Nivel',
               'Segundo Nivel',
               'Id Tercer Nivel',
               'Tercer Nivel',
               'Id Cuarto Nivel',
               'Cuarto Nivel',
               'Ejes',
               'Id Objetivo Estrategico',
               'Objetivo Estrategico',
               'Id Objetivo Tactico',
               'Objetivo Tactico',
               'Id Objetivo Tactico Terce Nivel',
               'Objetivo Tactico Tercer Nivel',
               'Id Objetivo Operativo',
               'Objetivo Operativo',
               'Id Actividad',
               'Descripcion Actividad',
               'Horas Actividad',
               'Volumen',
               'Personas Asignadas',
               'Total Horas',
               'Fecha Inicio Actividad',
               'Fecha Termino Actividad',
               'Fecha Registro',
               'Estado',
               'Fecha Real Inicio',
               'Fecha Reas Termino',
               'Fecha Reprogramacion Termino',
               'Justificacion',
               'Validada',
               'Flag Tmp',
               'Flag Reporta',
               'Descripcion Estado',
               'Descripcion Familia Cargo',
               'Descripcion Periodicidad',
               'Id Periodo',
               'Descripcion Periodo',
               'Año Periodo',
               'Fecha Inicio Periodo',
               'Fecha Termino Periodo',
               'Producto Estadistico',
               'Flag Finalizada',
               'Jefatura',
               'Estado Flujo Controlador',
               'Estado Plan Controlador',
               'Analista Asignado',
               'Primera Revision',
               'Segunda Revision'
               ]

    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    for campos in respaldo_periodo:
        row_num += 1

        row =''

        row = [
                str(campos.id_pn),
                str(campos.primer_nivel),
                str(campos.id_sn),
                str(campos.segundo_nivel),
                str(campos.id_tni),
                str(campos.tercer_nivel),
                str(campos.id_cn),
                str(campos.cuarto_nivel),
                str(campos.ejes),
                str(campos.id_oe),
                str(campos.objetivo_estrategico),
                str(campos.id_ot),
                str(campos.objetivo_tactico),
                str(campos.id_tn),
                str(campos.objetivo_tacticotn),
                str(campos.id_op),
                str(campos.objetivo_operativo),
                str(campos.id_actividad),
                str(campos.descripcion_actividad),
                str(campos.horas_actividad),
                str(campos.volumen),
                str(campos.personas_asignadas),
                str(campos.total_horas),
                str(campos.fecha_inicio_actividad),
                str(campos.fecha_termino_actividad),
                str(campos.fecha_registro),
                str(campos.estado),
                str(campos.fecha_real_inicio),
                str(campos.fecha_real_termino),
                str(campos.fecha_reprogramacion_termino),
                str(campos.justificacion),
                str(campos.validada),
                str(campos.flag_tmp),
                str(campos.flag_reporta),
                str(campos.descripcion_estado),
                str(campos.descripcion_familiacargo),
                str(campos.descripcion_periodicidad),
                str(campos.id_gp),
                str(campos.descripcion_periodo),
                str(campos.anio_periodo),
                str(campos.fecha_inicio_periodo),
                str(campos.fecha_termino_periodo),
                str(campos.producto_estadistico),
                str(campos.flag_finalizada),
                str(campos.jefatura),
                str(campos.estado_flujo_controlador),
                str(campos.estado_plan_controlador),
                str(campos.analista_asignado),
                str(campos.primera_revision),
                str(campos.segunda_revision),
            ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value


            if col_num == 23:
                cell.number_format = 'dd/mm/yyyy'
            if col_num == 24:
                cell.number_format = 'dd/mm/yyyy'
            if col_num == 25:
                cell.number_format = 'dd/mm/yyyy'
            if col_num == 27:
                cell.number_format = 'dd/mm/yyyy'
            if col_num == 28:
                cell.number_format = 'dd/mm/yyyy'
            if col_num == 29:
                cell.number_format = 'dd/mm/yyyy'
            if col_num == 40:
                cell.number_format = 'dd/mm/yyyy'
            if col_num == 41:
                cell.number_format = 'dd/mm/yyyy'

    workbook.save(response)


    return response




def logEventosCreate(tipo_evento, metodo ,usuario_evento, jefatura_dirigida):
    logEventos.objects.create(
        tipo_evento=tipo_evento,
        metodo=metodo,
        usuario_evento=usuario_evento,
        jefatura_dirigida=jefatura_dirigida,
    )
    return None

def AbrirPeriodoAnual(periodo_id):

    try:
        Ges_Observaciones.objects.all().delete()
        Ges_Observaciones.objects.all().delete()
        Ges_Observaciones_sr.objects.all().delete()
        Ges_Observaciones_valida.objects.all().delete()
        Ges_Actividad_Historia.objects.all().delete()
         #get(id= ) Ges_Actividad.objects.all().update
        Ges_Actividad.objects.all().update(fecha_inicio_actividad= None, fecha_termino_actividad  = None, fecha_real_inicio  = None,  fecha_real_termino  = None,
                                       fecha_reprogramacion_termino  = None, fecha_reprogramacion_inicio = None, validada = 0, flag_tmp   = 0,
                                       flag_reporta = 0, id_estado_actividad_id = 4, flag_finalizada  = 0, id_periodo_id = periodo_id, justificacion=None)

        Ges_Controlador.objects.all().update(estado_flujo_id = 2, id_periodo_id= periodo_id, id_estado_plan_id = 1)
        Ges_Ejes.objects.all().update( id_periodo_id=periodo_id)
        #Ges_Feriados.objects.get(id=123).update( id_periodo=periodo_id)
        Ges_Registro_Horas.objects.all().update(tiene_vacaciones= 0, fecha_inicio = None, fecha_termino=None,dias_habiles = 0, id_periodo_id=periodo_id)
        Ges_CuartoNivel.objects.all().update( id_periodo_id=periodo_id)
        Ges_TercerNivel.objects.all().update( id_periodo_id=periodo_id)
        Ges_SegundoNivel.objects.all().update( id_periodo_id=periodo_id)
        Ges_PrimerNivel.objects.all().update( id_periodo_id=periodo_id)
        Ges_Niveles.objects.all().update( id_periodo_id=periodo_id)
        Ges_Objetivo_Estrategico.objects.all().update( id_periodo_id=periodo_id)
        Ges_Objetivo_TacticoTN.objects.all().update( id_periodo_id=periodo_id)
        Ges_Objetivo_Tactico.objects.all().update( id_periodo_id=periodo_id)
        Ges_Objetivo_Operativo.objects.all().update( id_periodo_id=periodo_id)
        Ges_Jefatura.objects.all().update( id_periodo_id=periodo_id)



    except Exception as e:
        print % (e.message, type(e))











def respalda_periodo():


    try:
        resp= True
        connection = mysql.connector.connect(host='10.91.160.53',
                                             database='DJANGO_DES',
                                             user='capacity_des',
                                             password='2MOMO99L')
        if connection.is_connected():
            db_Info = connection.get_server_info()
            print("Connected to MySQL Server version ", db_Info)

            periodo_actual = Glo_Periodos.objects.get(id_estado=1)

            RespaldoPeriodo.objects.filter(id_gp=periodo_actual.id).delete()
            cursor = connection.cursor()
            #cursor.execute("delete from periodos_respaldoperiodo where id_gp = (SELECT id from periodos_glo_periodos WHERE id_estado_id= 1 )")
            print("respaldo periodo activo")
            #respaldo periodo activo
            cursor.execute("""INSERT into periodos_respaldoperiodo(
                             id_pn  ,
                            primer_nivel ,
                            id_sn ,
                             segundo_nivel ,
                             id_tni ,
                             tercer_nivel ,
                             id_cn , 
                            cuarto_nivel ,
                             ejes ,
                             id_oe  ,
                            objetivo_estrategico ,
                             id_ot  ,
                             objetivo_tactico ,
                             id_tn ,
                             objetivo_tacticotn ,
                             id_op ,
                            objetivo_operativo , 
                            id_actividad , 
                            descripcion_actividad ,
                            horas_actividad  ,
                             volumen ,
                             personas_asignadas  ,
                             total_horas ,
                             fecha_inicio_actividad ,
                             fecha_termino_actividad ,
                             fecha_registro ,
                             estado ,
                             fecha_real_inicio ,
                             fecha_real_termino ,
                             fecha_reprogramacion_termino ,
                             justificacion ,
                             validada ,
                             flag_tmp  ,
                             flag_reporta  ,
                             descripcion_estado ,
                             descripcion_familiacargo ,
                             descripcion_periodicidad ,
                              id_gp ,
                             descripcion_periodo ,
                             anio_periodo ,
                              fecha_inicio_periodo ,
                              fecha_termino_periodo , 
                              producto_estadistico , 
                             flag_finalizada ,
                             jefatura ,
                              estado_flujo_controlador ,
                             estado_plan_controlador ,
                            analista_asignado ,
                            primera_revision ,
                             segunda_revision )
                            SELECT 
                            pn.id id_pn,
                            pn.descripcion_nivel primer_nivel,
                            sn.id id_sn,
                             sn.descripcion_nivel segundo_nivel,
                             tni.id id_tni,
                             tni.descripcion_nivel tercer_nivel,
                             cn.id id_cn,
                             cn.descripcion_nivel cuarto_nivel,
                             (SELECT eje.descripcion_eje from eje_ges_ejes eje WHERE eje.id = oe.ges_eje_id) ejes,
                             oe.id id_oe,
                             oe.descripcion_objetivo objetivo_estrategico,
                             ot.id id_ot,
                             ot.descripcion_objetivo objetivo_tactico,
                             tn.id id_tn,
                             tn.descripcion_objetivo objetivo_tacticotn,
                             op.id id_op,
                             op.descripcion_objetivo objetivo_operativo,
                              ac.id id_actividad, 
                             ac.descripcion_actividad,
                             ac.horas_actividad,
                             ac.volumen,
                             ac.personas_asignadas,
                             ac.total_horas,
                             ac.fecha_inicio_actividad,
                             ac.fecha_termino_actividad,
                             ac.fecha_registro,
                             ac.estado,
                             ac.fecha_real_inicio,
                             ac.fecha_real_termino,
                             ac.fecha_reprogramacion_termino,
                             ac.justificacion,
                             case
                             when ac.validada= 0 then 'no validada'
                             when ac.validada = 1 then 'validada'
                             ELSE '' END validada,
                             ac.flag_tmp,
                             ac.flag_reporta,
                             CONCAT(ea.descripcion_estado, IFNULL(ea.descripcion_validada, ''))descripcion_estado,
                             fa.descripcion_familiacargo,
                             pe.descripcion_periodicidad,
                             gp.id id_gp,
                             gp.descripcion_periodo,
                             gp.anio_periodo anio_periodo,
                             gp.fecha_inicio fecha_inicio_periodo,
                             gp.fecha_termino fecha_termino_periodo, 
                             pes.descripcion_producto producto_estadistico,
                             ac.flag_finalizada,
                             CONCAT(usr.first_name, " ",usr.last_name) jefatura,
                             ef.descripcion_estado estado_flujo_controlador,
                             epl.descripcion_estado estado_plan_controlador,
                             
                             (SELECT CONCAT(us.first_name, " ", us.last_name) analista_asignado FROM auth_user us
                             where us.id = con.analista_asignado_id) analista_asignado,
                             
                              (SELECT CONCAT(us.first_name," ", us.last_name) primera_revision FROM auth_user us
                             where us.id = (SELECT je.id_user_id from jefaturas_ges_jefatura je WHERE je.id = con.jefatura_primerarevision_id)) primera_revision,
                             
                              (SELECT CONCAT(us.first_name," ", us.last_name) segunda_revision FROM auth_user us
                             where us.id = (SELECT je.id_user_id from jefaturas_ges_jefatura je WHERE je.id = con.jefatura_segundarevision_id)) segunda_revision
                            from objetivos_ges_objetivo_estrategico oe
                            INNER JOIN estructura_ges_primernivel pn ON oe.ges_primer_nivel_id = pn.id
                            INNER JOIN objetivos_ges_objetivo_tactico ot ON ot.id_objetivo_estrategico_id = oe.id
                            INNER JOIN estructura_ges_segundonivel sn ON sn.id = ot.id_segundo_nivel_id
                            INNER JOIN objetivos_ges_objetivo_tacticotn tn ON tn.id_objetivo_tactico_id = ot.id
                            INNER JOIN estructura_ges_tercernivel tni ON tni.id = tn.id_tercer_nivel_id 
                            INNER JOIN objetivos_ges_objetivo_operativo op ON op.id_objetivo_tacticotn_id = tn.id
                            INNER JOIN estructura_ges_cuartonivel cn ON op.id_cuarto_nivel_id = cn.id
                            inner JOIN actividades_ges_actividad ac ON ac.id_objetivo_operativo_id = op.id
                            INNER JOIN controlador_ges_controlador con ON con.id = ac.id_controlador_id
                            INNER JOIN jefaturas_ges_jefatura je ON je.id = con.id_jefatura_id
                            INNER JOIN auth_user usr ON usr.id= je.id_user_id
                            INNER JOIN estado_actividad_glo_estadoactividad ea ON ea.id = ac.id_estado_actividad_id
                            INNER JOIN familia_cargo_glo_familiacargo fa ON fa.id = ac.id_familia_cargo_id
                            INNER JOIN periodicidad_glo_periodicidad pe ON pe.id = ac.id_periodicidad_id
                            INNER JOIN periodos_glo_periodos gp ON gp.id = ac.id_periodo_id
                            LEFT JOIN productos_glo_productosestadisticos pes ON pes.id = ac.id_producto_estadistico_id
                            INNER JOIN estado_flujo_glo_estadoflujo ef ON  ef.id = con.estado_flujo_id
                            INNER JOIN estado_plan_glo_estadoplan epl ON epl.id = con.id_estado_plan_id
                            WHERE gp.id_estado_id = 1 
                            UNION ALL
                            SELECT 
                            pn.id id_pn,
                            pn.descripcion_nivel primer_nivel,
                            sn.id id_sn,
                             sn.descripcion_nivel segundo_nivel,
                             tni.id id_tni,
                             tni.descripcion_nivel tercer_nivel,
                             NULL id_cn, 
                            NULL cuarto_nivel,
                             (SELECT eje.descripcion_eje from eje_ges_ejes eje WHERE eje.id = oe.ges_eje_id) ejes,
                             oe.id id_oe,
                             oe.descripcion_objetivo objetivo_estrategico,
                             ot.id id_ot,
                             ot.descripcion_objetivo objetivo_tactico,
                             tn.id id_tn,
                             tn.descripcion_objetivo objetivo_tacticotn,
                             NULL id_op,
                            NULL objetivo_Operativo, 
                             ac.id id_actividad, 
                             ac.descripcion_actividad,
                             ac.horas_actividad,
                             ac.volumen,
                             ac.personas_asignadas,
                             ac.total_horas,
                             ac.fecha_inicio_actividad,
                             ac.fecha_termino_actividad,
                             ac.fecha_registro,
                             ac.estado,
                             ac.fecha_real_inicio,
                             ac.fecha_real_termino,
                             ac.fecha_reprogramacion_termino,
                             ac.justificacion,
                            case
                             when ac.validada= 0 then 'no validada'
                             when ac.validada = 1 then 'validada'
                             ELSE '' END validada,
                             ac.flag_tmp,
                             ac.flag_reporta,
                             CONCAT(ea.descripcion_estado, IFNULL(ea.descripcion_validada, ''))descripcion_estado,
                             fa.descripcion_familiacargo,
                             pe.descripcion_periodicidad,
                             gp.id id_gp,
                             gp.descripcion_periodo,
                             gp.anio_periodo anio_periodo,
                             gp.fecha_inicio fecha_inicio_periodo,
                             gp.fecha_termino fecha_termino_periodo, 
                             pes.descripcion_producto producto_estadistico,
                             ac.flag_finalizada,
                             CONCAT(usr.first_name, " ",usr.last_name) jefatura,
                             ef.descripcion_estado estado_flujo_controlador,
                             epl.descripcion_estado estado_plan_controlador,
                             
                               (SELECT CONCAT(us.first_name, " ",us.last_name) analista_asignado FROM auth_user us
                             where us.id = con.analista_asignado_id) analista_asignado,
                              (SELECT CONCAT(us.first_name, " ",us.last_name) primera_revision FROM auth_user us
                             where us.id = (SELECT je.id_user_id from jefaturas_ges_jefatura je WHERE je.id = con.jefatura_primerarevision_id)) primera_revision,
                             
                              (SELECT CONCAT(us.first_name, " ",us.last_name) segunda_revision FROM auth_user us
                             where us.id = (SELECT je.id_user_id from jefaturas_ges_jefatura je WHERE je.id = con.jefatura_segundarevision_id)) segunda_revision
                            from objetivos_ges_objetivo_estrategico oe
                            INNER JOIN estructura_ges_primernivel pn ON oe.ges_primer_nivel_id = pn.id
                            INNER JOIN objetivos_ges_objetivo_tactico ot ON ot.id_objetivo_estrategico_id = oe.id
                            INNER JOIN estructura_ges_segundonivel sn ON sn.id = ot.id_segundo_nivel_id
                            INNER JOIN objetivos_ges_objetivo_tacticotn tn ON tn.id_objetivo_tactico_id = ot.id
                            INNER JOIN estructura_ges_tercernivel tni ON tni.id = tn.id_tercer_nivel_id 
                            inner JOIN actividades_ges_actividad ac ON  ac.id_objetivo_tacticotn_id = tn.id
                            INNER JOIN controlador_ges_controlador con ON con.id = ac.id_controlador_id
                            INNER JOIN jefaturas_ges_jefatura je ON je.id = con.id_jefatura_id
                            INNER JOIN auth_user usr ON usr.id= je.id_user_id
                            INNER JOIN estado_actividad_glo_estadoactividad ea ON ea.id = ac.id_estado_actividad_id
                            INNER JOIN familia_cargo_glo_familiacargo fa ON fa.id = ac.id_familia_cargo_id
                            INNER JOIN periodicidad_glo_periodicidad pe ON pe.id = ac.id_periodicidad_id
                            INNER JOIN periodos_glo_periodos gp ON gp.id = ac.id_periodo_id
                            LEFT JOIN productos_glo_productosestadisticos pes ON pes.id = ac.id_producto_estadistico_id
                            INNER JOIN estado_flujo_glo_estadoflujo ef ON  ef.id = con.estado_flujo_id
                            INNER JOIN estado_plan_glo_estadoplan epl ON epl.id = con.id_estado_plan_id
                            WHERE gp.id_estado_id = 1 
                            UNION ALL 
                            SELECT 
                            pn.id id_pn,
                            pn.descripcion_nivel primer_nivel,
                            sn.id id_sn,
                             sn.descripcion_nivel segundo_nivel,
                             NULL id_tni,
                             NULL  tercer_nivel,
                             NULL id_cn, 
                            NULL cuarto_nivel,
                             (SELECT eje.descripcion_eje from eje_ges_ejes eje WHERE eje.id = oe.ges_eje_id) ejes,
                             oe.id id_oe,
                             oe.descripcion_objetivo objetivo_estrategico,
                             ot.id id_ot,
                             ot.descripcion_objetivo objetivo_tactico,
                             NULL id_tn,
                             NULL  objetivo_tacticotn,
                             NULL id_op,
                            NULL objetivo_Operativo, 
                             ac.id id_actividad, 
                             ac.descripcion_actividad,
                             ac.horas_actividad,
                             ac.volumen,
                             ac.personas_asignadas,
                             ac.total_horas,
                             ac.fecha_inicio_actividad,
                             ac.fecha_termino_actividad,
                             ac.fecha_registro,
                             ac.estado,
                             ac.fecha_real_inicio,
                             ac.fecha_real_termino,
                             ac.fecha_reprogramacion_termino,
                             ac.justificacion,
                            case
                             when ac.validada= 0 then 'no validada'
                             when ac.validada = 1 then 'validada'
                             ELSE '' END validada,
                             ac.flag_tmp,
                             ac.flag_reporta,
                             CONCAT(ea.descripcion_estado, IFNULL(ea.descripcion_validada, ''))descripcion_estado,
                             fa.descripcion_familiacargo,
                             pe.descripcion_periodicidad,
                             gp.id id_gp,
                             gp.descripcion_periodo,
                             gp.anio_periodo anio_periodo,
                             gp.fecha_inicio fecha_inicio_periodo,
                             gp.fecha_termino fecha_termino_periodo, 
                             pes.descripcion_producto producto_estadistico,
                             ac.flag_finalizada,
                              CONCAT(usr.first_name, " ",usr.last_name) jefatura,
                             ef.descripcion_estado estado_flujo_controlador,
                             epl.descripcion_estado estado_plan_controlador,
                            
                              (SELECT CONCAT(us.first_name, " ",us.last_name) analista_asignado FROM auth_user us
                             where us.id = con.analista_asignado_id) analista_asignado,
                              (SELECT CONCAT(us.first_name, " ",us.last_name) primera_revision FROM auth_user us
                             where us.id = (SELECT je.id_user_id from jefaturas_ges_jefatura je WHERE je.id = con.jefatura_primerarevision_id))primera_revision,
                             
                              (SELECT CONCAT(us.first_name, " ",us.last_name) segunda_revision FROM auth_user us
                             where us.id = (SELECT je.id_user_id from jefaturas_ges_jefatura je WHERE je.id = con.jefatura_segundarevision_id)) segunda_revision
                            from objetivos_ges_objetivo_estrategico oe
                            INNER JOIN estructura_ges_primernivel pn ON oe.ges_primer_nivel_id = pn.id
                            INNER JOIN objetivos_ges_objetivo_tactico ot ON ot.id_objetivo_estrategico_id = oe.id
                            INNER JOIN estructura_ges_segundonivel sn ON sn.id = ot.id_segundo_nivel_id
                            inner JOIN actividades_ges_actividad ac ON  ot.id= ac.id_objetivo_tactico_id
                            INNER JOIN controlador_ges_controlador con ON con.id = ac.id_controlador_id
                            INNER JOIN jefaturas_ges_jefatura je ON je.id = con.id_jefatura_id
                            INNER JOIN auth_user usr ON usr.id= je.id_user_id
                            INNER JOIN estado_actividad_glo_estadoactividad ea ON ea.id = ac.id_estado_actividad_id
                            INNER JOIN familia_cargo_glo_familiacargo fa ON fa.id = ac.id_familia_cargo_id
                            INNER JOIN periodicidad_glo_periodicidad pe ON pe.id = ac.id_periodicidad_id
                            INNER JOIN periodos_glo_periodos gp ON gp.id = ac.id_periodo_id
                            LEFT JOIN productos_glo_productosestadisticos pes ON pes.id = ac.id_producto_estadistico_id
                            INNER JOIN estado_flujo_glo_estadoflujo ef ON  ef.id = con.estado_flujo_id
                            INNER JOIN estado_plan_glo_estadoplan epl ON epl.id = con.id_estado_plan_id       
                            WHERE gp.id_estado_id = 1 
                            
                            
                            """)

            connection.commit()
            cursor.execute("delete from actividades_ges_actividad_historia_respaldo where id_periodo_id = (SELECT id from periodos_glo_periodos WHERE id_estado_id= 1 ) ")
            print("respaldo actividades_ges_actividad_historia_respaldo ")
            cursor.execute(""" 
            INSERT INTO actividades_ges_actividad_historia_respaldo  (
                          id ,
                          fecha_registro,
                          fecha_reprogramacion_inicio,
                          fecha_reprogramacion_termino,
                          fecha_real_termino,
                          justificacion,
                          id_actividad_id ,
                          id_estado_actividad_id,
                          id_periodo_id,
                          id_periodo_seguimiento_id,
                          id_controlador_id,
                          fecha_real_inicio,
                          validada
                        
                        )SELECT    id ,
                          fecha_registro,
                          fecha_reprogramacion_inicio,
                          fecha_reprogramacion_termino,
                          fecha_real_termino,
                          justificacion,
                          id_actividad_id ,
                          id_estado_actividad_id,
                          id_periodo_id,
                          id_periodo_seguimiento_id,
                          id_controlador_id,
                          fecha_real_inicio,
                          validada
                        FROM actividades_ges_actividad_historia where id_periodo_id =(SELECT id from periodos_glo_periodos WHERE id_estado_id= 1 )
            """)
            connection.commit()
            cursor.execute("delete from gestion_horas_ges_registro_horas_respaldo where id_periodo_id =(SELECT id from periodos_glo_periodos WHERE id_estado_id= 1 ) ")
            print("respaldo gestion_horas_ges_registro_horas_respaldo ")
            cursor.execute(""" 
                    INSERT INTO gestion_horas_ges_registro_horas_respaldo (
                                  id ,
                                  tiene_vacaciones ,
                                  fecha_inicio,
                                  fecha_termino,
                                  dias_habiles,
                                  notas,
                                  fecha_insercion ,
                                  id_familiacargo_id,
                                  id_nivel_id ,
                                  id_periodo_id,
                                  id_user_id 
                                ) 
                                SELECT id ,
                                  tiene_vacaciones ,
                                  fecha_inicio,
                                  fecha_termino,
                                  dias_habiles,
                                  notas,
                                  fecha_insercion ,
                                  id_familiacargo_id,
                                  id_nivel_id ,
                                  id_periodo_id,
                                  id_user_id
                                   FROM gestion_horas_ges_registro_horas where id_periodo_id =(SELECT id from periodos_glo_periodos WHERE id_estado_id= 1 ) """)
            connection.commit()
            print("respaldo valida_plan2_ges_observaciones_sr_respaldo ")
            cursor.execute( "delete from valida_plan2_ges_observaciones_sr_respaldo where id_periodo_id = (SELECT id from periodos_glo_periodos WHERE id_estado_id= 1 )")
            cursor.execute("""INSERT into valida_plan2_ges_observaciones_sr_respaldo  (
                                  id,
                                  observacion ,
                                  id_controlador ,
                                  fecha_registro ,
                                  observado ,
                                  id_jefe_observa_id ,
                                  id_periodo_id,
                                  user_observa_id,
                                  id_objetivo_operativo_id,
                                  id_objetivo_tactico_id,
                                  id_objetivo_tacticotn_id 
                                
                                ) 
                                SELECT   id,
                                  observacion ,
                                  id_controlador ,
                                  fecha_registro ,
                                  observado ,
                                  id_jefe_observa_id ,
                                  id_periodo_id,
                                  user_observa_id,
                                  id_objetivo_operativo_id,
                                  id_objetivo_tactico_id,
                                  id_objetivo_tacticotn_id 
                                  FROM valida_plan2_ges_observaciones_sr where id_periodo_id =(SELECT id from periodos_glo_periodos WHERE id_estado_id= 1 )""")
            connection.commit()
            print("respaldo valida_plan_ges_observaciones_respaldo ")
            cursor.execute(
                "delete from valida_plan_ges_observaciones_respaldo where id_periodo_id =(SELECT id from periodos_glo_periodos WHERE id_estado_id= 1 ) ")
            cursor.execute("""INSERT INTO valida_plan_ges_observaciones_respaldo(
                                      id  ,
                                      observacion,
                                      fecha_registro,
                                      id_actividad_id ,
                                      id_jefe_observa_id,
                                      id_periodo_id,
                                      user_observa_id,
                                      observado ,
                                      id_controlador,
                                      id_objetivo
                                    
                                    )
                                    SELECT   id  ,
                                      observacion,
                                      fecha_registro,
                                      id_actividad_id ,
                                      id_jefe_observa_id,
                                      id_periodo_id,
                                      user_observa_id,
                                      observado ,
                                      id_controlador,
                                      id_objetivo
                                    FROM valida_plan_ges_observaciones where id_periodo_id =(SELECT id from periodos_glo_periodos WHERE id_estado_id= 1 ) """)
            connection.commit()
            print("respaldo actividades_ges_observaciones_valida_respaldo ")
            cursor.execute( "delete from actividades_ges_observaciones_valida_respaldo  WHERE id_periodo_valida_id IN (SELECT id from periodos_glo_validacion gv WHERE gv.id_periodo_id IN (SELECT id FROM periodos_glo_periodos gp WHERE gp.id_estado_id = 1  ) ) ")
            cursor.execute("""INSERT INTO 
                            actividades_ges_observaciones_valida_respaldo
                            (                       
                             id  ,
                              descripcion_observacion,
                              fecha_registro,
                              id_actividad_id,
                              id_periodo_valida_id
                            
                            )                        
                            SELECT  id  ,
                              descripcion_observacion,
                              fecha_registro,
                              id_actividad_id,
                              id_periodo_valida_id FROM actividades_ges_observaciones_valida  WHERE id_periodo_valida_id IN (SELECT id from periodos_glo_validacion gv WHERE
										gv.id_periodo_id IN (SELECT id FROM periodos_glo_periodos gp WHERE gp.id_estado_id = 1  ) ) """)
            connection.commit()
            print("periodo respaldado")

    except Error as e:
        print("Error while connecting to MySQL", e)
        resp = False
        return resp
    finally:
        if connection.is_connected():
            cursor.close()
            connection.close()
            print("MySQL connection is closed")
            return resp
#resp = respalda_periodo()