from django.db.models import Q, Subquery, OuterRef, Count
from django.http import HttpResponseRedirect
from django.shortcuts import render
from apps.periodos.models import Glo_Seguimiento
# Create your views here.
from apps.actividades.models import Ges_Actividad, Ges_log_reportes, Ges_Actividad_Historia, Ges_Observaciones_valida
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView
from django.db.models import Case, CharField, Value, When
from apps.controlador.models import Ges_Controlador
from apps.estructura.models import Ges_Niveles
from apps.jefaturas.models import Ges_Jefatura
from apps.objetivos.models import Ges_Objetivo_Operativo, Ges_Objetivo_TacticoTN, Ges_Objetivo_Tactico
from apps.periodos.models import Glo_Periodos, Glo_Seguimiento
from django.contrib.messages.views import SuccessMessageMixin
from apps.estado_actividad.models import Glo_EstadoActividad
from apps.estado_plan.models import Glo_EstadoPlan
from apps.registration.models import logEventos
from apps.seguimiento_formula.forms import  GestionActividadesUpdateForm, PlanUpdateForm,PlanUpdateFormAcepta, GestionActividadesUpdateFormVer
from django.contrib import messages
from django.core.mail import EmailMessage,send_mass_mail
from django.contrib.auth.models import User, Group
from datetime import datetime
from django.http import HttpResponseRedirect, JsonResponse
from openpyxl import Workbook
from django.http import HttpResponse


class ActividadesObjetivosList(ListView): #clase modificada por JR- sprint 8 - Ok
    model= Ges_Actividad
    template_name = "seguimiento_formula/seguimiento_list.html"

    def get_context_data(self,   **kwargs,):
        context = super(ActividadesObjetivosList, self).get_context_data(**kwargs)
        id_usuario_actual= self.request.user.id #obtiene id usuario actual
        id_jefatura = Ges_Jefatura.objects.get(id_user=id_usuario_actual)

        try:
            periodo_actual = Glo_Periodos.objects.get(id_estado=1)
        except Glo_Periodos.DoesNotExist:
            return None
        try:
            #periodo_seguimiento = Glo_Seguimiento.objects.get(Q(id_estado_seguimiento=1) & Q(id_periodo=periodo_actual.id))
            periodo_seguimiento = Glo_Seguimiento.objects.filter(
                Q(id_periodo=periodo_actual.id)).order_by('-id')[0]
        except:
            periodo_seguimiento=None
            pass


        if periodo_seguimiento:

            try:
                id_controlador = Ges_Controlador.objects.get(Q(id_jefatura=id_jefatura) & Q(id_periodo=periodo_actual.id))
            except Ges_Controlador.DoesNotExist:
                id_controlador = 0
                pass

            if id_controlador != 0:
                 context['estado_flujo'] = {'estado':id_controlador.estado_flujo}

            if id_controlador == 0:
                context['error'] = {'id': 1}
                return context
                #sys.exit(1)
            else:

                try:
                    nivel_usuario= Ges_Jefatura.objects.get(Q(id_user=id_usuario_actual)  & Q(id_periodo=periodo_actual.id))

                except Ges_Jefatura.DoesNotExist:
                    context['habilitado'] = {'mensaje': False}
                    return None

                id_nivel =nivel_usuario.id_nivel.id
                replies = Ges_Niveles.objects.filter(Q(id=id_nivel) & Q(id_periodo=periodo_actual.id)).annotate(
                    nivel_order=Case(
                        When(orden_nivel=2, then='id_segundo_nivel'),
                        When(orden_nivel=3, then='id_tercer_nivel'),
                        When(orden_nivel=4, then='id_cuarto_nivel'),
                       output_field=CharField(),
                    )
                )
                for nivel in replies:
                    id_nivel_final = nivel.nivel_order
                    id_orden = nivel.orden_nivel

                    if id_orden == 2:
                        # replies2 = Ges_Objetivo_Tactico.objects.filter(Q(id_segundo_nivel=id_nivel_final) & Q(id_periodo=periodo_actual.id))

                        answer_subquery = Ges_Actividad.objects.values('id_objetivo_tactico_id').filter(
                            id_objetivo_tactico=OuterRef('pk')).annotate(
                            count_actividad=Count('id', filter=Q(id_periodo=periodo_actual.id)))



                        replies2 = Ges_Objetivo_Tactico.objects.filter(
                            Q(id_segundo_nivel=id_nivel_final) & Q(id_periodo=periodo_actual.id)).annotate(
                            count_actividad=Subquery(answer_subquery.values('count_actividad')))


                    if id_orden == 3:
                        # replies2 = Ges_Objetivo_TacticoTN.objects.filter(Q(id_tercer_nivel=id_nivel_final) & Q(id_periodo=periodo_actual.id))

                        answer_subquery = Ges_Actividad.objects.values('id_objetivo_tacticotn_id').filter(
                            id_objetivo_tacticotn=OuterRef('pk')).annotate(
                            count_actividad=Count('id', filter=Q(id_periodo=periodo_actual.id)))



                        replies2 = Ges_Objetivo_TacticoTN.objects.filter(
                            Q(id_tercer_nivel=id_nivel_final) & Q(id_periodo=periodo_actual.id)).annotate(
                            count_actividad=Subquery(answer_subquery.values('count_actividad')))

                    if id_orden == 4:
                        # replies2 = Ges_Objetivo_Operativo.objects.filter(Q(id_cuarto_nivel=id_nivel_final) & Q(id_periodo=periodo_actual.id))

                        answer_subquery = Ges_Actividad.objects.values('id_objetivo_operativo_id').filter(
                            id_objetivo_operativo=OuterRef('pk')).annotate(
                            count_actividad=Count('id', filter=Q(id_periodo=periodo_actual.id)))

                        replies2 = Ges_Objetivo_Operativo.objects.filter(
                            Q(id_cuarto_nivel=id_nivel_final) & Q(id_periodo=periodo_actual.id)).annotate(
                            count_actividad=Subquery(answer_subquery.values('count_actividad')))


                context['niveles'] = replies2

                context['orden'] = {'orden_nivel': id_orden}
                estado_seguimiento_id= periodo_seguimiento.id_estado_seguimiento_id
                context['total_disponible'] = {
                                                'id_jefatura': id_jefatura,
                                               'id_controlador': id_controlador,
                                                'estado_seguimiento_id': estado_seguimiento_id
                                               }

                context['nivel_usuario'] = replies

                self.request.session['id_orden'] = id_orden
                self.request.session['estado_seguimiento_id'] = estado_seguimiento_id
                return context
        else:
            self.request.session['message_class'] = "alert alert-warning"
            messages.error(self.request,
                           "Estimado usuario no existe un periodo de seguimiento activo. Favor comuniquese con el administrador")

            context['estado_seguimiento'] = 0
            return context



class ActividadesDetail(ListView): #clase modificada por JR- sprint 8 - Ok
    model = Ges_Actividad
    template_name = 'seguimiento_formula/seguimiento_actividades_detalle.html'

    def get_context_data(self,  **kwargs):
        context = super(ActividadesDetail, self).get_context_data(**kwargs)

        id_usuario_actual = self.request.user.id  # obtiene id usuario actual

        try:
            periodo_actual = Glo_Periodos.objects.get(id_estado=1)
        except Glo_Periodos.DoesNotExist:
            return None

        try:
            #periodo_seguimiento = Glo_Seguimiento.objects.get(Q(id_estado_seguimiento=1) & Q(id_periodo=periodo_actual.id))
            periodo_seguimiento = Glo_Seguimiento.objects.filter(
                Q(id_periodo=periodo_actual.id)).order_by('-id')[0]
        except:
            periodo_seguimiento=None
            pass

        nombre = ""
        if self.request.session['id_orden']==2:
            #lista_actividades = Ges_Actividad.objects.filter(Q(id_objetivo_tactico=self.kwargs['pk']) & Q(id_periodo=periodo_actual.id))

            lista_actividades = Ges_Actividad.objects.filter(
                Q(id_objetivo_tactico=self.kwargs['pk']) & Q(id_periodo=periodo_actual.id)).order_by('id_estado_actividad__orden','fecha_termino_actividad')

            nombre=  Ges_Objetivo_Tactico.objects.get(id=self.kwargs['pk'])




        if self.request.session['id_orden']==3:
           # lista_actividades = Ges_Actividad.objects.filter(Q(id_objetivo_tacticotn=self.kwargs['pk']) & Q(id_periodo=periodo_actual.id))


            lista_actividades = Ges_Actividad.objects.filter(
                Q(id_objetivo_tacticotn=self.kwargs['pk']) & Q(id_periodo=periodo_actual.id)).order_by('id_estado_actividad__orden','fecha_termino_actividad')


            nombre = Ges_Objetivo_TacticoTN.objects.get(id=self.kwargs['pk'])

        if self.request.session['id_orden']==4:
            #lista_actividades = Ges_Actividad.objects.filter(Q(id_objetivo_operativo=self.kwargs['pk']) & Q(id_periodo=periodo_actual.id))


            lista_actividades = Ges_Actividad.objects.filter(
                Q(id_objetivo_operativo=self.kwargs['pk']) & Q(id_periodo=periodo_actual.id)).order_by('id_estado_actividad__orden','fecha_termino_actividad')

            nombre = Ges_Objetivo_Operativo.objects.get(id=self.kwargs['pk'])





        self.request.session['tv'] = nombre.transversal

        try:
            nivel_usuario = Ges_Jefatura.objects.get(Q(id_user=id_usuario_actual) & Q(id_periodo=periodo_actual.id))
            id_nivel = nivel_usuario.id
        except Ges_Jefatura.DoesNotExist:
            context['habilitado'] = {'mensaje': False}
            return None

        try:
            id_jefatura = Ges_Jefatura.objects.get(id_user=id_usuario_actual)
            id_controlador = Ges_Controlador.objects.get(Q(id_jefatura=id_jefatura) & Q(id_periodo=periodo_actual.id))


        except Ges_Controlador.DoesNotExist:
            return None



        #estado_seguimiento_id = self.request.session['estado_seguimiento_id']
        context['object_list'] = lista_actividades
        context['nombre_objetivo'] = {'nombre': nombre}
        context['total_disponible'] = {'id_nivel':id_nivel,
                                       'id_controlador':id_controlador,
                                       'estado_seguimiento_id': periodo_seguimiento.id_estado_seguimiento_id
                                       }



        self.request.session['id_objetivo']=self.kwargs['pk']
        return context








class ActividadEdit(SuccessMessageMixin, UpdateView ):
    model = Ges_Actividad
    form_class = GestionActividadesUpdateForm
    template_name = 'seguimiento_formula/actividades_seguimiento_update.html'


    def get_context_data(self,  **kwargs):
        context = super(ActividadEdit, self).get_context_data(**kwargs)

        fechas_corte= Glo_Seguimiento.objects.order_by('-id')[0]
        flag_tmp= Ges_Actividad.objects.get(id=self.kwargs['pk'])




        context['fechas'] = {'fecha_inicio_corte_str':str(fechas_corte.fecha_inicio_corte)  ,
                             'fecha_termino_corte_str':str(fechas_corte.fecha_termino_corte),
                             'fecha_inicio_corte': fechas_corte.fecha_inicio_corte,
                             'fecha_termino_corte': fechas_corte.fecha_termino_corte,
                             'flag_tmp': flag_tmp.flag_tmp,
                             'fecha_real_inicio': flag_tmp.fecha_real_inicio,# sprint 3 - CI-21 -25012021
                             'fecha_real_inicio_str': str(flag_tmp.fecha_real_inicio),  # sprint 3 - CI-21 -25012021
                             'validada': str(flag_tmp.validada),  # sprint 3 - CI-21 -25012021

                             }
        return context



    def post(self, request, *args, **kwargs):
        form_class = self.get_form_class()
        self.object = self.get_object()
        id_actividad = kwargs['pk']
        instancia_nivel = self.model.objects.get(id=id_actividad)
        form = self.form_class(request.POST, instance=instancia_nivel)
        id_usuario_actual = self.request.user.id  # obtiene id usuario actual

        try:
            periodo_actual = Glo_Periodos.objects.get(id_estado=1)
        except Glo_Periodos.DoesNotExist:
            return None

        try:
            id_periodo_seguimiento = Glo_Seguimiento.objects.get(id_estado_seguimiento=1)
        except Glo_Seguimiento.DoesNotExist:
            return None

        try:
            id_jefatura = Ges_Jefatura.objects.get(Q(id_user=id_usuario_actual) & Q(id_periodo=periodo_actual.id))
        except Ges_Jefatura.DoesNotExist:
            return None

        try:
            usuario_controlador = Ges_Controlador.objects.get(id_jefatura=id_jefatura.id)
        except Ges_Controlador.DoesNotExist:
            return None


        fecha_real_inicio = request.POST['fecha_real_inicio'] #sprint 2 - CI -10 - 180120201


        fecha_real_termino = request.POST['fecha_real_termino']
        fecha_inicio_reprogramacion = request.POST['fecha_reprogramacion_inicio']
        fecha_termino_reprogramacion = request.POST['fecha_reprogramacion_termino']
        justificacion = request.POST['justificacion']

        id_estado_actividad = request.POST['id_estado_actividad']



        try:
            id_actividad_instancia = Ges_Actividad.objects.get(Q(id=id_actividad) & Q(id_periodo=periodo_actual.id))
        except id_actividad_instancia.DoesNotExist:
            return None

        try:
            actualiza_tmp = Ges_Actividad.objects.filter(Q(id=id_actividad) & Q(id_periodo=periodo_actual.id) & Q(flag_tmp=0))
        except id_actividad_instancia.DoesNotExist:
            return None

        if actualiza_tmp:
            form.instance.flag_tmp=0
        else:
            form.instance.flag_tmp=1

        try:
            id_estado_actividad_instancia = Glo_EstadoActividad.objects.get(id=id_estado_actividad)
        except id_estado_actividad_instancia.DoesNotExist:
            return None

        if fecha_real_inicio == '': #sprint 2 - CI -10 - 180120201
            form.instance.fecha_real_inicio = None
            fecha_real_inicio = None
        if fecha_real_termino == '':
            form.instance.fecha_real_termino = None
            fecha_real_termino = None
        if fecha_inicio_reprogramacion == '':
            form.instance.fecha_reprogramacion_inicio = None
            fecha_inicio_reprogramacion = None
        if fecha_termino_reprogramacion == '':
            form.instance.fecha_reprogramacion_termino = None
            fecha_termino_reprogramacion = None

        if form.is_valid():

            if self.request.session['id_orden'] == 2:
                id_objetivo = Ges_Objetivo_Tactico.objects.get(id=self.request.session['id_objetivo'])
                form.instance.id_objetivo_tactico = id_objetivo

            if self.request.session['id_orden'] == 3:
                id_objetivo = Ges_Objetivo_TacticoTN.objects.get(id=self.request.session['id_objetivo'])
                form.instance.id_objetivo_tacticotn = id_objetivo

            if self.request.session['id_orden'] == 4:
                id_objetivo = Ges_Objetivo_Operativo.objects.get(id=self.request.session['id_objetivo'])
                form.instance.id_objetivo_operativo = id_objetivo

            form.instance.flag_reporta = 1


            #form.instance.flag_tmp = 1 # Sprint 1 - CI-2 - 11012021

            id_controlador_id= id_actividad_instancia.id_controlador.id

            try:
                actualiza= Ges_Actividad_Historia.objects.get(Q(id_actividad=id_actividad) & Q(id_periodo_seguimiento=id_periodo_seguimiento) & Q(id_periodo=periodo_actual.id) & Q(id_controlador=id_actividad_instancia.id_controlador.id))

            except Ges_Actividad_Historia.DoesNotExist:
                actualiza = 0


            ActividadesHistoria(id_periodo_seguimiento, id_actividad_instancia, fecha_inicio_reprogramacion ,fecha_termino_reprogramacion, fecha_real_inicio, fecha_real_termino,
                                  id_estado_actividad_instancia, periodo_actual, justificacion, actualiza, id_controlador_id)

            form.save()
            request.session['message_class'] = "alert alert-success"
            messages.success(self.request, "Los datos fueron actualizados correctamente!")
            return HttpResponseRedirect('/seguimiento_formula/detalle/' + str(self.request.session['id_objetivo']))
        else:
            request.session['message_class'] = "alert alert-danger"
            messages.error(self.request,
                           "Error interno: No se ha creado el registro. Comuníquese con el administrador.")
            return HttpResponseRedirect('/seguimiento_formula/detalle/' + str(self.request.session['id_objetivo']))


class ObservacionesListar(ListView):
    model = Ges_Observaciones_valida
    template_name = 'seguimiento_formula/seguimiento_listar_comentarios.html'

    def get_context_data(self, **kwargs):
        context = super(ObservacionesListar, self).get_context_data(**kwargs)

        try:
            periodo_actual = Glo_Periodos.objects.get(id_estado=1)
        except Glo_Periodos.DoesNotExist:
            return None

        try:
            lista_comentarios = Ges_Observaciones_valida.objects.filter(id_actividad=self.kwargs['pk'])
        except:
            lista_comentarios = None
            pass

        nombre = ""

        nombre = Ges_Actividad.objects.get(id=self.kwargs['pk'])
        context['object_list'] = lista_comentarios
        context['nombre_actividad'] = {'nombre': nombre}

        return context


class ActividadDetallesVer(SuccessMessageMixin, UpdateView):
    model = Ges_Actividad
    form_class = GestionActividadesUpdateFormVer
    template_name = 'seguimiento_formula/seguimiento_formula_ver_detalle.html'



class iniciaSeguimiento(UpdateView):
    model = Ges_Controlador
    form_class = PlanUpdateFormAcepta
    template_name = 'seguimiento_formula/seguimiento_formula_inicia_form.html'

    def post(self, request, *args, **kwargs):
        self.object = self.get_object
        id_controlador = kwargs['pk']
        id_usuario_actual = self.request.user.id

        try:
            periodo_actual = Glo_Periodos.objects.get(id_estado=1)
        except Glo_Periodos.DoesNotExist:
            return None

        try:
            id_periodo_seguimiento = Glo_Seguimiento.objects.get(id_estado_seguimiento=1)
        except Glo_Seguimiento.DoesNotExist:
            return None

        controladorPlan = self.model.objects.get(Q(id=id_controlador) & Q(id_periodo=periodo_actual.id))

        area_plan = controladorPlan.id_jefatura.id_nivel

        #email_jefatura_primera = controladorPlan.jefatura_primerarevision.id_user.email  # correo de rechazo 1era revision

        try:
            email_jefatura_primera = controladorPlan.jefatura_primerarevision.id_user.email  # correo de rechazo 1era revision
        except:
            email_jefatura_primera = None
            pass

        analistas = list(User.objects.values_list('email', flat=True).filter(groups__in=Group.objects.filter(id='6')))
        email_jefatura= email_jefatura_primera
        emails_destino_analistas=analistas


        estado = 2
        controladorPlan.id_estado_plan_id = int(estado)

        try:

            controladorPlan.save()

            UpdateFlag(id_controlador, periodo_actual.id)

            tipo_evento = "Inicia seguimiento Plan"
            metodo = "Seguimiento - Formula"
            usuario_evento = self.request.user
            jefatura_dirigida = None
            logEventosCreate(tipo_evento, metodo, usuario_evento, jefatura_dirigida)
            CopiarActividades_a_Historial(id_periodo_seguimiento, id_controlador, periodo_actual)

        except:

            request.session['message_class'] = "alert alert-danger"
            messages.error(self.request,
                           "Error interno: El seguimiento no ha podido ser iniciado. Comuníquese con el administrador.")
            return HttpResponseRedirect('/seguimiento_formula/listar')


        try:
            EnviarCorreoInicioSeguimiento(emails_destino_analistas, email_jefatura, area_plan)

            request.session['message_class'] = "alert alert-success"  # Tipo mensaje
            messages.success(request,
                             "El periodo de seguimiento fue abierto correctamente! y se ha enviado un correo a su jefatura directa y a los analistas de gestión!")  # mensaje
            return HttpResponseRedirect('/seguimiento_formula/listar')  # Redirije a la pantalla principal

        except:

            request.session['message_class'] = "alert alert-warning"  # Tipo mensaje
            messages.success(request,
                             "El periodo de seguimiento fue abierto correctamente!, pero el servicio de correo tuvo un inconveniente favor comuníquese con la jefatura directa para informar de la apertura.")  # mensaje
            return HttpResponseRedirect('/seguimiento_formula/listar')  # Redirije a la pantalla principal


class cierraSeguimiento(UpdateView):
    model = Ges_Controlador
    form_class = PlanUpdateForm
    template_name = 'seguimiento_formula/seguimiento_formula_cierra_form.html'

    def post(self, request, *args, **kwargs):
        self.object = self.get_object
        id_controlador = kwargs['pk']
        id_usuario_actual = self.request.user.id

        try:
            periodo_actual = Glo_Periodos.objects.get(id_estado=1)
        except Glo_Periodos.DoesNotExist:
            return None

        controladorPlan = self.model.objects.get(Q(id=id_controlador) & Q(id_periodo=periodo_actual.id))


       # email_jefatura_primera = controladorPlan.jefatura_primerarevision.id_user.email  # correo de rechazo 1era revision

        try:
            email_jefatura_primera = controladorPlan.jefatura_primerarevision.id_user.email  # correo de rechazo 1era revision
        except:
            email_jefatura_primera = None
            pass

        analistas = list(User.objects.values_list('email', flat=True).filter(groups__in=Group.objects.filter(id='6')))
        email_jefatura= email_jefatura_primera
        emails_destino_analistas=analistas

        fechas_corte= Glo_Seguimiento.objects.order_by('-id')[0] # Sprint 1 - CI-3 - 12012021

        fecha_inicio_corte=fechas_corte.fecha_inicio_corte # Sprint 1 - CI-3 - 12012021
        fecha_termino_corte=fechas_corte.fecha_termino_corte # Sprint 1 - CI-3 - 12012021


        area_plan = controladorPlan.id_jefatura.id_nivel

        total_actividades_reportadas = Ges_Actividad.objects.filter(
            Q(id_controlador=id_controlador) & Q(fecha_real_inicio__isnull=True) & Q(id_periodo=periodo_actual) & Q(fecha_inicio_actividad__range=(fecha_inicio_corte, fecha_termino_corte)) & Q(flag_reporta=0) & (
                        ~Q(id_estado_actividad=7) & ~Q(id_estado_actividad=8) & ~Q(id_estado_actividad=9) & ~Q(Q(id_estado_actividad=5) & Q(id_periodicidad=9)))).count()

        # try:
        #
        #     total_actividades_reportadas = total_actividades_reportadas.get(
        #         Q(fecha_real_inicio__isnull=True) & ~Q(id_estado_actividad=8)).count()  # Sprint 2 - CI-18 - 20012021 nuevo
        # except total_actividades_reportadas.DoesNotExist:
        #     return 0




        estado = 3
        controladorPlan.id_estado_plan_id = int(estado)

        if total_actividades_reportadas == 0:

            try:
                controladorPlan.save()
                UpdateFlagCierre(id_controlador, periodo_actual.id)

                tipo_evento = "Cierra seguimiento Plan"
                metodo = "Seguimiento cierre - Formula"
                usuario_evento = self.request.user
                jefatura_dirigida = None
                logEventosCreate(tipo_evento, metodo, usuario_evento, jefatura_dirigida)

            except:

                request.session['message_class'] = "alert alert-danger"
                messages.error(self.request,
                               "Error interno: El seguimiento no ha podido ser cerrado. Comuníquese con el administrador.")
                return HttpResponseRedirect('/seguimiento_formula/listar')
        else:

            request.session['message_class'] = "alert alert-danger"
            messages.error(self.request,
                           "Estimado Usuario: Para cerrar el proceso de seguimiento debe actualizar todas las actividades que tengan fecha de inicio dentro de las fechas de corte.") # Sprint 1 - CI-3 - 12012021
            return HttpResponseRedirect('/seguimiento_formula/listar')


        try:
            EnviarCorreoCierreSeguimiento(emails_destino_analistas, email_jefatura, area_plan)

            request.session['message_class'] = "alert alert-success"  # Tipo mensaje
            messages.success(request,
                             "El periodo de seguimiento fue cerrado correctamente! y se ha enviado un correo a su jefatura directa y a los analistas de gestión!")  # mensaje
            return HttpResponseRedirect('/seguimiento_formula/listar')  # Redirije a la pantalla principal

        except:

            request.session['message_class'] = "alert alert-warning"  # Tipo mensaje
            messages.success(request,
                             "El periodo de seguimiento fue cerrado correctamente!, pero el servicio de correo tuvo un inconveniente favor comuníquese con la jefatura directa para informar de la apertura.")  # mensaje
            return HttpResponseRedirect('/seguimiento_formula/listar')  # Redirije a la pantalla principal




def logEventosCreate(tipo_evento, metodo ,usuario_evento, jefatura_dirigida):
    logEventos.objects.create(
        tipo_evento=tipo_evento,
        metodo=metodo,
        usuario_evento=usuario_evento,
        jefatura_dirigida=jefatura_dirigida,
    )
    return None

def ActividadesHistoria(id_periodo_seguimiento, id_actividad,fecha_reprogramacion_inicio, fecha_reprogramacion_termino, fecha_real_inicio,fecha_real_termino, id_estado_actividad,id_periodo, justificacion, actualiza, id_controlador_id):


    if actualiza == 0:
        Ges_Actividad_Historia.objects.create(
            id_periodo_seguimiento=id_periodo_seguimiento,
            id_actividad=id_actividad,
            fecha_reprogramacion_inicio=fecha_reprogramacion_inicio,
            fecha_reprogramacion_termino=fecha_reprogramacion_termino,
            fecha_real_inicio=fecha_real_inicio, #sprint 2 - CI -10 - 180120201
            fecha_real_termino=fecha_real_termino,
            id_estado_actividad=id_estado_actividad,
            id_periodo=id_periodo,
            justificacion=justificacion,

        )
        return None
    else:
        Id = Ges_Actividad_Historia.objects.get(Q(id_actividad=id_actividad) & Q(id_periodo_seguimiento=id_periodo_seguimiento) & Q(id_periodo=id_periodo) & Q(id_controlador=id_controlador_id))

        Ges_Actividad_Historia.objects.filter(id=Id.id).update(
            fecha_reprogramacion_inicio=fecha_reprogramacion_inicio,
            fecha_reprogramacion_termino=fecha_reprogramacion_termino,
            fecha_real_inicio=fecha_real_inicio, #sprint 2 - CI -10 - 180120201
            fecha_real_termino=fecha_real_termino,
            id_estado_actividad=id_estado_actividad,
            justificacion=justificacion,


        )
        return None




def CopiarActividades_a_Historial(id_periodo_seguimiento, id_controlador, periodo_actual):

    no_vacio_periodo= Ges_Actividad_Historia.objects.filter(Q(id_controlador=id_controlador) & Q(id_periodo_seguimiento=id_periodo_seguimiento) & Q(id_periodo=periodo_actual))

    if no_vacio_periodo:

        Ges_Actividad_Historia.objects.filter(
            Q(id_controlador=id_controlador) & Q(id_periodo_seguimiento=id_periodo_seguimiento) & Q(
                id_periodo=periodo_actual)).delete()

        model = Ges_Actividad.objects.filter(Q(id_controlador=id_controlador) & Q(id_periodo=periodo_actual))
        for actividad in model:
            Ges_Actividad_Historia.objects.create(
                id_actividad_id=actividad.id,
                id_estado_actividad=actividad.id_estado_actividad,
                id_periodo_seguimiento=id_periodo_seguimiento,
                id_controlador=actividad.id_controlador,
                id_periodo=actividad.id_periodo,
                validada=actividad.validada,

            )

    else:
        model = Ges_Actividad.objects.filter(Q(id_controlador=id_controlador) & Q(id_periodo=periodo_actual))
        for actividad in model:
            Ges_Actividad_Historia.objects.create(
                id_actividad_id=actividad.id,
                id_estado_actividad=actividad.id_estado_actividad,
                id_periodo_seguimiento=id_periodo_seguimiento,
                id_controlador=actividad.id_controlador,
                id_periodo=actividad.id_periodo,
                validada=actividad.validada,

            )

def UpdateFlag(id_controlador,periodo_actual):


    Ges_Actividad.objects.filter(Q(id_controlador=id_controlador) & Q(id_periodo=periodo_actual) & (~Q(id_estado_actividad=7) & ~Q(id_estado_actividad=9))).update(flag_reporta=0)
   # Ges_Actividad.objects.filter(Q(id_controlador=id_controlador) & Q(id_periodo=periodo_actual) & (~Q(fecha_real_inicio__isnull=True) | ~Q(id_estado_actividad=2) & ~Q(id_estado_actividad=3) & (~Q(id_estado_actividad=5) & (~Q(fecha_reprogramacion_inicio__isnull=True) & ~Q(fecha_reprogramacion_termino__isnull=True)))  & ~Q(id_estado_actividad=10) & ~Q(id_estado_actividad=4))).update(flag_tmp=1) # Sprint 1 - CI-2 - 11012021
    Ges_Actividad.objects.filter(Q(id_controlador=id_controlador) & Q(id_periodo=periodo_actual) & (Q(id_estado_actividad=1) | Q(id_estado_actividad=6) | Q(id_estado_actividad=7) | Q(id_estado_actividad=8) | Q(id_estado_actividad=9))).update(flag_tmp=1)

    Ges_Actividad.objects.filter(Q(id_controlador=id_controlador) & Q(id_periodo=periodo_actual) & (
                Q(id_estado_actividad=7)  | Q(id_estado_actividad=9) | (Q(id_estado_actividad=5) & (Q(id_periodicidad=9))) )).update(flag_finalizada=1)

def UpdateFlagCierre(id_controlador,periodo_actual):
    Ges_Actividad.objects.filter(Q(id_controlador=id_controlador) & Q(id_periodo=periodo_actual) & (
                ~Q(id_estado_actividad=7) | ~Q(id_estado_actividad=9))).update(validada=0)
    Ges_Actividad.objects.filter(Q(id_controlador=id_controlador) & Q(id_periodo=periodo_actual) & (
                Q(id_estado_actividad=3) & (~Q(fecha_real_inicio__isnull=True) & Q(fecha_reprogramacion_termino__isnull=True)))).update(flag_tmp=1)


    #Ges_Actividad.objects.filter(Q(id_controlador=id_controlador) & Q(id_periodo=periodo_actual)).update(validada=0)
def EnviarCorreoInicioSeguimiento(emails_destino_analista, email_jefatura,area_plan):

    ahora = datetime.now()
    fecha = ahora.strftime("%d" + "/" + "%m" + "/" + "%Y" + " a las " + "%H:%M")
    area_plan = str(area_plan)
    idcorreoJefatura=emails_destino_analista + [email_jefatura]

    subject = 'Inicio Etapa de Seguimiento'
    messageHtml = '<b>Estimada(o) Usuaria(o) del Sistema Capacity Institucional</b>, <br> Le informamos que con fecha <b>'+ str(fecha) +'</b> , el proceso de seguimiento del área: <b>' + area_plan + '</b> fue abierto. <br><p style="font-size:12px;color:red;">correo generado automaticamente favor no responder.'

    email = EmailMessage(subject, messageHtml ,to=idcorreoJefatura)
    email.content_subtype='html'
    email.send()


def EnviarCorreoCierreSeguimiento(emails_destino_analista, email_jefatura, area_plan):

    ahora = datetime.now()
    fecha = ahora.strftime("%d" + "/" + "%m" + "/" + "%Y" + " a las " + "%H:%M")

    area_plan = str(area_plan)
    idcorreoJefatura=emails_destino_analista + [email_jefatura]

    subject = 'Cierre Etapa de Seguimiento'
    messageHtml = '<b>Estimada(o) Usuaria(o) del Sistema Capacity Institucional</b>, <br> Le informamos que con fecha <b>'+ str(fecha) +'</b> , el proceso de seguimiento del área: <b>' + area_plan + '</b> fue cerrado. <br><p style="font-size:12px;color:red;">correo generado automaticamente favor no responder.'

    email = EmailMessage(subject, messageHtml ,to=idcorreoJefatura)
    email.content_subtype='html'
    email.send()

def export_users_xls_seguimiento(request, *args, **kwargs):
    try:
        periodo_actual = Glo_Periodos.objects.get(id_estado=1)
    except Glo_Periodos.DoesNotExist:
        return None

    id_controlador = kwargs['pk']


    nivel=Ges_Controlador.objects.get(Q(id=id_controlador) & Q(id_periodo=periodo_actual))

    nivel= nivel.nivel_inicial


    actividades = Ges_Actividad.objects.filter(Q(id_controlador=id_controlador) &
                                                            Q(id_periodo=periodo_actual))

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-Plan_de_Gestion.xlsx'.format(
        date=datetime.now().strftime('%d/%m/%Y'),
    )
    workbook = Workbook()


    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'reporte_seguimiento'

    # Define the titles for columns

    columns = ['Unidad',
               'Id Actividad',
               'Actividad',
               'Objetivo Vinculado',
               'Periodicidad',
               'Producto Estadístico',
               'Hora x Actividad',
               'Volumen',
               'N° Personas Asignadas',
               'Total Horas',
               'Cargo',
               'Fecha Incio Actividad',
               'Fecha Término Actividad',
               'Estado Actividad',
               'Fecha Real Inicio',
               'Fecha Real Finalización',
               'Reprogramación Fecha Inicio',
               'Reprogramación Fecha Término',
               'Justificación Desviación',
               'Respuesta Jefatura'

               ]

    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    for actividad in actividades:
        row_num += 1

        row =''

        if nivel==4:

            row = [
                actividad.id_controlador.id_jefatura.id_nivel.descripcion_nivel,
                actividad.id,
                actividad.descripcion_actividad,
                str(actividad.id_objetivo_operativo),
                str(actividad.id_periodicidad),
                str(actividad.id_producto_estadistico),
                actividad.horas_actividad,
                actividad.volumen,
                actividad.personas_asignadas,
                actividad.total_horas,
                str(actividad.id_familia_cargo),
                actividad.fecha_inicio_actividad,
                actividad.fecha_termino_actividad,
                str(actividad.id_estado_actividad),
                actividad.fecha_real_inicio,
                actividad.fecha_real_termino,
                actividad.fecha_reprogramacion_inicio,
                actividad.fecha_reprogramacion_termino,
                actividad.justificacion,
                actividad.validada,


            ]

        if nivel==3:

            row = [
                actividad.id_controlador.id_jefatura.id_nivel.descripcion_nivel,
                actividad.id,
                actividad.descripcion_actividad,
                str(actividad.id_objetivo_tacticotn) ,
                str(actividad.id_periodicidad),
                str(actividad.id_producto_estadistico),
                actividad.horas_actividad,
                actividad.volumen,
                actividad.personas_asignadas,
                actividad.total_horas,
                str(actividad.id_familia_cargo),
                actividad.fecha_inicio_actividad,
                actividad.fecha_termino_actividad,
                str(actividad.id_estado_actividad),
                actividad.fecha_real_inicio,
                actividad.fecha_real_termino,
                actividad.fecha_reprogramacion_inicio,
                actividad.fecha_reprogramacion_termino,
                actividad.justificacion,
                actividad.validada,



            ]



        if nivel==2:

            row = [
                actividad.id_controlador.id_jefatura.id_nivel.descripcion_nivel,
                actividad.id,
                actividad.descripcion_actividad,
                str(actividad.id_objetivo_tactico) ,
                str(actividad.id_periodicidad),
                str(actividad.id_producto_estadistico),
                actividad.horas_actividad,
                actividad.volumen,
                actividad.personas_asignadas,
                actividad.total_horas,
                str(actividad.id_familia_cargo),
                actividad.fecha_inicio_actividad,
                actividad.fecha_termino_actividad,
                str(actividad.id_estado_actividad),
                actividad.fecha_real_inicio,
                actividad.fecha_real_termino,
                actividad.fecha_reprogramacion_inicio,
                actividad.fecha_reprogramacion_termino,
                actividad.justificacion,
                actividad.validada,

            ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

            if col_num == 12:
                cell.number_format = 'dd/mm/yyyy'
            if col_num == 13:
                cell.number_format = 'dd/mm/yyyy'
            if col_num == 15:
                cell.number_format = 'dd/mm/yyyy'
            if col_num == 16:
                cell.number_format = 'dd/mm/yyyy'
            if col_num == 17:
                cell.number_format = 'dd/mm/yyyy'
            if col_num == 18:
                cell.number_format = 'dd/mm/yyyy'

            if col_num == 20:
                if cell.value == 0:
                    cell.value='No Reportado'
                if cell.value == 1:
                    cell.value='Aceptado'
                if cell.value == 2:
                    cell.value='Rechazado'


    workbook.save(response)


    return response


def export_users_xls_seguimiento_comentarios(request, *args, **kwargs):
    try:
        periodo_actual = Glo_Periodos.objects.get(id_estado=1)
    except Glo_Periodos.DoesNotExist:
        return None

    id_controlador = kwargs['pk']


    nivel=Ges_Controlador.objects.get(Q(id=id_controlador) & Q(id_periodo=periodo_actual))

    nivel= nivel.nivel_inicial


    observaciones = Ges_Observaciones_valida.objects.filter(Q(id_controlador=id_controlador) &
                                                            Q(id_periodo=periodo_actual))

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-Plan_de_Gestion.xlsx'.format(
        date=datetime.now().strftime('%d/%m/%Y'),
    )
    workbook = Workbook()


    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'reporte_seguimiento'

    # Define the titles for columns

    columns = ['Id Actividad',
               'Actividad',
               'Observación',
               'Fecha Registro',
               'Periodo',
               'Periodo Validación',
               'Jefatura',
               'Area'
               ]

    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    for observacion in observaciones:
        row_num += 1

        row =''

        row = [str(observacion.id_actividad.id),
               observacion.id_actividad.descripcion_actividad,
               observacion.descripcion_observacion,
               observacion.fecha_registro,
               observacion.id_periodo.descripcion_periodo,
               observacion.id_periodo_valida.descripcion_validacion,
               observacion.jefatura_primerarevision.id_user.username,
               str(observacion.id_controlador.id_jefatura.id_nivel.descripcion_nivel),

               ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

            if col_num == 12:
                cell.number_format = 'dd/mm/yyyy'
            if col_num == 13:
                cell.number_format = 'dd/mm/yyyy'
            if col_num == 15:
                cell.number_format = 'dd/mm/yyyy'
            if col_num == 16:
                cell.number_format = 'dd/mm/yyyy'
            if col_num == 17:
                cell.number_format = 'dd/mm/yyyy'
            if col_num == 18:
                cell.number_format = 'dd/mm/yyyy'

            if col_num == 20:
                if cell.value == 0:
                    cell.value='No Reportado'
                if cell.value == 1:
                    cell.value='Aceptado'
                if cell.value == 2:
                    cell.value='Rechazado'


    workbook.save(response)


    return response