from django.shortcuts import render
from django.views.generic import ListView, CreateView, UpdateView, DeleteView
# Create your views here.
from apps.actividades.models import Ges_Actividad, Ges_Observaciones_valida, Ges_log_reportes, Ges_Actividad_Historia
from apps.controlador.models import Ges_Controlador
from apps.estado_actividad.models import Glo_EstadoActividad
from apps.estructura.models import Ges_Niveles
from apps.jefaturas.models import Ges_Jefatura
from apps.objetivos.models import Ges_Objetivo_TacticoTN, Ges_Objetivo_Tactico, Ges_Objetivo_Operativo
from apps.periodos.models import Glo_Periodos, Glo_Seguimiento, Glo_validacion
from apps.valida_plan.models import Ges_Observaciones
from django.db.models import Subquery, OuterRef, Count
from django.db.models import Q
from django.http import HttpResponseRedirect, JsonResponse
from django.contrib import messages
from django.contrib.messages.views import SuccessMessageMixin
from apps.valida_plan2.models import Ges_Observaciones_sr
from apps.valida_seguimiento.forms import  ValidaActividadesUpdateForm,PlanUpdateForm, ValidaSeguimientoUpdateFormVer
from django.http import HttpResponseRedirect, JsonResponse
from openpyxl import Workbook
from django.http import HttpResponse
from datetime import datetime


class UnidadesList(ListView): #Modificado por JR- sprint 8 - OK
    model = Ges_Niveles
    template_name = 'valida_seguimiento/valida_seguimiento_list.html'

    def get_context_data(self, **kwargs):
        context = super(UnidadesList, self).get_context_data(**kwargs)
        id_usuario_actual = self.request.user.id  # obtiene id usuario actual
        id_jefatura = Ges_Jefatura.objects.get(id_user=id_usuario_actual)
        try:
            periodo_actual = Glo_Periodos.objects.get(id_estado=1)
        except Glo_Periodos.DoesNotExist:
            return None

        try:
            periodo_valida = Glo_validacion.objects.get(
                Q(id_estado_periodo=1) & Q(id_periodo=periodo_actual.id))
        except:
            periodo_valida = None
            pass




        if periodo_valida:
            try:

                id_controlador = Ges_Controlador.objects.filter(
                    Q(jefatura_primerarevision=id_jefatura) & Q(id_periodo=periodo_actual.id)
                    & Q(id_estado_plan=3))

            except Ges_Controlador.DoesNotExist:
                id_controlador = 0
                pass

            if id_controlador == 0:
                context['error'] = {'id': 1}
                return context

            else:
                try:
                    id_jefatura = Ges_Jefatura.objects.get(id_user=id_usuario_actual)

                    count_no_vistos = Ges_Observaciones.objects.values('id_controlador').filter(
                        id_controlador=OuterRef('pk')).annotate(
                        count_id_actividad=Count('id', filter=Q(observado=1)  & Q(id_periodo=periodo_actual.id) & (~Q(user_observa=id_usuario_actual))))

                    count_observaciones = Ges_Observaciones.objects.values('id_controlador').filter(
                        id_controlador=OuterRef('pk')).annotate(
                        count_id_actividad=Count('id'))

                    count_actividades = Ges_Actividad.objects.values('id_controlador').filter(Q(
                        id_controlador=OuterRef('pk')) & Q(id_estado_actividad__in=[2,3,5,10])).annotate(
                        count_id_actividad_valida=Count('id'))

                    id_controlador = Ges_Controlador.objects.filter(
                        Q(jefatura_primerarevision=id_jefatura.id) & Q(id_periodo=periodo_actual.id) & Q(id_estado_plan=3)).annotate(
                        count_no_vistos=Subquery(count_no_vistos.values('count_id_actividad')),
                        count_actividad=Count(Subquery(count_actividades.values('count_id_actividad_valida'))),
                        count_observaciones=Subquery(count_observaciones.values('count_id_actividad'))).order_by(
                        '-count_no_vistos', '-count_observaciones')

                    id_controladorfiltrado = id_controlador.filter(count_actividad__gte=1)

                    context['jefatura_primerarevision'] = {'id': id_jefatura.id}

                    #id_jefatura = Ges_Jefatura.objects.filter(id_user=id_usuario_actual)

                    context['object_list'] = id_controladorfiltrado

                   # context['object_list3'] = id_jefatura

                    return context

                except Ges_Jefatura.DoesNotExist:
                    context['habilitado'] = {'mensaje': False}
                    return None
        else:
            self.request.session['message_class'] = "alert alert-warning"
            messages.error(self.request,
                           "Estimado usuario no existe un periodo de seguimiento activo. Favor comuniquese con el administrador")

            context['estado_seguimiento'] = 0
            return context

class Objetivos(ListView): #Modificado por JR- sprint 8 - OK
    model = Ges_Actividad
    template_name = 'valida_seguimiento/valida_seguimiento_detalle.html'

    def get_context_data(self, **kwargs):
        context = super(Objetivos, self).get_context_data(**kwargs)
        id_usuario_actual = self.request.user.id  # obtiene id usuario actual
        self.request.session['id_nivel_controlador'] = self.kwargs['pk'] #guarda id  controlador

        try:
            periodo_actual = Glo_Periodos.objects.get(id_estado=1)
        except Glo_Periodos.DoesNotExist:
            return None

        try:
            controlador = Ges_Controlador.objects.get(Q(id=self.kwargs['pk']) & Q(id_periodo=periodo_actual.id))
        except Ges_Controlador.DoesNotExist:
            return None

        id_orden = controlador.nivel_inicial
        self.request.session['id_controlador_real'] = controlador.id
       # vaar= controlador.id_jefatura.id_nivel_id

        id_nivel = Ges_Niveles.objects.get(
            Q(id=controlador.id_jefatura.id_nivel_id) & Q(id_periodo=periodo_actual.id))


        try:
           # id_jefatura = Ges_Jefatura.objects.get(id_user=id_usuario_actual)
           # id_controlador = Ges_Controlador.objects.get(Q(jefatura_primerarevision=id_jefatura.id) & Q(id_periodo=periodo_actual.id) & Q(id=controlador.id)) #que pasa cuando hay más de una jefatura?

            if id_orden == 3:
               # replies2 = Ges_Objetivo_TacticoTN.objects.filter(id_tercer_nivel_id=id_nivel.id_tercer_nivel_id)
                count_no_vistos = Ges_Observaciones.objects.values('id_objetivo').filter(
                    id_objetivo=OuterRef('pk')).annotate(
                    count_id_actividad=Count('id', filter=Q(observado=1)  & Q(id_periodo=periodo_actual.id) & (~Q(user_observa=id_usuario_actual))))

                count_observaciones = Ges_Observaciones.objects.values('id_objetivo').filter(
                    id_objetivo=OuterRef('pk')).annotate(
                    count_id_actividad=Count('id'))

                count_actividades = Ges_Actividad.objects.values('id_objetivo_tacticotn').filter(
                    id_objetivo_tacticotn=OuterRef('pk')).annotate(
                    count_id_actividad=Count('id', filter=Q(id_periodo=periodo_actual.id) & (Q(id_estado_actividad=3) & ~Q(fecha_reprogramacion_termino=None)) | (Q(id_estado_actividad=5) & ~Q(fecha_reprogramacion_inicio=None) & ~Q(fecha_reprogramacion_termino=None)) | (Q(id_estado_actividad=2)  | Q(id_estado_actividad=10))))

                count_no_vistos_obj = Ges_Observaciones_sr.objects.values('id_objetivo_tacticotn').filter(
                    id_objetivo_tacticotn=OuterRef('pk')).annotate(
                    count_id_actividad=Count('id', filter=Q(observado=1) & Q(id_periodo=periodo_actual.id) & (
                        ~Q(user_observa=id_usuario_actual))))

                replies2 = Ges_Objetivo_TacticoTN.objects.filter(
                    Q(id_tercer_nivel_id=id_nivel.id_tercer_nivel_id) & Q(id_periodo=periodo_actual.id)).annotate(
                    count_no_vistos=Subquery(count_no_vistos.values('count_id_actividad')),
                    count_actividades=Subquery(count_actividades.values('count_id_actividad')),
                    count_no_vistos_obj=Subquery(count_no_vistos_obj.values('count_id_actividad')),
                    count_observaciones=Subquery(count_observaciones.values('count_id_actividad'))).order_by(
                    '-count_actividades', '-count_observaciones')


            if id_orden == 2:
               # replies2 = Ges_Objetivo_Tactico.objects.filter(id_segundo_nivel_id=id_nivel.id_segundo_nivel_id)
                count_no_vistos = Ges_Observaciones.objects.values('id_objetivo').filter(
                    id_objetivo=OuterRef('pk')).annotate(
                    count_id_actividad=Count('id', filter=Q(observado=1)  & Q(id_periodo=periodo_actual.id) & (~Q(user_observa=id_usuario_actual))))

                count_observaciones = Ges_Observaciones.objects.values('id_objetivo').filter(
                    id_objetivo=OuterRef('pk')).annotate(
                    count_id_actividad=Count('id'))

                count_actividades = Ges_Actividad.objects.values('id_objetivo_tactico').filter(
                    id_objetivo_tactico=OuterRef('pk')).annotate(
                    count_id_actividad=Count('id', filter=Q(id_periodo=periodo_actual.id) & (Q(id_estado_actividad=3) & ~Q(fecha_reprogramacion_termino=None)) | (Q(id_estado_actividad=5) & ~Q(fecha_reprogramacion_inicio=None) & ~Q(fecha_reprogramacion_termino=None)) | (Q(id_estado_actividad=2)  | Q(id_estado_actividad=10))))

                count_no_vistos_obj = Ges_Observaciones_sr.objects.values('id_objetivo_tactico').filter(
                    id_objetivo_tactico=OuterRef('pk')).annotate(
                    count_id_actividad=Count('id', filter=Q(observado=1) & Q(id_periodo=periodo_actual.id) & (
                       ~Q(user_observa=id_usuario_actual))))

                replies2 = Ges_Objetivo_Tactico.objects.filter(
                    Q(id_segundo_nivel_id=id_nivel.id_segundo_nivel_id) & Q(id_periodo=periodo_actual.id)).annotate(
                    count_no_vistos=Subquery(count_no_vistos.values('count_id_actividad')),
                    count_actividades=Subquery(count_actividades.values('count_id_actividad')),
                    count_no_vistos_obj=Subquery(count_no_vistos_obj.values('count_id_actividad')),
                    count_observaciones=Subquery(count_observaciones.values('count_id_actividad'))).order_by(
                    '-count_actividades', '-count_observaciones')


            if id_orden == 4:
               # replies2 = Ges_Objetivo_Operativo.objects.filter(id_cuarto_nivel_id=id_nivel.id_cuarto_nivel_id)

                count_no_vistos = Ges_Observaciones.objects.values('id_objetivo').filter(
                    id_objetivo=OuterRef('pk')).annotate(
                    count_id_actividad=Count('id', filter=Q(observado=1)  & Q(id_periodo=periodo_actual.id) & (~Q(user_observa=id_usuario_actual))))

                count_observaciones = Ges_Observaciones.objects.values('id_objetivo').filter(
                    id_objetivo=OuterRef('pk')).annotate(
                    count_id_actividad=Count('id'))

                count_actividades = Ges_Actividad.objects.values('id_objetivo_operativo').filter(
                    id_objetivo_operativo=OuterRef('pk')).annotate(
                    count_id_actividad=Count('id', filter=Q(id_periodo=periodo_actual.id) & (Q(id_estado_actividad=3) & ~Q(fecha_reprogramacion_termino=None)) | (Q(id_estado_actividad=5) & ~Q(fecha_reprogramacion_inicio=None) & ~Q(fecha_reprogramacion_termino=None)) | (Q(id_estado_actividad=2)  | Q(id_estado_actividad=10))))

                count_no_vistos_obj = Ges_Observaciones_sr.objects.values('id_objetivo_operativo').filter(
                    id_objetivo_operativo=OuterRef('pk')).annotate(
                    count_id_actividad=Count('id', filter=Q(observado=1) & Q(id_periodo=periodo_actual.id) & (
                        ~Q(user_observa=id_usuario_actual))))

                replies2 = Ges_Objetivo_Operativo.objects.filter(
                    Q(id_cuarto_nivel_id=id_nivel.id_cuarto_nivel_id) & Q(id_periodo=periodo_actual.id)).annotate(
                    count_no_vistos=Subquery(count_no_vistos.values('count_id_actividad')),
                    count_actividades=Subquery(count_actividades.values('count_id_actividad')),
                    count_no_vistos_obj=Subquery(count_no_vistos_obj.values('count_id_actividad')),
                    count_observaciones=Subquery(count_observaciones.values('count_id_actividad'))).order_by(
                    '-count_actividades', '-count_observaciones')



            context['orden'] = {'orden_nivel': id_orden}
            context['niveles'] = replies2
            context['objetivo'] = id_nivel


            self.request.session['id_orden'] = id_orden


            return context

        except Ges_Jefatura.DoesNotExist:
            context['habilitado'] = {'mensaje': False}


            return None

class Actividades(ListView): # Modificado por JR- sprint 8 - OK
    model = Ges_Actividad
    template_name = 'valida_seguimiento/valida_seguimiento_list_actividades.html'

    def get_context_data(self,  **kwargs):
        context = super(Actividades, self).get_context_data(**kwargs)
        id_usuario_actual = self.request.user.id  # obtiene id usuario actual

        try:
            periodo_actual = Glo_Periodos.objects.get(id_estado=1)
        except Glo_Periodos.DoesNotExist:
            return None

        self.request.session['id_objetivo']=self.kwargs['pk']#315

        nombre = ""
        if self.request.session['id_orden']==2:
           # lista_actividades = Ges_Actividad.objects.filter(Q(id_objetivo_tactico=self.kwargs['pk']) & Q(id_periodo=periodo_actual.id))

            count_no_vistos = Ges_Observaciones.objects.values('id_actividad_id').filter(id_actividad=OuterRef('pk')).annotate(
              count_id_actividad=Count('id', filter=Q(observado=1) & Q(id_periodo=periodo_actual.id) & (~Q(user_observa=id_usuario_actual))))

            count_observaciones = Ges_Observaciones.objects.values('id_actividad_id').filter(id_actividad=OuterRef('pk')).annotate(
            count_id_actividad=Count('id'))

            lista_actividades = Ges_Actividad.objects.filter(
              Q(id_objetivo_tactico=self.kwargs['pk']) & Q(id_periodo=periodo_actual.id) & Q(id_estado_actividad__in=[1,2,3,4,5,6,8,9,10])  ).annotate(
              count_no_vistos=Subquery(count_no_vistos.values('count_id_actividad')),

                count_observaciones=Subquery(count_observaciones.values('count_id_actividad'))).order_by('-count_no_vistos','-fecha_registro')

            nombre=  Ges_Objetivo_Tactico.objects.get(id=self.kwargs['pk'])

        if self.request.session['id_orden']==3:
           # lista_actividades = Ges_Actividad.objects.filter(Q(id_objetivo_tacticotn=self.kwargs['pk']) & Q(id_periodo=periodo_actual.id))

            count_no_vistos = Ges_Observaciones.objects.values('id_actividad_id').filter(id_actividad=OuterRef('pk')).annotate(
              count_id_actividad=Count('id', filter=Q(observado=1) & Q(id_periodo=periodo_actual.id) & (~Q(user_observa=id_usuario_actual))))
            count_observaciones = Ges_Observaciones.objects.values('id_actividad_id').filter(id_actividad=OuterRef('pk')).annotate(
            count_id_actividad=Count('id'))


            lista_actividades = Ges_Actividad.objects.filter(
              Q(id_objetivo_tacticotn=self.kwargs['pk']) & Q(id_periodo=periodo_actual.id)  & Q(id_estado_actividad__in=[1,2,3,4,5,6,8,9,10]) ).annotate(
              count_no_vistos=Subquery(count_no_vistos.values('count_id_actividad')),


                count_observaciones=Subquery(count_observaciones.values('count_id_actividad'))).order_by('-count_no_vistos','-fecha_registro')


            nombre = Ges_Objetivo_TacticoTN.objects.get(id=self.kwargs['pk'])

        if self.request.session['id_orden']==4:

            count_no_vistos = Ges_Observaciones.objects.values('id_actividad_id').filter(id_actividad=OuterRef('pk')).annotate(
              count_id_actividad=Count('id', filter=Q(observado=1) & Q(id_periodo=periodo_actual.id) & (~Q(user_observa=id_usuario_actual))))
            count_observaciones = Ges_Observaciones.objects.values('id_actividad_id').filter(id_actividad=OuterRef('pk')).annotate(
            count_id_actividad=Count('id'))

            lista_actividades = Ges_Actividad.objects.filter(
              Q(id_objetivo_operativo=self.kwargs['pk']) & Q(id_periodo=periodo_actual.id) & Q(id_estado_actividad__in=[1,2,3,4,5,6,8,9,10]) ).annotate(
              count_no_vistos=Subquery(count_no_vistos.values('count_id_actividad')),

                count_observaciones=Subquery(count_observaciones.values('count_id_actividad'))).order_by('-count_no_vistos','-fecha_registro')


            nombre = Ges_Objetivo_Operativo.objects.get(id=self.kwargs['pk'])

        try:
            nivel_usuario = Ges_Jefatura.objects.get(Q(id_user=id_usuario_actual) & Q(id_periodo=periodo_actual.id))
        except Ges_Jefatura.DoesNotExist:
            context['habilitado'] = {'mensaje': False}
            return None


        context['object_list'] = lista_actividades
        context['nombre_objetivo'] = {'nombre': nombre}
        context['id_orden'] = {'id_orden': self.request.session['id_orden']}
        context['id_nivel_controlador']={'id': self.request.session['id_nivel_controlador']}

        #context['objetivo'] = id_nivel


        return context

class ActividadEdit(SuccessMessageMixin, UpdateView ):
    model = Ges_Actividad
    form_class = ValidaActividadesUpdateForm
    template_name = 'valida_seguimiento/validactividades_seguimiento_update.html'

    def get_form_kwargs(self, **kwargs):
        kwargs = super(ActividadEdit, self).get_form_kwargs()

        try:
            estado_actividad = Ges_Actividad.objects.get(id=self.kwargs['pk'])
        except Glo_Periodos.DoesNotExist:
            return None
        kwargs['id_estado_actual'] = estado_actividad.id_estado_actividad_id

        return kwargs

    def get_context_data(self,  **kwargs):
        context = super(ActividadEdit, self).get_context_data(**kwargs)



        try:
            periodo_actual = Glo_Periodos.objects.get(id_estado=1)
        except Glo_Periodos.DoesNotExist:
            return None
        try:
            actividad = Ges_Actividad.objects.get(id=self.kwargs['pk'])
        except Glo_Periodos.DoesNotExist:
            return None
        try:
            periodo_validacion = Glo_validacion.objects.get(
                Q(id_estado_periodo=1) & Q(id_periodo=periodo_actual.id))
        except Glo_Periodos.DoesNotExist:
            return None
        fechas_corte= Glo_Seguimiento.objects.order_by('-id')[0]

        context['fechas'] = {'fecha_inicio_corte_str':str(fechas_corte.fecha_inicio_corte)  ,
                             'fecha_termino_corte_str':str(fechas_corte.fecha_termino_corte),
                             'fecha_inicio_corte': fechas_corte.fecha_inicio_corte,
                             'fecha_termino_corte': fechas_corte.fecha_termino_corte,
                             'id_actividad': actividad.id,
                             'estado_valida': actividad.validada,
                             'id_periodo_validacion': periodo_validacion.id,
                             'estado_actividad':actividad.id_estado_actividad_id,
                             }


        return context

    def post(self, request, *args, **kwargs):
        form_class = self.get_form_class()
        self.object = self.get_object()
        id_actividad = kwargs['pk']
        instancia_nivel = self.model.objects.get(id=id_actividad)
        form = self.form_class(request.POST, instance=instancia_nivel)
        id_usuario_actual = self.request.user.id  # obtiene id usuario actual

        try:
            periodo_actual = Glo_Periodos.objects.get(id_estado=1)
        except Glo_Periodos.DoesNotExist:
            return None

        try:
            id_jefatura = Ges_Jefatura.objects.get(Q(id_user=id_usuario_actual) & Q(id_periodo=periodo_actual.id))
        except Ges_Jefatura.DoesNotExist:
            return None

        try:
            usuario_controlador = Ges_Controlador.objects.get(id_jefatura=id_jefatura.id)
        except Ges_Controlador.DoesNotExist:
            return None
        fecha_real_termino = request.POST['fecha_real_termino']
        fecha_inicio_reprogramacion = request.POST['fecha_reprogramacion_inicio']
        fecha_termino_reprogramacion = request.POST['fecha_reprogramacion_termino']

        if fecha_real_termino == '':
            form.instance.fecha_real_termino = None
        if fecha_inicio_reprogramacion == '':
            form.instance.fecha_reprogramacion_inicio = None
        if fecha_termino_reprogramacion == '':
            form.instance.fecha_reprogramacion_termino = None

        if form.is_valid():

            if self.request.session['id_orden'] == 2:
                id_objetivo = Ges_Objetivo_Tactico.objects.get(id=self.request.session['id_objetivo'])
                form.instance.id_objetivo_tactico = id_objetivo

            if self.request.session['id_orden'] == 3:
                id_objetivo = Ges_Objetivo_TacticoTN.objects.get(id=self.request.session['id_objetivo'])
                form.instance.id_objetivo_tacticotn = id_objetivo

            if self.request.session['id_orden'] == 4:
                id_objetivo = Ges_Objetivo_Operativo.objects.get(id=self.request.session['id_objetivo'])
                form.instance.id_objetivo_operativo = id_objetivo

            form.save()
            request.session['message_class'] = "alert alert-success"
            messages.success(self.request, "Los datos fueron actualizados correctamente!")
            return HttpResponseRedirect('/seguimiento_formula/detalle/' + str(self.request.session['id_objetivo']))
        else:
            request.session['message_class'] = "alert alert-danger"
            messages.error(self.request,
                           "Error interno: No se ha creado el registro. Comuníquese con el administrador.")
            return HttpResponseRedirect('/seguimiento_formula/detalle/' + str(self.request.session['id_objetivo']))


def update_actividad(request):
    id_actividad = request.POST.get('id')
    response_data = {}
    if request.POST.get('action') == 'post':
        #Valores a recibir del form
        id_actividad = request.POST.get('id_actividad')
        id_periodo_validacion = request.POST.get('id_periodo_validacion')
        estado_validacion = request.POST.get('estado_validacion')
        fecha_inicio= request.POST.get('fecha_inicio')
        fecha_termino= request.POST.get('fecha_termino')
        estado_final= request.POST.get('estado_final')
        observacion= request.POST.get('observacion')


        #buscar el modelo a actualizar
        actividad = Ges_Actividad.objects.get(id=id_actividad)
        actividad.validada  = estado_validacion
        actividad.observacion_valida = observacion
        actividad.id_estado_actividad_id = int(estado_final)

        if (fecha_inicio != ''):
            actividad.fecha_inicio_actividad = fecha_inicio
            actividad.fecha_reprogramacion_inicio = None  # Sprint 1 - CI-18 - 20012021 nuevo
            actividad.fecha_reprogramacion_termino = None  # Sprint 1 - CI-18 - 20012021 nuevo

        if (fecha_termino != ''):
            actividad.fecha_termino_actividad = fecha_termino
            actividad.fecha_reprogramacion_inicio = None  # Sprint 2 - CI-18 - 20012021 nuevo
            actividad.fecha_reprogramacion_termino = None  # Sprint 2 - CI-18- 20012021 nuevo

        if(fecha_termino != ''):
            actividad.fecha_termino_actividad = fecha_termino

        try:
          actividad.save()
          response_data['error'] = 'La actividad fue validada correctamente.'
        except:
          response_data['error'] = 'Error al intentar validar la actividad, intente nuevamente o comuniquese con el administrador.'
        #
        try:
            periodo_actual = Glo_Periodos.objects.get(id_estado=1)
        except Glo_Periodos.DoesNotExist:
            return None
        try:
            id_periodo_seguimiento = Glo_Seguimiento.objects.filter(id_estado_seguimiento=2).order_by('-id')[0]
        except Glo_Seguimiento.DoesNotExist:
            return None
        #
        try:
            id_estado_final_instancia = Glo_EstadoActividad.objects.get(id=int(estado_final))
        except id_estado_final_instancia.DoesNotExist:
            return None
        #
        try:
            periodo_valida = Glo_validacion.objects.get(
                Q(id_estado_periodo=1) & Q(id_periodo=periodo_actual.id))
        except:
            periodo_valida = None
            pass


        agregarObservacion(observacion, int(id_actividad), int(periodo_valida.id) )
        Log_valida(int(periodo_valida.id),  int(id_actividad),  fecha_inicio, fecha_termino, None, int(estado_final))
        #
        ActividadesHistoria(id_periodo_seguimiento, id_actividad,  periodo_actual)

        return JsonResponse(response_data)

    return render(request, 'valida_seguimiento/listarActividades/' + str(id_actividad), {'success': 'true'})

class ValidaSeguimientoActividadDetallesVer(SuccessMessageMixin, UpdateView):
    model = Ges_Actividad
    form_class = ValidaSeguimientoUpdateFormVer
    template_name = 'valida_seguimiento/valida_seguimiento_ver_detalle.html'


def update_actividad_rechaza(request):
    id_actividad = request.POST.get('id')
    response_data = {}
    if request.POST.get('action') == 'post':
        #Valores a recibir del form
        id_actividad = request.POST.get('id_actividad')
        id_periodo_validacion = request.POST.get('id_periodo_validacion')
        estado_validacion = request.POST.get('estado_validacion')
        estado_final= request.POST.get('estado_final')
        observacion= request.POST.get('observacion')


        #buscar el modelo a actualizar
        actividad = Ges_Actividad.objects.get(id=id_actividad)
        actividad.validada  = estado_validacion
        actividad.observacion_valida = observacion
        actividad.id_estado_actividad_id = estado_final

        actividad.fecha_reprogramacion_inicio = None
        actividad.fecha_reprogramacion_termino = None

        try:
          actividad.save()
          response_data['error'] = 'La actividad fue rechazada correctamente.'
        except:
          response_data['error'] = 'Error al intentar validar la actividad, intente nuevamente o comuniquese con el administrador.'



        try:
            periodo_actual = Glo_Periodos.objects.get(id_estado=1)
        except Glo_Periodos.DoesNotExist:
            return None


        try:
            id_periodo_seguimiento = Glo_Seguimiento.objects.filter(id_estado_seguimiento=2).order_by('-id')[0]
        except Glo_Seguimiento.DoesNotExist:
            return None

        try:
            id_estado_final_instancia = Glo_EstadoActividad.objects.get(id=int(estado_final))
        except id_estado_final_instancia.DoesNotExist:
            return None

        try:
            periodo_valida = Glo_validacion.objects.get(
                Q(id_estado_periodo=1) & Q(id_periodo=periodo_actual.id))
        except:
            periodo_valida = None
            pass

        agregarObservacion(observacion, int(id_actividad), int(periodo_valida.id))
        Log_valida(int(periodo_valida.id), int(id_actividad),  None, None, None, int(estado_final))
        ActividadesHistoria(int(id_periodo_seguimiento.id), id_actividad, periodo_actual)


        return JsonResponse(response_data)

    return render(request, 'valida_seguimiento/listarActividades/' + str(id_actividad), {'success': 'true'})


def agregarObservacion(descripcion , id_actividad, id_valida):

    Ges_Observaciones_valida.objects.create(
        descripcion_observacion=descripcion,
        id_actividad_id=id_actividad,
        id_periodo_valida_id =id_valida,
    )
    return None

from datetime import date

def Log_valida(id_periodo_valida, id_actividad, fecha_inicio, fecha_termino, fecha_real_termino, id_estado_actividad ):

    if (fecha_inicio == ''):
        fecha_inicio = None
    if(fecha_termino== ''):
        fecha_termino = None
    if(fecha_real_termino == ''):
        fecha_real_termino = None

    Ges_log_reportes.objects.create(
    id_periodo_valida_id=id_periodo_valida,
    id_actividad_id= id_actividad,
    fecha_inicio = fecha_inicio,
    fecha_termino =fecha_termino,
    fecha_real_termino= fecha_real_termino,
    id_estado_actividad_id =id_estado_actividad,
    )
    return None


def ActividadesHistoria(id_periodo_seguimiento, id_actividad,id_periodo):

    Id = Ges_Actividad_Historia.objects.get(
        Q(id_actividad=id_actividad) & Q(id_periodo_seguimiento=id_periodo_seguimiento) & Q(id_periodo=id_periodo))

    model = Ges_Actividad.objects.filter(Q(id=id_actividad) & Q(id_periodo=id_periodo))

    for actividad in model:
        Ges_Actividad_Historia.objects.filter(id=Id.id).update(
            fecha_reprogramacion_inicio=actividad.fecha_reprogramacion_inicio,
            fecha_reprogramacion_termino=actividad.fecha_reprogramacion_termino,
            fecha_real_inicio=actividad.fecha_real_inicio,
            fecha_real_termino=actividad.fecha_real_termino,
            id_estado_actividad=actividad.id_estado_actividad,
            justificacion=actividad.justificacion,

        )



    return None


def export_users_xls_valida_seguimiento(request, *args, **kwargs):
    try:
        periodo_actual = Glo_Periodos.objects.get(id_estado=1)
    except Glo_Periodos.DoesNotExist:
        return None

    id_jefatura_primera = kwargs['pk']



    nivel=Ges_Controlador.objects.filter(Q(jefatura_primerarevision=id_jefatura_primera) & Q(id_periodo=periodo_actual)).order_by('-id')[0]


    nivel= nivel.nivel_inicial

    controladores = Ges_Controlador.objects.filter(Q(jefatura_primerarevision=id_jefatura_primera) & Q(id_periodo=periodo_actual))


    actividades = Ges_Actividad.objects.filter(Q(id_controlador__in=controladores) &
                                                            Q(id_periodo=periodo_actual) & Q(validada__in =[1,2]))

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-Plan_de_Gestion.xlsx'.format(
        date=datetime.now().strftime('%d/%m/%Y'),
    )
    workbook = Workbook()


    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'reporte_seguimiento'

    # Define the titles for columns

    columns = ['Unidad',
               'Objetivo Vinculado',
               'Actividad',

               'Periodicidad',
               'Producto Estadístico',
               'Hora x Actividad',
               'Volumen',
               'N° Personas Asignadas',
               'Total Horas',
               'Cargo',
               'Fecha Incio Actividad',
               'Fecha Término Actividad',
               'Estado Actividad',
               'Fecha Real Inicio',
               'Fecha Real Finalización',
               'Reprogramación Fecha Inicio',
               'Reprogramación Fecha Término',
               'Justificación Desviación',
               'Resultado Validación',
               'Fecha_Reg'



               ]

    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    for actividad in actividades:
        row_num += 1

        row =''

        if nivel==4:

            row = [
                str(actividad.id_objetivo_operativo.id_cuarto_nivel),
                str(actividad.id_objetivo_operativo),
                actividad.descripcion_actividad,
                str(actividad.id_periodicidad),
                str(actividad.id_producto_estadistico),
                actividad.horas_actividad,
                actividad.volumen,
                actividad.personas_asignadas,
                actividad.total_horas,
                str(actividad.id_familia_cargo),
                actividad.fecha_inicio_actividad,
                actividad.fecha_termino_actividad,
                str(actividad.id_estado_actividad),
                actividad.fecha_real_inicio,
                actividad.fecha_real_termino,
                actividad.fecha_reprogramacion_inicio,
                actividad.fecha_reprogramacion_termino,
                actividad.justificacion,
                actividad.fecha_registro,
                actividad.fecha_registro,



            ]

        if nivel==3:

            row = [
                str(actividad.id_objetivo_tacticotn.id_tercer_nivel),
                str(actividad.id_objetivo_tacticotn),
                actividad.descripcion_actividad,
                str(actividad.id_periodicidad),
                str(actividad.id_producto_estadistico),
                actividad.horas_actividad,
                actividad.volumen,
                actividad.personas_asignadas,
                actividad.total_horas,
                str(actividad.id_familia_cargo),
                actividad.fecha_inicio_actividad,
                actividad.fecha_termino_actividad,
                str(actividad.id_estado_actividad),
                actividad.fecha_real_inicio,
                actividad.fecha_real_termino,
                actividad.fecha_reprogramacion_inicio,
                actividad.fecha_reprogramacion_termino,
                actividad.justificacion,
                actividad.fecha_registro,



            ]



        if nivel==2:

            row = [
                str(actividad.id_objetivo_tactico.id_segundo_nivel),
                str(actividad.id_objetivo_tactico),
                actividad.descripcion_actividad,
                str(actividad.id_periodicidad),
                str(actividad.id_producto_estadistico),
                actividad.horas_actividad,
                actividad.volumen,
                actividad.personas_asignadas,
                actividad.total_horas,
                str(actividad.id_familia_cargo),
                actividad.fecha_inicio_actividad,
                actividad.fecha_termino_actividad,
                str(actividad.id_estado_actividad),
                actividad.fecha_real_inicio,
                actividad.fecha_real_termino,
                actividad.fecha_reprogramacion_inicio,
                actividad.fecha_reprogramacion_termino,
                actividad.justificacion,
                actividad.fecha_registro,

            ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value


            if col_num == 11:
                cell.number_format = 'dd/mm/yyyy'
            if col_num == 12:
                cell.number_format = 'dd/mm/yyyy'
            if col_num == 14:
                cell.number_format = 'dd/mm/yyyy'
            if col_num == 15:
                cell.number_format = 'dd/mm/yyyy'
            if col_num == 16:
                cell.number_format = 'dd/mm/yyyy'
            if col_num == 17:
                cell.number_format = 'dd/mm/yyyy:HH:MM:SS'

    workbook.save(response)


    return response

