from django.contrib.auth.models import User
from django.core import mail
from django.core.mail import EmailMultiAlternatives
from django.db.models import Q, OuterRef, Count, Subquery
from django.shortcuts import render
from django.utils.module_loading import import_string
from django.views.generic import ListView, CreateView, UpdateView, DeleteView

from apps.actividades.models import Ges_Actividad
from apps.controlador.models import Ges_Controlador
from apps.estructura.models import Ges_Niveles
from apps.objetivos.models import Ges_Objetivo_Tactico, Ges_Objetivo_TacticoTN, Ges_Objetivo_Operativo
from apps.periodos.models import Glo_Periodos
from django.http import HttpResponseRedirect, JsonResponse
from django.contrib import messages

from apps.registration.models import logEventos
from apps.revision_planificacion.forms import   PlanUpdateForm
from apps.valida_plan.models import Ges_Observaciones
from apps.valida_plan2.models import Ges_Observaciones_sr
from django.conf import settings
from datetime import datetime
from django.contrib import messages
from django.core.mail import EmailMessage,send_mass_mail
from openpyxl import Workbook
from django.http import HttpResponse
from decimal import *
from django.db.models import Sum
from apps.gestion_horas.models import Ges_Registro_Horas

def EnviarCorreoRechazo_jefaturas(emails_jefaturas, area_plan):
    # try:
    #     periodo_actual = Glo_Periodos.objects.get(id_estado=1)
    # except Glo_Periodos.DoesNotExist:
    #     return None

    ahora = datetime.now()
    fecha = ahora.strftime("%d" + "/" + "%m" + "/" + "%Y" + " a las " + "%H:%M")

    #controladorPlan = Ges_Jefatura.objects.values_list('id_user__email' , flat=True).filter(Q(id_periodo=periodo_actual) & Q(id=id_jefatura))
    idcorreoJefatura=emails_jefaturas
    area_plan=str(area_plan)

    subject = 'Rechazo Plan de Gestión '
    messageHtml = 'Estimada(o) Usuaria(o),<br> Se informa que con fecha <b>'+ fecha +'</b> se ha rechazado con observaciones el plan de gestión para correpondiente a <b>' + area_plan + '</b> . <br> Atte. <br>Subdpto. de Planificación Institucional.<br><p style="font-size:12px;color:red;">correo generado automaticamente favor no responder.'

    email = EmailMessage(subject, messageHtml ,to=idcorreoJefatura)

    email.content_subtype='html'
    email.send()

def EnviarCorreoAcepta_jefaturas(emails_jefaturas, area_plan):
    # try:
    #     periodo_actual = Glo_Periodos.objects.get(id_estado=1)
    # except Glo_Periodos.DoesNotExist:
    #     return None

    ahora = datetime.now()
    fecha = ahora.strftime("%d" + "/" + "%m" + "/" + "%Y" + " a las " + "%H:%M")

    #controladorPlan = Ges_Jefatura.objects.values_list('id_user__email' , flat=True).filter(Q(id_periodo=periodo_actual) & Q(id=id_jefatura))
    idcorreoJefatura=emails_jefaturas
    area_plan=str(area_plan)

    subject = 'Aprobación Plan de Gestión '
    messageHtml = 'Estimada(o) Usuaria(o),<br> Se informa que con fecha <b>'+ fecha +'</b> se ha aprobado el plan de gestión para correpondiente a <b>' + area_plan + '</b> . <br> Atte. <br>Subdpto. de Planificación Institucional.<br><p style="font-size:12px;color:red;">correo generado automaticamente favor no responder.'

    email = EmailMessage(subject, messageHtml ,to=idcorreoJefatura)

    email.content_subtype='html'
    email.send()

class UnidadesListar(ListView):
    model = Ges_Niveles
    template_name = 'revision_planificacion/revision_planificacion_list.html'

    def get_context_data(self, **kwargs):
        context = super(UnidadesListar, self).get_context_data(**kwargs)
        id_usuario_actual = self.request.user.id  # obtiene id usuario actual

        try:
            periodo_actual = Glo_Periodos.objects.get(id_estado=1)
        except Glo_Periodos.DoesNotExist:
            return None

        try:

            count_no_vistos = Ges_Observaciones.objects.values('id_controlador').filter(
                id_controlador=OuterRef('pk')).annotate(
                count_id_actividad=Count('id', filter=Q(observado=1) & Q(id_periodo=periodo_actual.id) & (
                    ~Q(user_observa=id_usuario_actual))))

            count_no_vistos_generales = Ges_Observaciones_sr.objects.values('id_controlador').filter(
                id_controlador=OuterRef('pk')).annotate(
                count_id_actividad=Count('id', filter=Q(observado=1) & Q(id_periodo=periodo_actual.id) & (
                    ~Q(user_observa=id_usuario_actual))))

            count_observaciones = Ges_Observaciones.objects.values('id_controlador').filter(
                id_controlador=OuterRef('pk')).annotate(
                count_id_actividad=Count('id'))

            id_controlador = Ges_Controlador.objects.filter(
                Q(analista_asignado=id_usuario_actual) & Q(id_periodo=periodo_actual.id) & Q(
                    estado_flujo__in= [6,7,11])).annotate(
                count_no_vistos=Subquery(count_no_vistos.values('count_id_actividad')),
                count_no_vistos_generales=Subquery(count_no_vistos_generales.values('count_id_actividad')),
                count_observaciones=Subquery(count_observaciones.values('count_id_actividad')[0:1])).order_by(
                '-count_no_vistos', '-count_observaciones')

        except Ges_Controlador.DoesNotExist:
            return None

        context['object_list'] = id_controlador
        return context


class UnidadesListarNoFinalizadas(ListView):
    model = Ges_Niveles
    template_name = 'revision_planificacion/revision_planificacion_list_no_finalizadas.html'

    def get_context_data(self, **kwargs):
        context = super(UnidadesListarNoFinalizadas, self).get_context_data(**kwargs)
        id_usuario_actual = self.request.user.id  # obtiene id usuario actual

        try:
            periodo_actual = Glo_Periodos.objects.get(id_estado=1)
        except Glo_Periodos.DoesNotExist:
            return None

        try:

            id_controlador = Ges_Controlador.objects.filter(
                 Q(id_periodo=periodo_actual.id) & ~Q(estado_flujo__in= [6,7,2,10]))

        except Ges_Controlador.DoesNotExist:
            return None

        context['object_list'] = id_controlador
        return context



class ObjetivosListar(ListView):
    model = Ges_Actividad
    template_name = 'revision_planificacion/revision_planificacion_detalle.html'

    def get_context_data(self, **kwargs):
        context = super(ObjetivosListar, self).get_context_data(**kwargs)
        id_usuario_actual = self.request.user.id  # obtiene id usuario actual
        self.request.session['id_nivel_controlador']=self.kwargs['pk'] #guarda id  controlador

        try:
            periodo_actual = Glo_Periodos.objects.get(id_estado=1)
        except Glo_Periodos.DoesNotExist:
            return None

        try:
            controlador = Ges_Controlador.objects.get(Q(id=self.kwargs['pk']) & Q(id_periodo=periodo_actual.id))
        except Ges_Controlador.DoesNotExist:
            return None

        id_orden = controlador.nivel_inicial
        self.request.session['id_controlador_real'] = controlador.id

        try:

            total_dias = list(Ges_Registro_Horas.objects.filter(
                Q(id_nivel=controlador.id_jefatura.id_nivel_id) & Q(id_periodo=periodo_actual.id)).aggregate(
                Sum('dias_habiles')).values())[0]

            total_utilizado = list(Ges_Actividad.objects.filter(
                Q(id_controlador=controlador.id) & Q(id_periodo=periodo_actual.id)).aggregate(
                Sum('total_horas')).values())[0]

            if total_dias != None:
                total_horas = total_dias * 8
            else:
                total_horas = 0

            if total_utilizado != None:
                consumidas = total_utilizado
                total_utilizado = total_horas - total_utilizado
                try:
                    avance_porcentaje = (consumidas / total_horas) * 100
                    avance_porcentaje = Decimal(avance_porcentaje)
                    avance_porcentaje = avance_porcentaje.quantize(Decimal('0.1'), rounding=ROUND_UP)
                except ZeroDivisionError:
                    avance_porcentaje = 0

            else:
                total_utilizado = total_horas
                consumidas = 0
                avance_porcentaje = 0
            total_utilizado = int(total_utilizado)
            consumidas = int(consumidas)

        except Ges_Registro_Horas.DoesNotExist:
            return None


       # vaar= controlador.id_jefatura.id_nivel_id

        id_nivel = Ges_Niveles.objects.get(
            Q(id=controlador.id_jefatura.id_nivel_id) & Q(id_periodo=periodo_actual.id))
        try:
            if id_orden == 2:
                answer_subquery = Ges_Actividad.objects.values('id_objetivo_tactico_id').filter(
                    id_objetivo_tactico=OuterRef('pk')).annotate(
                    count_actividad=Count('id', filter=Q(id_periodo=periodo_actual.id)))

                count_no_vistos_obj = Ges_Observaciones_sr.objects.values('id_objetivo_tactico').filter(
                    id_objetivo_tactico=OuterRef('pk')).annotate(
                    count_id_actividad=Count('id', filter=Q(observado=1) & Q(id_controlador=controlador.id) & Q(id_periodo=periodo_actual.id) & (
                        ~Q(user_observa=id_usuario_actual))))

                count_no_vistos = Ges_Observaciones.objects.values('id_objetivo').filter(
                    id_objetivo=OuterRef('pk')).annotate(
                    count_id_actividad=Count('id', filter=Q(observado=1) & Q(id_controlador=controlador.id) & Q(id_periodo=periodo_actual.id) & (
                        ~Q(user_observa=id_usuario_actual))))

                count_observaciones_obj = Ges_Observaciones_sr.objects.values('id_objetivo_tactico').filter(
                    id_objetivo_tactico=OuterRef('pk')).annotate(
                    count_id_actividad=Count('id'))

                count_observaciones = Ges_Observaciones.objects.values('id_objetivo').filter(Q(
                    id_objetivo=OuterRef('pk')) & Q(id_controlador=controlador.id)).annotate(
                    count_id_actividad=Count('id'))

                replies2 = Ges_Objetivo_Tactico.objects.filter(
                    Q(id_segundo_nivel_id=id_nivel.id_segundo_nivel_id) & Q(id_periodo=periodo_actual.id)).annotate(
                    count_no_vistos_obj=Subquery(count_no_vistos_obj.values('count_id_actividad')),
                    count_no_vistos=Subquery(count_no_vistos.values('count_id_actividad')),
                    count_observaciones_obj=Subquery(count_observaciones_obj.values('count_id_actividad')),
                    count_observaciones=Subquery(count_observaciones.values('count_id_actividad')),
                    count_actividad=Subquery(answer_subquery.values('count_actividad')[0:1])).order_by(
                    '-count_no_vistos', '-count_observaciones')

            if id_orden == 3:
                answer_subquery = Ges_Actividad.objects.values('id_objetivo_tacticotn_id').filter(
                    id_objetivo_tacticotn=OuterRef('pk')).annotate(
                    count_actividad=Count('id', filter=Q(id_periodo=periodo_actual.id)))

                count_no_vistos_obj = Ges_Observaciones_sr.objects.values('id_objetivo_tacticotn').filter(
                    id_objetivo_tacticotn=OuterRef('pk')).annotate(
                    count_id_actividad=Count('id', filter=Q(observado=1) & Q(id_controlador=controlador.id) & Q(id_periodo=periodo_actual.id) & (
                        ~Q(user_observa=id_usuario_actual))))

                count_no_vistos = Ges_Observaciones.objects.values('id_objetivo').filter(
                    id_objetivo=OuterRef('pk')).annotate(
                    count_id_actividad=Count('id', filter=Q(observado=1) & Q(id_controlador=controlador.id) & Q(id_periodo=periodo_actual.id) & (
                        ~Q(user_observa=id_usuario_actual))))

                count_observaciones_obj = Ges_Observaciones_sr.objects.values('id_objetivo_tacticotn').filter(
                    id_objetivo_tacticotn=OuterRef('pk')).annotate(
                    count_id_actividad=Count('id'))

                count_observaciones = Ges_Observaciones.objects.values('id_objetivo').filter(Q(
                    id_objetivo=OuterRef('pk')) & Q(id_controlador=controlador.id)).annotate(
                    count_id_actividad=Count('id'))

                replies2 = Ges_Objetivo_TacticoTN.objects.filter(
                    Q(id_tercer_nivel_id=id_nivel.id_tercer_nivel_id) & Q(id_periodo=periodo_actual.id)).annotate(
                    count_no_vistos_obj=Subquery(count_no_vistos_obj.values('count_id_actividad')),
                    count_no_vistos=Subquery(count_no_vistos.values('count_id_actividad')),
                    count_observaciones_obj=Subquery(count_observaciones_obj.values('count_id_actividad')),
                    count_observaciones=Subquery(count_observaciones.values('count_id_actividad')),
                    count_actividad=Subquery(answer_subquery.values('count_actividad')[0:1])).order_by(
                    '-count_no_vistos', '-count_observaciones')

            if id_orden == 4:
                answer_subquery = Ges_Actividad.objects.values('id_objetivo_operativo_id').filter(
                    id_objetivo_operativo=OuterRef('pk')).annotate(
                    count_actividad=Count('id', filter=Q(id_periodo=periodo_actual.id)))

                count_no_vistos_obj = Ges_Observaciones_sr.objects.values('id_objetivo_operativo').filter(
                    id_objetivo_operativo=OuterRef('pk')).annotate(
                    count_id_actividad=Count('id', filter=Q(observado=1) & Q(id_controlador=controlador.id) & Q(id_periodo=periodo_actual.id) & (
                        ~Q(user_observa=id_usuario_actual))))

                count_no_vistos = Ges_Observaciones.objects.values('id_objetivo').filter(
                    id_objetivo=OuterRef('pk')).annotate(
                    count_id_actividad=Count('id', filter=Q(observado=1) & Q(id_controlador=controlador.id) & Q(id_periodo=periodo_actual.id) & (
                        ~Q(user_observa=id_usuario_actual))))


                count_observaciones_obj = Ges_Observaciones_sr.objects.values('id_objetivo_operativo').filter(Q(
                    id_objetivo_operativo=OuterRef('pk'))).annotate(
                    count_id_actividad=Count('id'))

                count_observaciones = Ges_Observaciones.objects.values('id_objetivo').filter(Q(
                    id_objetivo=OuterRef('pk')) & Q(id_controlador=controlador.id) ).annotate(
                    count_id_actividad=Count('id'))

                replies2 = Ges_Objetivo_Operativo.objects.filter(
                    Q(id_cuarto_nivel_id=id_nivel.id_cuarto_nivel_id) & Q(id_periodo=periodo_actual.id)).annotate(
                    count_no_vistos_obj=Subquery(count_no_vistos_obj.values('count_id_actividad')),
                    count_no_vistos=Subquery(count_no_vistos.values('count_id_actividad')),
                    count_observaciones_obj=Subquery(count_observaciones_obj.values('count_id_actividad')),
                    count_observaciones=Subquery(count_observaciones.values('count_id_actividad')),
                    count_actividad=Subquery(answer_subquery.values('count_actividad')[0:1])).order_by(
                    '-count_no_vistos', '-count_observaciones')

        except:
            return None
        context['orden'] = {'orden_nivel': id_orden}
        context['niveles'] = replies2
        context['objetivo'] = id_nivel
        context['total_disponible'] = {'total_dias': total_dias, 'total_horas': total_horas,
                                       'total_utilizado': total_utilizado,
                                       'consumidas': consumidas,
                                       'avance_porcentaje': avance_porcentaje

                                       }
        self.request.session['id_orden'] = id_orden
        return context

class ObjetivosListarNoFinalizadas(ListView):
    model = Ges_Actividad
    template_name = 'revision_planificacion/revision_planificacion_detalle_no_finalizadas.html'

    def get_context_data(self, **kwargs):
        context = super(ObjetivosListarNoFinalizadas, self).get_context_data(**kwargs)
        id_usuario_actual = self.request.user.id  # obtiene id usuario actual
        self.request.session['id_nivel_controlador']=self.kwargs['pk'] #guarda id  controlador

        try:
            periodo_actual = Glo_Periodos.objects.get(id_estado=1)
        except Glo_Periodos.DoesNotExist:
            return None

        try:
            controlador = Ges_Controlador.objects.get(Q(id=self.kwargs['pk']) & Q(id_periodo=periodo_actual.id))
        except Ges_Controlador.DoesNotExist:
            return None

        id_orden = controlador.nivel_inicial
        self.request.session['id_controlador_real'] = controlador.id
       # vaar= controlador.id_jefatura.id_nivel_id

        id_nivel = Ges_Niveles.objects.get(
            Q(id=controlador.id_jefatura.id_nivel_id) & Q(id_periodo=periodo_actual.id))
        try:
            if id_orden == 2:
                answer_subquery = Ges_Actividad.objects.values('id_objetivo_tactico_id').filter(
                    id_objetivo_tactico=OuterRef('pk')).annotate(
                    count_actividad=Count('id', filter=Q(id_periodo=periodo_actual.id)))

                count_no_vistos = Ges_Observaciones_sr.objects.values('id_objetivo_tactico').filter(
                    id_objetivo_tactico=OuterRef('pk')).annotate(
                    count_id_actividad=Count('id', filter=Q(observado=1) & Q(id_periodo=periodo_actual.id) & (
                        ~Q(user_observa=id_usuario_actual))))

                count_observaciones = Ges_Observaciones_sr.objects.values('id_objetivo_tactico').filter(
                    id_objetivo_tactico=OuterRef('pk')).annotate(
                    count_id_actividad=Count('id'))

                replies2 = Ges_Objetivo_Tactico.objects.filter(
                    Q(id_segundo_nivel_id=id_nivel.id_segundo_nivel_id) & Q(id_periodo=periodo_actual.id)).annotate(
                    count_no_vistos=Subquery(count_no_vistos.values('count_id_actividad')),
                    count_observaciones=Subquery(count_observaciones.values('count_id_actividad')),
                    count_actividad=Subquery(answer_subquery.values('count_actividad')[0:1])).order_by(
                    '-count_no_vistos', '-count_observaciones')

            if id_orden == 3:
                answer_subquery = Ges_Actividad.objects.values('id_objetivo_tacticotn_id').filter(
                    id_objetivo_tacticotn=OuterRef('pk')).annotate(
                    count_actividad=Count('id', filter=Q(id_periodo=periodo_actual.id)))

                count_no_vistos = Ges_Observaciones_sr.objects.values('id_objetivo_tacticotn').filter(
                    id_objetivo_tacticotn=OuterRef('pk')).annotate(
                    count_id_actividad=Count('id', filter=Q(observado=1) & Q(id_periodo=periodo_actual.id) & (
                        ~Q(user_observa=id_usuario_actual))))

                count_observaciones = Ges_Observaciones_sr.objects.values('id_objetivo_tacticotn').filter(
                    id_objetivo_tacticotn=OuterRef('pk')).annotate(
                    count_id_actividad=Count('id'))

                replies2 = Ges_Objetivo_TacticoTN.objects.filter(
                    Q(id_tercer_nivel_id=id_nivel.id_tercer_nivel_id) & Q(id_periodo=periodo_actual.id)).annotate(
                    count_no_vistos=Subquery(count_no_vistos.values('count_id_actividad')),
                    count_observaciones=Subquery(count_observaciones.values('count_id_actividad')),
                    count_actividad=Subquery(answer_subquery.values('count_actividad')[0:1])).order_by(
                    '-count_no_vistos', '-count_observaciones')

            if id_orden == 4:
                answer_subquery = Ges_Actividad.objects.values('id_objetivo_operativo_id').filter(
                    id_objetivo_operativo=OuterRef('pk')).annotate(
                    count_actividad=Count('id', filter=Q(id_periodo=periodo_actual.id)))

                count_no_vistos = Ges_Observaciones_sr.objects.values('id_objetivo_operativo').filter(
                    id_objetivo_operativo=OuterRef('pk')).annotate(
                    count_id_actividad=Count('id', filter=Q(observado=1) & Q(id_periodo=periodo_actual.id) & (
                        ~Q(user_observa=id_usuario_actual))))

                count_observaciones = Ges_Observaciones_sr.objects.values('id_objetivo_operativo').filter(
                    id_objetivo_operativo=OuterRef('pk')).annotate(
                    count_id_actividad=Count('id'))

                replies2 = Ges_Objetivo_Operativo.objects.filter(
                    Q(id_cuarto_nivel_id=id_nivel.id_cuarto_nivel_id) & Q(id_periodo=periodo_actual.id)).annotate(
                    count_no_vistos=Subquery(count_no_vistos.values('count_id_actividad')),
                    count_observaciones=Subquery(count_observaciones.values('count_id_actividad')),
                    count_actividad=Subquery(answer_subquery.values('count_actividad')[0:1])).order_by(
                    '-count_no_vistos', '-count_observaciones')

        except:
            return None
        context['orden'] = {'orden_nivel': id_orden}
        context['niveles'] = replies2
        context['objetivo'] = id_nivel
        self.request.session['id_orden'] = id_orden
        return context


class ActividadesListar(ListView):
    model = Ges_Actividad
    template_name = 'revision_planificacion/revision_planificacion_actividades.html'

    def get_context_data(self,  **kwargs):
        context = super(ActividadesListar, self).get_context_data(**kwargs)
        id_usuario_actual = self.request.user.id  # obtiene id usuario actual
        try:
            periodo_actual = Glo_Periodos.objects.get(id_estado=1)
        except Glo_Periodos.DoesNotExist:
            return None

        self.request.session['id_objetivo']=self.kwargs['pk']

        nombre = ""
        try:
            if self.request.session['id_orden']==2:
               # lista_actividades = Ges_Actividad.objects.filter(Q(id_objetivo_tactico=self.kwargs['pk']) & Q(id_periodo=periodo_actual.id))

                count_no_vistos = Ges_Observaciones.objects.values('id_actividad_id').filter(
                   id_actividad=OuterRef('pk')).annotate(
                   count_id_actividad=Count('id', filter=Q(observado=1) & Q(id_periodo=periodo_actual.id) & (
                       ~Q(user_observa=id_usuario_actual))))

                count_observaciones = Ges_Observaciones.objects.values('id_actividad').filter(id_actividad=OuterRef('pk')).annotate(
                count_id_actividad=Count('id'))

                lista_actividades = Ges_Actividad.objects.filter(
                  Q(id_objetivo_tactico=self.kwargs['pk']) & Q(id_periodo=periodo_actual.id)).annotate(count_no_vistos=Subquery(count_no_vistos.values('count_id_actividad')),
                 count_observaciones=Subquery(count_observaciones.values('count_id_actividad'))).order_by('-count_observaciones')

                nombre=  Ges_Objetivo_Tactico.objects.get(id=self.kwargs['pk'])

            if self.request.session['id_orden']==3:
               # lista_actividades = Ges_Actividad.objects.filter(Q(id_objetivo_tacticotn=self.kwargs['pk']) & Q(id_periodo=periodo_actual.id))

                count_no_vistos = Ges_Observaciones.objects.values('id_actividad_id').filter(
                   id_actividad=OuterRef('pk')).annotate(
                   count_id_actividad=Count('id', filter=Q(observado=1) & Q(id_periodo=periodo_actual.id) & (
                       ~Q(user_observa=id_usuario_actual))))

                count_observaciones = Ges_Observaciones.objects.values('id_actividad').filter(id_actividad=OuterRef('pk')).annotate(
                count_id_actividad=Count('id'))

                lista_actividades = Ges_Actividad.objects.filter(
                  Q(id_objetivo_tacticotn=self.kwargs['pk']) & Q(id_periodo=periodo_actual.id)).annotate(count_no_vistos=Subquery(count_no_vistos.values('count_id_actividad')),
                 count_observaciones=Subquery(count_observaciones.values('count_id_actividad'))).order_by('-count_observaciones')


                nombre = Ges_Objetivo_TacticoTN.objects.get(id=self.kwargs['pk'])

            if self.request.session['id_orden']==4:


                count_observaciones = Ges_Observaciones.objects.values('id_actividad').filter(id_actividad=OuterRef('pk')).annotate(
                count_id_actividad=Count('id'))

                count_no_vistos = Ges_Observaciones.objects.values('id_actividad_id').filter(
                    id_actividad=OuterRef('pk')).annotate(
                    count_id_actividad=Count('id', filter=Q(observado=1) & Q(id_periodo=periodo_actual.id) & (
                        ~Q(user_observa=id_usuario_actual))))

                lista_actividades = Ges_Actividad.objects.filter(
                  Q(id_objetivo_operativo=self.kwargs['pk']) & Q(id_periodo=periodo_actual.id)).annotate(count_no_vistos=Subquery(count_no_vistos.values('count_id_actividad')),
                 count_observaciones=Subquery(count_observaciones.values('count_id_actividad'))).order_by('-count_observaciones')
                nombre = Ges_Objetivo_Operativo.objects.get(id=self.kwargs['pk'])
        except:
            return None



        context['object_list'] = lista_actividades

        context['nombre_objetivo'] = {'nombre': nombre}
        context['id_orden'] = {'id_orden': self.request.session['id_orden']}


        context['id_nivel_controlador']={'id': self.request.session['id_nivel_controlador']}


        return context


class ActividadesListarNoFinalizadas(ListView):
    model = Ges_Actividad
    template_name = 'revision_planificacion/revision_planificacion_actividades_no_finalizadas.html'

    def get_context_data(self,  **kwargs):
        context = super(ActividadesListarNoFinalizadas, self).get_context_data(**kwargs)
        id_usuario_actual = self.request.user.id  # obtiene id usuario actual
        try:
            periodo_actual = Glo_Periodos.objects.get(id_estado=1)
        except Glo_Periodos.DoesNotExist:
            return None

        self.request.session['id_objetivo']=self.kwargs['pk']

        nombre = ""
        try:
            if self.request.session['id_orden']==2:
               # lista_actividades = Ges_Actividad.objects.filter(Q(id_objetivo_tactico=self.kwargs['pk']) & Q(id_periodo=periodo_actual.id))

                count_no_vistos = Ges_Observaciones.objects.values('id_actividad_id').filter(
                   id_actividad=OuterRef('pk')).annotate(
                   count_id_actividad=Count('id', filter=Q(observado=1) & Q(id_periodo=periodo_actual.id) & (
                       ~Q(user_observa=id_usuario_actual))))

                count_observaciones = Ges_Observaciones.objects.values('id_actividad').filter(id_actividad=OuterRef('pk')).annotate(
                count_id_actividad=Count('id'))

                lista_actividades = Ges_Actividad.objects.filter(
                  Q(id_objetivo_tactico=self.kwargs['pk']) & Q(id_periodo=periodo_actual.id)).annotate(count_no_vistos=Subquery(count_no_vistos.values('count_id_actividad')),
                 count_observaciones=Subquery(count_observaciones.values('count_id_actividad'))).order_by('-count_observaciones')

                nombre=  Ges_Objetivo_Tactico.objects.get(id=self.kwargs['pk'])

            if self.request.session['id_orden']==3:
               # lista_actividades = Ges_Actividad.objects.filter(Q(id_objetivo_tacticotn=self.kwargs['pk']) & Q(id_periodo=periodo_actual.id))

                count_no_vistos = Ges_Observaciones.objects.values('id_actividad_id').filter(
                   id_actividad=OuterRef('pk')).annotate(
                   count_id_actividad=Count('id', filter=Q(observado=1) & Q(id_periodo=periodo_actual.id) & (
                       ~Q(user_observa=id_usuario_actual))))

                count_observaciones = Ges_Observaciones.objects.values('id_actividad').filter(id_actividad=OuterRef('pk')).annotate(
                count_id_actividad=Count('id'))

                lista_actividades = Ges_Actividad.objects.filter(
                  Q(id_objetivo_tacticotn=self.kwargs['pk']) & Q(id_periodo=periodo_actual.id)).annotate(count_no_vistos=Subquery(count_no_vistos.values('count_id_actividad')),
                 count_observaciones=Subquery(count_observaciones.values('count_id_actividad'))).order_by('-count_observaciones')


                nombre = Ges_Objetivo_TacticoTN.objects.get(id=self.kwargs['pk'])

            if self.request.session['id_orden']==4:


                count_observaciones = Ges_Observaciones.objects.values('id_actividad').filter(id_actividad=OuterRef('pk')).annotate(
                count_id_actividad=Count('id'))

                count_no_vistos = Ges_Observaciones.objects.values('id_actividad_id').filter(
                    id_actividad=OuterRef('pk')).annotate(
                    count_id_actividad=Count('id', filter=Q(observado=1) & Q(id_periodo=periodo_actual.id) & (
                        ~Q(user_observa=id_usuario_actual))))

                lista_actividades = Ges_Actividad.objects.filter(
                  Q(id_objetivo_operativo=self.kwargs['pk']) & Q(id_periodo=periodo_actual.id)).annotate(count_no_vistos=Subquery(count_no_vistos.values('count_id_actividad')),
                 count_observaciones=Subquery(count_observaciones.values('count_id_actividad'))).order_by('-count_observaciones')
                nombre = Ges_Objetivo_Operativo.objects.get(id=self.kwargs['pk'])
        except:
            return None



        context['object_list'] = lista_actividades

        context['nombre_objetivo'] = {'nombre': nombre}
        context['id_orden'] = {'id_orden': self.request.session['id_orden']}


        context['id_nivel_controlador']={'id': self.request.session['id_nivel_controlador']}


        return context



class RechazaPlan(UpdateView):
    model = Ges_Controlador
    form_class = PlanUpdateForm
    template_name = 'revision_planificacion/revision_planificacion_rechaza_form.html'

    def post(self, request, *args, **kwargs):
        self.object = self.get_object
        id_controlador = kwargs['pk']
        id_usuario_actual = self.request.user.id  # obtiene id usuario actual


        try:
            periodo_actual = Glo_Periodos.objects.get(id_estado=1)
        except Glo_Periodos.DoesNotExist:
            return None

        controladorPlan = self.model.objects.get(id=id_controlador)
        email_jefatura = controladorPlan.id_jefatura.id_user.email    #correo de rechazo


        try:
            email_jefatura_primera = controladorPlan.jefatura_primerarevision.id_user.email  # correo de rechazo 2da revision
        except:
            email_jefatura_primera = None
            pass


        try:
            email_jefatura_segunda = controladorPlan.jefatura_segundarevision.id_user.email  # correo de rechazo 2da revision
        except:
            email_jefatura_segunda = None
            pass

        emails_jefaturas=[email_jefatura,email_jefatura_primera,email_jefatura_segunda ]
        area_plan = controladorPlan.id_jefatura.id_nivel.descripcion_nivel

        lista_observaciones = int(Ges_Observaciones_sr.objects.filter(Q(id_controlador=id_controlador) & Q(id_periodo=periodo_actual.id)
                                                         & Q(observado=1) & Q(user_observa=id_usuario_actual)).count() ) + int(Ges_Observaciones.objects.filter(
            Q(id_controlador=id_controlador) & Q(id_periodo=periodo_actual.id)
            & Q(observado=1) & Q(user_observa=id_usuario_actual)).count())

        estado = 11
        controladorPlan.estado_flujo_id = int(estado)

        count_obs_no_vistas = Ges_Observaciones.objects.filter(
            Q(id_controlador=id_controlador) & Q(observado=1) & ~Q(user_observa_id=id_usuario_actual) & Q(
                id_periodo=periodo_actual.id)).count()
        count_obs_no_vistas_generales = Ges_Observaciones_sr.objects.filter(
            Q(id_controlador=id_controlador) & Q(observado=1) & ~Q(user_observa_id=id_usuario_actual) & Q(
                id_periodo=periodo_actual.id)).count()

        count_obs_nuevas = Ges_Observaciones.objects.filter(
            Q(id_controlador=id_controlador) & Q(observado=1) & Q(user_observa_id=id_usuario_actual) & Q(
                id_periodo=periodo_actual.id)).count()
        count_obs_nuevas_generales = Ges_Observaciones_sr.objects.filter(
            Q(id_controlador=id_controlador) & Q(observado=1) & Q(user_observa_id=id_usuario_actual) & Q(
                id_periodo=periodo_actual.id)).count()

        if count_obs_no_vistas == 0 and count_obs_no_vistas_generales == 0:
            if count_obs_nuevas > 0 or count_obs_nuevas_generales > 0:
                try:
                    controladorPlan.save()

                    tipo_evento = "Rechaza plan Analista"
                    metodo = "Revisión - Planificación"
                    usuario_evento = self.request.user
                    jefatura_dirigida = None
                    logEventosCreate(tipo_evento, metodo, usuario_evento, jefatura_dirigida)

                    try:

                        EnviarCorreoRechazo_jefaturas(emails_jefaturas, area_plan)

                        request.session['message_class'] = "alert alert-success"  # Tipo mensaje
                        messages.success(request,
                                         "El Plan fue rechazado correctamente y el correo de notificación fue enviado a las jefaturas!.")  # mensaje
                        return HttpResponseRedirect(
                            '/revision_planificacion/listarUnidadesAnalista')  # Redirije a la pantalla principal

                    except:

                        request.session['message_class'] = "alert alert-warning"  # Tipo mensaje
                        messages.success(request,
                                         "El Plan fue rechazado correctamente!, pero el servicio de correo tuvo un inconveniente favor comuníquese con las jefaturas para informar del rechazo.")  # mensaje
                        return HttpResponseRedirect(
                            '/revision_planificacion/listarUnidadesAnalista')  # Redirije a la pantalla principal


                except:
                    request.session['message_class'] = "alert alert-danger"
                    messages.error(self.request,
                                   "Error interno: El plan no fue rechazado, intente nuevamente o comuníquese con el administrador.")
                    return HttpResponseRedirect('/revision_planificacion/listarUnidadesAnalista')
            else:
                request.session['message_class'] = "alert alert-danger"
                messages.error(self.request, "Para rechazar el plan, primero debe ingresar al menos una observacion.")
                return HttpResponseRedirect('/revision_planificacion/listarUnidadesAnalista')

        else:
            request.session['message_class'] = "alert alert-danger"
            messages.error(self.request, "Para rechazar el plan, primero debe leer todas las observaciones.")
            return HttpResponseRedirect('/revision_planificacion/listarUnidadesAnalista')







class AceptaPlan(UpdateView):
    model = Ges_Controlador
    form_class = PlanUpdateForm
    template_name = 'revision_planificacion/revision_planificacion_acepta_form.html'

    def post(self, request, *args, **kwargs):
        self.object = self.get_object
        id_controlador = kwargs['pk']
        id_usuario_actual = self.request.user.id

        try:
            periodo_actual = Glo_Periodos.objects.get(id_estado=1)
        except Glo_Periodos.DoesNotExist:
            return None

        controladorPlan = self.model.objects.get(Q(id=id_controlador) & Q(id_periodo=periodo_actual.id))
        #Se debe cambiar por email Analistas
        email_jefatura = controladorPlan.id_jefatura.id_user.email  # correo de rechazo


        try:
            email_jefatura_primera = controladorPlan.jefatura_primerarevision.id_user.email  # correo de rechazo 1era revision
        except:
            email_jefatura_primera = None
            pass



        try:
            email_jefatura_segunda = controladorPlan.jefatura_segundarevision.id_user.email  # correo de rechazo 2da revision
        except:
            email_jefatura_segunda = None
            pass

        emails_jefaturas=[email_jefatura,email_jefatura_primera,email_jefatura_segunda ]
        area_plan = controladorPlan.id_jefatura.id_nivel.descripcion_nivel


        #instancia_nivel = self.model.objects.get(id=id_controlador)
        #form = self.form_class(request.POST, instance=controladorPlan)
        estado = 7
        controladorPlan.estado_flujo_id = int(estado)
        count_obs_no_vistas = Ges_Observaciones.objects.filter(
            Q(id_controlador=id_controlador) & Q(observado=1) & ~Q(user_observa_id=id_usuario_actual) & Q(id_periodo=periodo_actual.id)).count()
        count_obs_no_vistas_generales = Ges_Observaciones_sr.objects.filter(
            Q(id_controlador=id_controlador) & Q(observado=1) & ~Q(user_observa_id=id_usuario_actual) & Q(id_periodo=periodo_actual.id)).count()

        if count_obs_no_vistas == 0 and count_obs_no_vistas_generales == 0:
            try:
                controladorPlan.save()

                tipo_evento = "Aprueba plan Analista"
                metodo = "Revisión - Planificación"
                usuario_evento = self.request.user
                jefatura_dirigida = None
                logEventosCreate(tipo_evento, metodo, usuario_evento, jefatura_dirigida)

                try:

                    EnviarCorreoAcepta_jefaturas(emails_jefaturas, area_plan)

                    request.session['message_class'] = "alert alert-success"  # Tipo mensaje
                    messages.success(request,
                                     "El Plan fue aceptado correctamente y el correo de notificación fue enviado a las jefaturas!.")  # mensaje
                    return HttpResponseRedirect(
                        '/revision_planificacion/listarUnidadesAnalista')  # Redirije a la pantalla principal

                except:

                    request.session['message_class'] = "alert alert-warning"  # Tipo mensaje
                    messages.success(request,
                                     "El Plan fue aceptado correctamente!, pero el servicio de correo tuvo un inconveniente favor comuníquese con las jefaturas para informar del rechazo.")  # mensaje
                    return HttpResponseRedirect(
                        '/revision_planificacion/listarUnidadesAnalista')  # Redirije a la pantalla principal


            except:
                request.session['message_class'] = "alert alert-danger"
                messages.error(self.request,
                               "Error interno: El plan no ha podido ser aceptado. Comuníquese con el administrador.")
                return HttpResponseRedirect('/revision_planificacion/listarUnidadesAnalista')
        else:

            request.session['message_class'] = "alert alert-danger"
            messages.error(self.request,
                           "Estimado usuario, para aceptar el plan debe leer todas las observaciones relacionadas.")
            return HttpResponseRedirect('/revision_planificacion/listarUnidadesAnalista')


class EnviarPlanAdministrador(UpdateView):
    model = Ges_Controlador
    form_class = PlanUpdateForm
    template_name = 'revision_planificacion/enviar_plan_administrador_form.html'

    def post(self, request, *args, **kwargs):
        self.object = self.get_object
        id_controlador = kwargs['pk']
        id_usuario_actual = self.request.user.id

        try:
            periodo_actual = Glo_Periodos.objects.get(id_estado=1)
        except Glo_Periodos.DoesNotExist:
            return None


        controladorPlan = self.model.objects.get(Q(id=id_controlador) & Q(id_periodo=periodo_actual.id))
        estado = 10
        controladorPlan.estado_flujo_id = int(estado)



        try:

            controladorPlan.save()
            QuitarAnalista(id_controlador)

            request.session['message_class'] = "alert alert-success"
            messages.error(self.request,
                           "El plan fue enviado a la bandeja del administrador para ser asignado a un analista.")
            return HttpResponseRedirect('/revision_planificacion/listarUnidadesNoFinalizadas')

        except:
            request.session['message_class'] = "alert alert-danger"
            messages.error(self.request,
                           "Error interno: El plan no fue enviado al administrador. Comuníquese con el administrador.")
            return HttpResponseRedirect('/revision_planificacion/listarUnidadesNoFinalizadas')


def export_users_xls(request, *args, **kwargs):
    try:
        periodo_actual = Glo_Periodos.objects.get(id_estado=1)
    except Glo_Periodos.DoesNotExist:
        return None

    id_controlador = kwargs['pk']


    nivel=Ges_Controlador.objects.get(Q(id=id_controlador) & Q(id_periodo=periodo_actual))

    nivel= nivel.nivel_inicial


    actividades = Ges_Actividad.objects.filter(Q(id_controlador=id_controlador) &
                                                            Q(id_periodo=periodo_actual))

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-Plan_de_Gestion.xlsx'.format(
        date=datetime.now().strftime('%d/%m/%Y'),
    )
    workbook = Workbook()


    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'reporte_seguimiento'

    # Define the titles for columns

    columns = ['Actividad',
               'Objetivo Vinculado',
               'Periodicidad',
               'Producto Estadístico',
               'Hora x Actividad',
               'Volumen',
               'N° Personas Asignadas',
               'Total Horas',
               'Cargo',
               'Fecha Incio Actividad',
               'Fecha Término Actividad',
               'Estado Actividad',
               'Fecha Real Finalización',
               'Reprogramación Fecha Inicio',
               'Reprogramación Fecha Término',
               'Justificación Desviación'

               ]

    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    for actividad in actividades:
        row_num += 1

        row =''

        if nivel==4:

            row = [

                actividad.descripcion_actividad,
                str(actividad.id_objetivo_operativo) ,
                str(actividad.id_periodicidad),
                str(actividad.id_producto_estadistico),
                actividad.horas_actividad,
                actividad.volumen,
                actividad.personas_asignadas,
                actividad.total_horas,
                str(actividad.id_familia_cargo),
                actividad.fecha_inicio_actividad,
                actividad.fecha_termino_actividad,
                str(actividad.id_estado_actividad),
                actividad.fecha_real_termino,
                actividad.fecha_reprogramacion_inicio,
                actividad.fecha_reprogramacion_termino,
                actividad.justificacion,


            ]

        if nivel==3:

            row = [

                actividad.descripcion_actividad,
                str(actividad.id_objetivo_tacticotn) ,
                str(actividad.id_periodicidad),
                str(actividad.id_producto_estadistico),
                actividad.horas_actividad,
                actividad.volumen,
                actividad.personas_asignadas,
                actividad.total_horas,
                str(actividad.id_familia_cargo),
                actividad.fecha_inicio_actividad,
                actividad.fecha_termino_actividad,
                str(actividad.id_estado_actividad),
                actividad.fecha_real_termino,
                actividad.fecha_reprogramacion_inicio,
                actividad.fecha_reprogramacion_termino,
                actividad.justificacion,



            ]



        if nivel==2:

            row = [

                actividad.descripcion_actividad,
                str(actividad.id_objetivo_tactico) ,
                str(actividad.id_periodicidad),
                str(actividad.id_producto_estadistico),
                actividad.horas_actividad,
                actividad.volumen,
                actividad.personas_asignadas,
                actividad.total_horas,
                str(actividad.id_familia_cargo),
                actividad.fecha_inicio_actividad,
                actividad.fecha_termino_actividad,
                str(actividad.id_estado_actividad),
                actividad.fecha_real_termino,
                actividad.fecha_reprogramacion_inicio,
                actividad.fecha_reprogramacion_termino,
                actividad.justificacion,

            ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value


            if col_num == 10:
                cell.number_format = 'dd/mm/yyyy'
            if col_num == 11:
                cell.number_format = 'dd/mm/yyyy'
            if col_num == 13:
                cell.number_format = 'dd/mm/yyyy'
            if col_num == 14:
                cell.number_format = 'dd/mm/yyyy'
            if col_num == 15:
                cell.number_format = 'dd/mm/yyyy'

    workbook.save(response)


    return response




def ActividadDetalle(request, id):

    template_name = "revision_planificacion/modal_revision_planificacion_actividad_detalle.html"
    qs = Ges_Actividad.objects.get(id=id) #Cualquier QS con la que quiera obtener datos para enviar al modal.
    context = {"qs": qs} # aquí le envío lo que quiero al modal para que lo muestre, incluso una lista.



    return render(request, template_name, context)




def GestionObservacionesActividades(request, id):
    template_name = "revision_planificacion/modal_observaciones_actividad.html"
    response_data = {}
    ahora = datetime.now()
    fecha = ahora.strftime("%d" + " de " + "%B" + " de " + "%Y" + " a las " + "%H:%M")

    try:
        periodo_activo = Glo_Periodos.objects.get(id_estado=1)
    except Glo_Periodos.DoesNotExist:
        return None

    id_controller = request.session['id_controlador_real']

    try:
        controlador_id = Ges_Controlador.objects.get(id=id_controller)
    except Ges_Controlador.DoesNotExist:
        controlador_id = None

    id_usuario_redacta_plan = controlador_id.id_jefatura
    id_usuario_actual = request.user.id  # obtiene id usuario actual

    try:
        usuario_id = User.objects.get(id=id_usuario_actual)
    except User.DoesNotExist:
        usuario_id = None


    try:
        actividad = Ges_Actividad.objects.get(id=id)
    except Ges_Actividad.DoesNotExist:
        actividad = 0



    var = id_usuario_redacta_plan.id_nivel.orden_nivel  # Nivel del usuario

    if var==2:
        id_objetivo= actividad.id_objetivo_tactico_id
    if var==3:
        id_objetivo = actividad.id_objetivo_tacticotn_id
    if var==4:
        id_objetivo = actividad.id_objetivo_operativo_id


    novacio= Ges_Observaciones.objects.filter(Q(id_actividad=id) & Q(id_periodo=periodo_activo.id) & (~Q(user_observa=id_usuario_actual)))
    #actualiza el estado a leido
    if novacio:

        Ges_Observaciones.objects.filter(Q(id_actividad=id) & Q(id_periodo=periodo_activo.id) & (~Q(user_observa=id_usuario_actual))).update(observado=0,)


    qs = Ges_Observaciones.objects.filter(Q(id_actividad=id) & Q(id_periodo=periodo_activo.id))

    args = {}

    args['id_actividad'] = id
    args['estado_plan'] = controlador_id.estado_flujo_id
    args['object_list']= qs


    if request.POST.get('action') == 'post':


        observacion = request.POST.get('observacion')

        response_data['observacion'] = observacion
        response_data['id_actividad'] = id
        response_data['fecha'] = fecha


        Ges_Observaciones.objects.create(
            observacion = observacion, id_controlador=id_controller, user_observa= usuario_id, id_actividad_id=id, id_periodo_id=periodo_activo.id,
           id_objetivo=id_objetivo, observado=1,

            )


        return JsonResponse(response_data)

    return render(request, template_name, args)


def GestionObservacionesObjetivosVp2(request, id):
    template_name = "revision_planificacion/modal_observaciones_objetivos_Analista.html"
    response_data = {}
    ahora = datetime.now()
    fecha = ahora.strftime("%d" + " de " + "%B" + " de " + "%Y" + " a las " + "%H:%M")

    id_controller = request.session['id_controlador_real']

    try:
        controlador_id = Ges_Controlador.objects.get(id=id_controller)
    except Ges_Controlador.DoesNotExist:
        controlador_id = None

    try:
        periodo_activo = Glo_Periodos.objects.get(id_estado=1)
    except Glo_Periodos.DoesNotExist:
        return None

    id_usuario_redacta_plan= controlador_id.id_jefatura
    id_usuario_actual = request.user.id  # obtiene id usuario actual

    try:
        usuario_id = User.objects.get(id=id_usuario_actual)
    except User.DoesNotExist:
        usuario_id = None

    var= id_usuario_redacta_plan.id_nivel.orden_nivel # Nivel del usuario

    # actualiza el estado a leido
    if var == 2:
        novacio= Ges_Observaciones_sr.objects.filter(
            Q(id_objetivo_tactico=id) & Q(id_periodo=periodo_activo.id) & (~Q(user_observa = id_usuario_actual)))
        if novacio:
            Ges_Observaciones_sr.objects.filter(Q(id_objetivo_tactico=id) & Q(id_periodo=periodo_activo.id) & (~Q(user_observa = id_usuario_actual))).update(
                observado=0,
            )
    if var == 3:
        novacio = Ges_Observaciones_sr.objects.filter(
            Q(id_objetivo_tacticotn=id) & Q(id_periodo=periodo_activo.id)  & (~Q(user_observa = id_usuario_actual)))
        if novacio:
            Ges_Observaciones_sr.objects.filter(Q(id_objetivo_tacticotn=id) & Q(id_periodo=periodo_activo.id) & (~Q(user_observa = id_usuario_actual))).update(
                observado=0,
            )
    if var == 4:
        novacio = Ges_Observaciones_sr.objects.filter(
            Q(id_objetivo_operativo=id) & Q(id_periodo=periodo_activo.id)  & (~Q(user_observa = id_usuario_actual)))
        if novacio:
            Ges_Observaciones_sr.objects.filter(Q(id_objetivo_operativo=id) & Q(id_periodo=periodo_activo.id) & (~Q(user_observa = id_usuario_actual))).update(
                observado=0,
            )

    if var == 2:
        qs = Ges_Observaciones_sr.objects.filter(Q(id_objetivo_tactico=id) & Q(id_periodo=periodo_activo.id))
    if var == 3:
        qs = Ges_Observaciones_sr.objects.filter(Q(id_objetivo_tacticotn=id) & Q(id_periodo=periodo_activo.id))
    if var == 4:
        qs = Ges_Observaciones_sr.objects.filter(Q(id_objetivo_operativo=id) & Q(id_periodo=periodo_activo.id))


    args = {}

    args['id_objetivo'] = id
    args['estado_plan'] = controlador_id.estado_flujo_id
    args['object_list']= qs


    if request.POST.get('action') == 'post':


        observacion = request.POST.get('observacion')

        response_data['observacion'] = observacion
        response_data['id_actividad'] = id
        response_data['fecha'] = fecha

        if var == 2:
            Ges_Observaciones_sr.objects.create(
                observacion=observacion, id_controlador=id_controller, user_observa=usuario_id,
                id_objetivo_tactico_id=id, id_periodo_id=periodo_activo.id, observado=1,
            )
        if var == 3:
            Ges_Observaciones_sr.objects.create(
                observacion=observacion, id_controlador=id_controller, user_observa=usuario_id,
                id_objetivo_tacticotn_id=id, id_periodo_id=periodo_activo.id, observado=1,
            )
        if var == 4:
            Ges_Observaciones_sr.objects.create(
                observacion=observacion, id_controlador=id_controller, user_observa=usuario_id,
                id_objetivo_operativo_id=id, id_periodo_id=periodo_activo.id, observado=1,
            )



        return JsonResponse(response_data)

    return render(request, template_name, args)

def QuitarAnalista(id_controlador):

    Ges_Controlador.objects.filter(id=id_controlador).update(
        analista_asignado=None,
    )

def logEventosCreate(tipo_evento, metodo ,usuario_evento, jefatura_dirigida):
    logEventos.objects.create(
        tipo_evento=tipo_evento,
        metodo=metodo,
        usuario_evento=usuario_evento,
        jefatura_dirigida=jefatura_dirigida,
    )
    return None








